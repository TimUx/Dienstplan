"""
Shifts router: shift types, schedule, planning, assignments, exports, shift exchanges.
"""

import logging
import threading as _threading
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime, date, timedelta
from typing import Optional
import json
import uuid

from fastapi import APIRouter, Request, Depends
from fastapi.responses import JSONResponse, Response

from .shared import (
    get_db, require_auth, require_role, log_audit,
    get_row_value, _paginate,
    extend_planning_dates_to_complete_weeks, validate_monthly_date_range,
    check_csrf, parse_json_body
)
from .repositories.shift_repository import ShiftRepository

logger = logging.getLogger(__name__)
router = APIRouter()

# ============================================================================
# PROCESS POOL – replaces threading.Thread for solver jobs.
# max_workers=4 limits concurrent solver processes; the GIL is bypassed entirely.
# ============================================================================
_solver_pool = ProcessPoolExecutor(max_workers=4)
_active_futures: dict = {}
_futures_lock = _threading.Lock()
MAX_CONCURRENT_JOBS = 4


@router.get('/api/shifttypes')
def get_shift_types(request: Request):
    """Get all shift types"""
    db = get_db()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    shift_types = []
    for row in ShiftRepository.get_all_shift_types(cursor):
        # Handle MaxConsecutiveDays for backward compatibility
        try:
            max_consecutive_days = row['MaxConsecutiveDays']
        except (KeyError, IndexError):
            max_consecutive_days = 6  # Default value
        
        shift_types.append({
            'id': row['Id'],
            'code': row['Code'],
            'name': row['Name'],
            'startTime': row['StartTime'],
            'endTime': row['EndTime'],
            'colorCode': row['ColorCode'],
            'durationHours': row['DurationHours'],
            'weeklyWorkingHours': row['WeeklyWorkingHours'],
            'isActive': bool(row['IsActive']),
            'worksMonday': bool(row['WorksMonday']),
            'worksTuesday': bool(row['WorksTuesday']),
            'worksWednesday': bool(row['WorksWednesday']),
            'worksThursday': bool(row['WorksThursday']),
            'worksFriday': bool(row['WorksFriday']),
            'worksSaturday': bool(row['WorksSaturday']),
            'worksSunday': bool(row['WorksSunday']),
            'minStaffWeekday': row['MinStaffWeekday'],
            'maxStaffWeekday': row['MaxStaffWeekday'],
            'minStaffWeekend': row['MinStaffWeekend'],
            'maxStaffWeekend': row['MaxStaffWeekend'],
            'maxConsecutiveDays': max_consecutive_days
        })
    
    conn.close()
    return shift_types


@router.post('/api/shifttypes', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])
def create_shift_type(request: Request, data: dict = Depends(parse_json_body)):
    """Create new shift type (Admin only)"""
    try:
        
        # Validate required fields
        required_fields = ['code', 'name', 'startTime', 'endTime', 'durationHours']
        for field in required_fields:
            if not data.get(field):
                return JSONResponse(content={'error': f'{field} ist Pflichtfeld'}, status_code=400)
        
        # Validate staffing requirements
        min_staff_weekday = data.get('minStaffWeekday', 3)
        max_staff_weekday = data.get('maxStaffWeekday', 5)
        min_staff_weekend = data.get('minStaffWeekend', 2)
        max_staff_weekend = data.get('maxStaffWeekend', 3)
        max_consecutive_days = data.get('maxConsecutiveDays', 6)
        
        if min_staff_weekday > max_staff_weekday:
            return JSONResponse(content={'error': 'Minimale Personalstärke an Wochentagen darf nicht größer sein als die maximale Personalstärke'}, status_code=400)
        if min_staff_weekend > max_staff_weekend:
            return JSONResponse(content={'error': 'Minimale Personalstärke am Wochenende darf nicht größer sein als die maximale Personalstärke'}, status_code=400)
        if max_consecutive_days < 1 or max_consecutive_days > 10:
            return JSONResponse(content={'error': 'Maximale aufeinanderfolgende Tage muss zwischen 1 und 10 liegen'}, status_code=400)
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Check if code already exists
        cursor.execute("SELECT Id FROM ShiftTypes WHERE Code = ?", (data.get('code'),))
        if cursor.fetchone():
            conn.close()
            return JSONResponse(content={'error': 'Schichtkürzel bereits vorhanden'}, status_code=400)
        
        # Insert shift type
        cursor.execute("""
            INSERT INTO ShiftTypes (Code, Name, StartTime, EndTime, DurationHours, ColorCode, IsActive,
                                  WorksMonday, WorksTuesday, WorksWednesday, WorksThursday, WorksFriday,
                                  WorksSaturday, WorksSunday, WeeklyWorkingHours, 
                                  MinStaffWeekday, MaxStaffWeekday, MinStaffWeekend, MaxStaffWeekend, 
                                  MaxConsecutiveDays, CreatedBy)
            VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data.get('code'),
            data.get('name'),
            data.get('startTime'),
            data.get('endTime'),
            data.get('durationHours'),
            data.get('colorCode', '#808080'),
            1 if data.get('worksMonday', True) else 0,
            1 if data.get('worksTuesday', True) else 0,
            1 if data.get('worksWednesday', True) else 0,
            1 if data.get('worksThursday', True) else 0,
            1 if data.get('worksFriday', True) else 0,
            1 if data.get('worksSaturday', False) else 0,
            1 if data.get('worksSunday', False) else 0,
            data.get('weeklyWorkingHours', 40.0),
            min_staff_weekday,
            max_staff_weekday,
            min_staff_weekend,
            max_staff_weekend,
            max_consecutive_days,
            request.session.get('user_email', 'system')
        ))
        
        shift_type_id = cursor.lastrowid
        
        # Log audit entry
        changes = json.dumps(data, ensure_ascii=False)
        log_audit(conn, 'ShiftType', shift_type_id, 'Created', changes)
        
        conn.commit()
        conn.close()
        
        return JSONResponse(content={'success': True, 'id': shift_type_id}, status_code=201)
        
    except Exception as e:
        logger.error(f"Create shift type error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Erstellen: {str(e)}'}, status_code=500)


@router.get('/api/shifttypes/{id:int}')

def get_shift_type(request: Request, id):
    """Get single shift type by ID"""
    db = get_db()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    row = ShiftRepository.get_shift_type_by_id(cursor, id)
    
    if not row:
        conn.close()
        return JSONResponse(content={'error': 'Schichttyp nicht gefunden'}, status_code=404)
    
    shift_type = {
        'id': row['Id'],
        'code': row['Code'],
        'name': row['Name'],
        'startTime': row['StartTime'],
        'endTime': row['EndTime'],
        'durationHours': row['DurationHours'],
        'colorCode': row['ColorCode'],
        'isActive': bool(row['IsActive']),
        'worksMonday': bool(row['WorksMonday']),
        'worksTuesday': bool(row['WorksTuesday']),
        'worksWednesday': bool(row['WorksWednesday']),
        'worksThursday': bool(row['WorksThursday']),
        'worksFriday': bool(row['WorksFriday']),
        'worksSaturday': bool(row['WorksSaturday']),
        'worksSunday': bool(row['WorksSunday']),
        'weeklyWorkingHours': row['WeeklyWorkingHours'],
        'minStaffWeekday': row['MinStaffWeekday'],
        'maxStaffWeekday': row['MaxStaffWeekday'],
        'minStaffWeekend': row['MinStaffWeekend'],
        'maxStaffWeekend': row['MaxStaffWeekend']
    }
    
    # Handle MaxConsecutiveDays for backward compatibility
    try:
        shift_type['maxConsecutiveDays'] = row['MaxConsecutiveDays']
    except (KeyError, IndexError):
        shift_type['maxConsecutiveDays'] = 6  # Default value
    
    conn.close()
    return shift_type


@router.put('/api/shifttypes/{id:int}', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def update_shift_type(request: Request, id):
    """Update shift type (Admin only)"""
    try:
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Check if shift type exists
        cursor.execute("SELECT * FROM ShiftTypes WHERE Id = ?", (id,))
        old_row = cursor.fetchone()
        if not old_row:
            conn.close()
            return JSONResponse(content={'error': 'Schichttyp nicht gefunden'}, status_code=404)
        
        # Check if new code conflicts with existing
        if data.get('code') and data.get('code') != old_row['Code']:
            cursor.execute("SELECT Id FROM ShiftTypes WHERE Code = ? AND Id != ?", 
                         (data.get('code'), id))
            if cursor.fetchone():
                conn.close()
                return JSONResponse(content={'error': 'Schichtkürzel bereits vorhanden'}, status_code=400)
        
        # Validate staffing requirements using the helper function
        min_staff_weekday = data.get('minStaffWeekday', get_row_value(old_row, 'MinStaffWeekday', 3))
        max_staff_weekday = data.get('maxStaffWeekday', get_row_value(old_row, 'MaxStaffWeekday', 5))
        min_staff_weekend = data.get('minStaffWeekend', get_row_value(old_row, 'MinStaffWeekend', 2))
        max_staff_weekend = data.get('maxStaffWeekend', get_row_value(old_row, 'MaxStaffWeekend', 3))
        max_consecutive_days = data.get('maxConsecutiveDays', get_row_value(old_row, 'MaxConsecutiveDays', 6))
        
        if min_staff_weekday > max_staff_weekday:
            conn.close()
            return JSONResponse(content={'error': 'Minimale Personalstärke an Wochentagen darf nicht größer sein als die maximale Personalstärke'}, status_code=400)
        if min_staff_weekend > max_staff_weekend:
            conn.close()
            return JSONResponse(content={'error': 'Minimale Personalstärke am Wochenende darf nicht größer sein als die maximale Personalstärke'}, status_code=400)
        if max_consecutive_days < 1 or max_consecutive_days > 10:
            conn.close()
            return JSONResponse(content={'error': 'Maximale aufeinanderfolgende Tage muss zwischen 1 und 10 liegen'}, status_code=400)
        
        # Update shift type
        cursor.execute("""
            UPDATE ShiftTypes 
            SET Code = ?, Name = ?, StartTime = ?, EndTime = ?, 
                DurationHours = ?, ColorCode = ?, IsActive = ?,
                WorksMonday = ?, WorksTuesday = ?, WorksWednesday = ?, WorksThursday = ?, WorksFriday = ?,
                WorksSaturday = ?, WorksSunday = ?, WeeklyWorkingHours = ?,
                MinStaffWeekday = ?, MaxStaffWeekday = ?, MinStaffWeekend = ?, MaxStaffWeekend = ?,
                MaxConsecutiveDays = ?, ModifiedAt = ?, ModifiedBy = ?
            WHERE Id = ?
        """, (
            data.get('code', old_row['Code']),
            data.get('name', old_row['Name']),
            data.get('startTime', old_row['StartTime']),
            data.get('endTime', old_row['EndTime']),
            data.get('durationHours', old_row['DurationHours']),
            data.get('colorCode', old_row['ColorCode']),
            1 if data.get('isActive', True) else 0,
            1 if data.get('worksMonday', get_row_value(old_row, 'WorksMonday', True)) else 0,
            1 if data.get('worksTuesday', get_row_value(old_row, 'WorksTuesday', True)) else 0,
            1 if data.get('worksWednesday', get_row_value(old_row, 'WorksWednesday', True)) else 0,
            1 if data.get('worksThursday', get_row_value(old_row, 'WorksThursday', True)) else 0,
            1 if data.get('worksFriday', get_row_value(old_row, 'WorksFriday', True)) else 0,
            1 if data.get('worksSaturday', get_row_value(old_row, 'WorksSaturday', False)) else 0,
            1 if data.get('worksSunday', get_row_value(old_row, 'WorksSunday', False)) else 0,
            data.get('weeklyWorkingHours', get_row_value(old_row, 'WeeklyWorkingHours', 40.0)),
            min_staff_weekday,
            max_staff_weekday,
            min_staff_weekend,
            max_staff_weekend,
            max_consecutive_days,
            datetime.utcnow().isoformat(),
            request.session.get('user_email', 'system'),
            id
        ))
        
        # Log audit entry
        changes_dict = {}
        for field in ['code', 'name', 'startTime', 'endTime', 'durationHours', 'colorCode', 'isActive']:
            db_field = field[0].upper() + field[1:]
            if field == 'isActive':
                db_field = 'IsActive'
            elif field == 'startTime':
                db_field = 'StartTime'
            elif field == 'endTime':
                db_field = 'EndTime'
            elif field == 'durationHours':
                db_field = 'DurationHours'
            elif field == 'colorCode':
                db_field = 'ColorCode'
            else:
                db_field = field[0].upper() + field[1:]
            
            if field in data:
                old_val = old_row[db_field]
                new_val = data[field]
                if field == 'isActive':
                    old_val = bool(old_val)
                if old_val != new_val:
                    changes_dict[field] = {'old': old_val, 'new': new_val}
        
        if changes_dict:
            changes = json.dumps(changes_dict, ensure_ascii=False)
            log_audit(conn, 'ShiftType', id, 'Updated', changes)
        
        conn.commit()
        conn.close()
        
        return {'success': True}
        
    except Exception as e:
        logger.error(f"Update shift type error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Aktualisieren: {str(e)}'}, status_code=500)


@router.delete('/api/shifttypes/{id:int}', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def delete_shift_type(request: Request, id):
    """Delete shift type (Admin only)"""
    try:
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Check if shift type exists
        cursor.execute("SELECT Code, Name FROM ShiftTypes WHERE Id = ?", (id,))
        shift_row = cursor.fetchone()
        if not shift_row:
            conn.close()
            return JSONResponse(content={'error': 'Schichttyp nicht gefunden'}, status_code=404)
        
        # Check if shift type is used in assignments
        cursor.execute("SELECT COUNT(*) as count FROM ShiftAssignments WHERE ShiftTypeId = ?", (id,))
        assignment_count = cursor.fetchone()['count']
        
        if assignment_count > 0:
            conn.close()
            return JSONResponse(content={'error': f'Schichttyp wird in {assignment_count} Zuweisungen verwendet und kann nicht gelöscht werden'}, status_code=400)
        
        # Delete shift type (cascade will delete relationships and team assignments)
        cursor.execute("DELETE FROM ShiftTypes WHERE Id = ?", (id,))
        
        # Log audit entry
        changes = json.dumps({'code': shift_row['Code'], 'name': shift_row['Name']}, ensure_ascii=False)
        log_audit(conn, 'ShiftType', id, 'Deleted', changes)
        
        conn.commit()
        conn.close()
        
        return {'success': True}
        
    except Exception as e:
        logger.error(f"Delete shift type error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Löschen: {str(e)}'}, status_code=500)


# Team-Shift Assignment endpoints
@router.get('/api/shifttypes/{shift_id:int}/teams')

def get_shift_type_teams(request: Request, shift_id):
    """Get teams assigned to a shift type"""
    db = get_db()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT t.Id, t.Name
        FROM Teams t
        INNER JOIN TeamShiftAssignments tsa ON t.Id = tsa.TeamId
        WHERE tsa.ShiftTypeId = ?
        ORDER BY t.Name
    """, (shift_id,))
    
    teams = []
    for row in cursor.fetchall():
        teams.append({
            'id': row['Id'],
            'name': row['Name']
        })
    
    conn.close()
    return teams


@router.put('/api/shifttypes/{shift_id:int}/teams', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def update_shift_type_teams(request: Request, shift_id):
    """Update teams assigned to a shift type (Admin only)"""
    try:
        team_ids = data.get('teamIds', [])
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Check if shift type exists
        cursor.execute("SELECT Id FROM ShiftTypes WHERE Id = ?", (shift_id,))
        if not cursor.fetchone():
            conn.close()
            return JSONResponse(content={'error': 'Schichttyp nicht gefunden'}, status_code=404)
        
        # Delete all existing assignments for this shift
        cursor.execute("DELETE FROM TeamShiftAssignments WHERE ShiftTypeId = ?", (shift_id,))
        
        # Insert new assignments
        for team_id in team_ids:
            cursor.execute("""
                INSERT INTO TeamShiftAssignments (TeamId, ShiftTypeId, CreatedBy)
                VALUES (?, ?, ?)
            """, (team_id, shift_id, request.session.get('user_email', 'system')))
        
        # Log audit entry
        changes = json.dumps({'shiftTypeId': shift_id, 'teamIds': team_ids}, ensure_ascii=False)
        log_audit(conn, 'TeamShiftAssignment', shift_id, 'Updated', changes)
        
        conn.commit()
        conn.close()
        
        return {'success': True}
        
    except Exception as e:
        logger.error(f"Update shift type teams error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Aktualisieren: {str(e)}'}, status_code=500)


@router.get('/api/teams/{team_id:int}/shifttypes')

def get_team_shift_types(request: Request, team_id):
    """Get shift types assigned to a team"""
    db = get_db()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT st.Id, st.Code, st.Name, st.ColorCode
        FROM ShiftTypes st
        INNER JOIN TeamShiftAssignments tsa ON st.Id = tsa.ShiftTypeId
        WHERE tsa.TeamId = ? AND st.IsActive = 1
        ORDER BY st.Code
    """, (team_id,))
    
    shift_types = []
    for row in cursor.fetchall():
        shift_types.append({
            'id': row['Id'],
            'code': row['Code'],
            'name': row['Name'],
            'colorCode': row['ColorCode']
        })
    
    conn.close()
    return shift_types


@router.put('/api/teams/{team_id:int}/shifttypes', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def update_team_shift_types(request: Request, team_id):
    """Update shift types assigned to a team (Admin only)"""
    try:
        shift_type_ids = data.get('shiftTypeIds', [])
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Check if team exists
        cursor.execute("SELECT Id FROM Teams WHERE Id = ?", (team_id,))
        if not cursor.fetchone():
            conn.close()
            return JSONResponse(content={'error': 'Team nicht gefunden'}, status_code=404)
        
        # Delete all existing assignments for this team
        cursor.execute("DELETE FROM TeamShiftAssignments WHERE TeamId = ?", (team_id,))
        
        # Insert new assignments
        for shift_type_id in shift_type_ids:
            cursor.execute("""
                INSERT INTO TeamShiftAssignments (TeamId, ShiftTypeId, CreatedBy)
                VALUES (?, ?, ?)
            """, (team_id, shift_type_id, request.session.get('user_email', 'system')))
        
        # Log audit entry
        changes = json.dumps({'teamId': team_id, 'shiftTypeIds': shift_type_ids}, ensure_ascii=False)
        log_audit(conn, 'TeamShiftAssignment', team_id, 'Updated', changes)
        
        conn.commit()
        conn.close()
        
        return {'success': True}
        
    except Exception as e:
        logger.error(f"Update team shift types error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Aktualisieren: {str(e)}'}, status_code=500)


# ============================================================================
# SHIFT SCHEDULE ENDPOINT
# ============================================================================

@router.get('/api/shifts/schedule')

def get_schedule(request: Request):
    """Get shift schedule for a date range.

    Optional pagination query parameters:
      - page  (int, default 1): 1-based page number for assignments
      - limit (int, default 0): assignments per page; 0 means return all
    """
    start_date_str = request.query_params.get('startDate')
    end_date_str = request.query_params.get('endDate')
    view = request.query_params.get('view', 'week')

    # Parse pagination parameters
    try:
        page = max(1, int(request.query_params.get('page', 1)))
        limit = max(0, int(request.query_params.get('limit', 0)))
    except (ValueError, TypeError):
        return JSONResponse(content={'error': 'page and limit must be integers'}, status_code=400)

    if not start_date_str:
        return JSONResponse(content={'error': 'startDate is required'}, status_code=400)
    
    # Validate and parse dates
    try:
        start_date = date.fromisoformat(start_date_str)
    except (ValueError, TypeError):
        return JSONResponse(content={'error': 'Invalid startDate format. Use YYYY-MM-DD'}, status_code=400)
    
    # Calculate end date based on view
    if not end_date_str:
        if view == 'week':
            end_date = start_date + timedelta(days=6)
        elif view == 'month':
            # Get last day of month
            if start_date.month == 12:
                end_date = date(start_date.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(start_date.year, start_date.month + 1, 1) - timedelta(days=1)
            
            # Expand to complete calendar weeks (Sunday-Saturday)
            # Find Sunday of the week containing start_date (first day of month)
            start_weekday = start_date.weekday()  # Monday=0, Sunday=6
            if start_weekday != 6:  # Not Sunday
                days_back = start_weekday + 1
                start_date = start_date - timedelta(days=days_back)
            
            # Find Saturday of the week containing end_date (last day of month)
            end_weekday = end_date.weekday()  # Monday=0, Sunday=6
            if end_weekday != 5:  # If not Saturday already
                days_forward = (5 - end_weekday + 7) % 7
                end_date = end_date + timedelta(days=days_forward)
        elif view == 'year':
            end_date = date(start_date.year, 12, 31)
        else:
            end_date = start_date + timedelta(days=30)
    else:
        try:
            end_date = date.fromisoformat(end_date_str)
        except (ValueError, TypeError):
            return JSONResponse(content={'error': 'Invalid endDate format. Use YYYY-MM-DD'}, status_code=400)
    
    conn = None
    try:
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Check if user is admin (can see unapproved plans)
        user_roles = request.session.get('user_roles', [])
        is_admin = 'Admin' in user_roles
        
        # Get approved months if user is not admin
        approved_months = set()
        if not is_admin:
            cursor.execute("""
                SELECT Year, Month FROM ShiftPlanApprovals
                WHERE IsApproved = 1
            """)
            for row in cursor.fetchall():
                approved_months.add((row['Year'], row['Month']))
        
        # Get assignments
        cursor.execute("""
            SELECT sa.*, e.Vorname, e.Name, e.TeamId,
                   st.Code, st.Name as ShiftName, st.ColorCode
            FROM ShiftAssignments sa
            JOIN Employees e ON sa.EmployeeId = e.Id
            JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
            WHERE sa.Date >= ? AND sa.Date <= ?
            ORDER BY sa.Date, e.TeamId, e.Name, e.Vorname
        """, (start_date.isoformat(), end_date.isoformat()))
        
        assignments = []
        for row in cursor.fetchall():
            assignment_date = date.fromisoformat(row['Date'])
            year_month = (assignment_date.year, assignment_date.month)
            
            # Filter based on approval status for non-admin users
            if not is_admin and year_month not in approved_months:
                # Skip unapproved plans for regular users
                continue
            
            assignments.append({
                'id': row['Id'],
                'employeeId': row['EmployeeId'],
                'employeeName': f"{row['Vorname']} {row['Name']}",
                'teamId': row['TeamId'],
                'shiftTypeId': row['ShiftTypeId'],
                'shiftCode': row['Code'],
                'shiftName': row['ShiftName'],
                'colorCode': row['ColorCode'],
                'date': row['Date'],
                'isManual': bool(row['IsManual']),
                'isFixed': bool(row['IsFixed']),
                'notes': row['Notes']
            })
        
        # Get absences from Absences table
        cursor.execute("""
            SELECT a.*, e.Vorname, e.Name, e.TeamId
            FROM Absences a
            JOIN Employees e ON a.EmployeeId = e.Id
            WHERE (a.StartDate <= ? AND a.EndDate >= ?)
               OR (a.StartDate >= ? AND a.StartDate <= ?)
        """, (end_date.isoformat(), start_date.isoformat(), 
              start_date.isoformat(), end_date.isoformat()))
        
        absences = []
        for row in cursor.fetchall():
            type_index = row['Type']
            absences.append({
                'id': row['Id'],
                'employeeId': row['EmployeeId'],
                'employeeName': f"{row['Vorname']} {row['Name']}",
                'teamId': row['TeamId'],
                'type': ['', 'Krank', 'Urlaub', 'Lehrgang'][type_index],
                'status': 'Genehmigt' if type_index == 2 else None,  # Only Urlaub type has status
                'startDate': row['StartDate'],
                'endDate': row['EndDate'],
                'notes': row['Notes']
            })
        
        # Also get vacation requests (all statuses) and add them as absences
        cursor.execute("""
            SELECT vr.Id, vr.EmployeeId, vr.StartDate, vr.EndDate, vr.Notes, vr.Status,
                   e.Vorname, e.Name, e.TeamId
            FROM VacationRequests vr
            JOIN Employees e ON vr.EmployeeId = e.Id
            WHERE ((vr.StartDate <= ? AND vr.EndDate >= ?)
               OR (vr.StartDate >= ? AND vr.StartDate <= ?))
        """, (end_date.isoformat(), start_date.isoformat(),
              start_date.isoformat(), end_date.isoformat()))
        
        vacation_id_offset = 10000  # Offset to avoid ID conflicts
        for row in cursor.fetchall():
            # Determine the type label based on status
            if row['Status'] == 'Genehmigt':
                type_label = 'Urlaub'
                notes = row['Notes'] or 'Genehmigter Urlaub'
            elif row['Status'] == 'InBearbeitung':
                type_label = 'Urlaub (in Genehmigung)'
                notes = row['Notes'] or 'Urlaubsantrag in Bearbeitung'
            else:  # Abgelehnt
                type_label = 'Urlaub (abgelehnt)'
                notes = row['Notes'] or 'Urlaubsantrag abgelehnt'
            
            absences.append({
                'id': vacation_id_offset + row['Id'],
                'employeeId': row['EmployeeId'],
                'employeeName': f"{row['Vorname']} {row['Name']}",
                'teamId': row['TeamId'],
                'type': type_label,
                'status': row['Status'],  # Include status for color-coding
                'startDate': row['StartDate'],
                'endDate': row['EndDate'],
                'notes': notes
            })
        
        # Get vacation periods (Ferienzeiten) that overlap with the date range
        cursor.execute("""
            SELECT Id, Name, StartDate, EndDate, ColorCode
            FROM VacationPeriods
            WHERE (StartDate <= ? AND EndDate >= ?)
               OR (StartDate >= ? AND StartDate <= ?)
            ORDER BY StartDate
        """, (end_date.isoformat(), start_date.isoformat(),
              start_date.isoformat(), end_date.isoformat()))
        
        vacation_periods = []
        for row in cursor.fetchall():
            vacation_periods.append({
                'id': row['Id'],
                'name': row['Name'],
                'startDate': row['StartDate'],
                'endDate': row['EndDate'],
                'colorCode': row['ColorCode'] or '#E8F5E9'
            })
        
        pagination = _paginate(assignments, page, limit)

        return {
            'startDate': start_date.isoformat(),
            'endDate': end_date.isoformat(),
            'assignments': pagination['data'],
            'absences': absences,
            'vacationPeriods': vacation_periods,
            'pagination': {
                'total': pagination['total'],
                'page': pagination['page'],
                'limit': pagination['limit'],
                'totalPages': pagination['totalPages'],
            }
        }
        
    except Exception as e:
        return JSONResponse(content={'error': f'Database error: {str(e)}'}, status_code=500)
    finally:
        if conn:
            conn.close()


def _serialize_planning_report(report) -> str:
    """
    Serialize a PlanningReport dataclass instance to a JSON string.

    All ``date`` objects are converted to ISO-8601 strings so they round-trip
    safely through the database and the REST API.
    """
    from datetime import date as _date

    def _date_to_str(d) -> str:
        return d.isoformat() if isinstance(d, _date) else d

    data = {
        'planning_period': [
            _date_to_str(report.planning_period[0]),
            _date_to_str(report.planning_period[1]),
        ],
        'status': report.status,
        'total_employees': report.total_employees,
        'available_employees': report.available_employees,
        'absent_employees': [
            {
                'employee_name': a.employee_name,
                'absence_type': a.absence_type,
                'start_date': _date_to_str(a.start_date),
                'end_date': _date_to_str(a.end_date),
                'notes': a.notes,
            }
            for a in report.absent_employees
        ],
        'shifts_assigned': report.shifts_assigned,
        'uncovered_shifts': [
            {
                'date': _date_to_str(u.date),
                'shift_code': u.shift_code,
                'reason': u.reason,
            }
            for u in report.uncovered_shifts
        ],
        'rule_violations': [
            {
                'rule_id': v.rule_id,
                'description': v.description,
                'severity': v.severity,
                'affected_dates': [_date_to_str(d) for d in v.affected_dates],
                'cause': v.cause,
                'impact': v.impact,
            }
            for v in report.rule_violations
        ],
        'relaxed_constraints': [
            {
                'constraint_name': rc.constraint_name,
                'reason': rc.reason,
                'description': rc.description,
            }
            for rc in report.relaxed_constraints
        ],
        'objective_value': report.objective_value,
        'solver_time_seconds': report.solver_time_seconds,
        'penalty_breakdown': report.penalty_breakdown,
    }
    return json.dumps(data, ensure_ascii=False)


def _save_planning_report(db, year: int, month: int, report) -> None:
    """
    Persist a PlanningReport to the PlanningReports table.

    If a report already exists for the given year/month it is replaced.
    Errors are logged but do not propagate so that a serialization failure
    never prevents the caller from returning a successful response.
    """
    try:
        report_json = _serialize_planning_report(report)
        conn = db.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO PlanningReports (year, month, status, created_at, report_json)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(year, month) DO UPDATE SET
                    status      = excluded.status,
                    created_at  = excluded.created_at,
                    report_json = excluded.report_json
            """, (year, month, report.status, datetime.utcnow().isoformat(), report_json))
            conn.commit()
        finally:
            conn.close()
    except Exception as exc:
        logger.warning(f"Failed to save PlanningReport for {year}/{month}: {exc}")


def _cleanup_old_jobs(db):
    """Remove finished jobs older than 24 hours."""
    cutoff = (datetime.utcnow() - timedelta(hours=24)).isoformat()
    with db.connection() as conn:
        conn.execute(
            "DELETE FROM PlanningJobs WHERE finished_at IS NOT NULL AND finished_at < ?",
            (cutoff,)
        )
        conn.commit()

def _create_job(db, job_id: str):
    _cleanup_old_jobs(db)
    with db.connection() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO PlanningJobs (id, status, started_at) VALUES (?, 'running', ?)",
            (job_id, datetime.utcnow().isoformat())
        )
        conn.commit()

def _update_job(db, job_id: str, status: str, message: str = None, result_json: str = None):
    finished_at = datetime.utcnow().isoformat() if status in ('completed', 'error', 'cancelled', 'success') else None
    with db.connection() as conn:
        conn.execute(
            "UPDATE PlanningJobs SET status=?, message=?, finished_at=? WHERE id=?",
            (status, message, finished_at, job_id)
        )
        if result_json is not None:
            conn.execute("UPDATE PlanningJobs SET result_json=? WHERE id=?", (result_json, job_id))
        conn.commit()

def _get_job(db, job_id: str):
    with db.connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM PlanningJobs WHERE id=?", (job_id,))
        return cursor.fetchone()


def _run_planning_job(job_id: str, start_date, end_date, force: bool, db_path: str):
    """
    Standalone worker executed in a subprocess via ProcessPoolExecutor.
    Must not reference any FastAPI context objects – all imports are done locally
    and db is accessed directly via Database(db_path).
    """
    import logging as _logging
    import json as _json
    import sqlite3 as _sqlite3
    from datetime import datetime as _datetime, date as _date, timedelta as _timedelta

    _logging.basicConfig(
        level=_logging.INFO,
        format='%(asctime)s [%(levelname)s] %(name)s: %(message)s'
    )
    _logger = _logging.getLogger(__name__)

    try:
        from api.shared import Database, extend_planning_dates_to_complete_weeks
        db = Database(db_path)

        # Planning steps for progress display (1-based, shown in UI)
        _TOTAL_STEPS = 4

        def _update(status: str, message: str, step: int = None, **kwargs):
            data = {}
            if step is not None:
                data['planningStep'] = step
                data['planningTotalSteps'] = _TOTAL_STEPS
            data.update(kwargs)
            result_json = _json.dumps(data) if data else None
            _update_job(db, job_id, status, message, result_json)

        _update('running', 'Daten werden geladen…', step=1)

        # Extend planning dates to complete weeks (may extend into next month)
        extended_start, extended_end = extend_planning_dates_to_complete_weeks(start_date, end_date)
    
        # Log the extension for transparency
        _logger.info(f"Planning for {start_date} to {end_date}")
        if extended_end > end_date:
            _logger.info(f"Extended to complete week: {extended_start} to {extended_end} (added {(extended_end - end_date).days} days from next month)")
        
        # Load data
        from data_loader import load_from_database, load_global_settings
        employees, teams, absences, shift_types = load_from_database(db.db_path)
        
        # Load global settings (consecutive shifts limits, rest time, etc.)
        global_settings = load_global_settings(db.db_path)
        
        # Load existing assignments for the extended period (to lock days from adjacent months)
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Get existing assignments for days that extend beyond the current month
        # These will be locked so we don't overwrite already-planned shifts
        locked_team_shift = {}
        locked_employee_weekend = {}
        locked_employee_shift = {}  # NEW: Lock individual employee shifts to prevent double shifts
        
        # Query ALL existing shift assignments in the extended planning period
        # This prevents double shifts when planning across months
        # NOTE: This is separate from the team-level locking below because we need to
        # lock individual employee assignments for the ENTIRE period, not just adjacent months
        cursor.execute("""
            SELECT sa.EmployeeId, sa.Date, st.Code
            FROM ShiftAssignments sa
            INNER JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
                WHERE sa.Date >= ? AND sa.Date <= ?
            """, (extended_start.isoformat(), extended_end.isoformat()))
            
        existing_employee_assignments = cursor.fetchall()
        
        # Calculate weeks for boundary detection (needed for employee locks)
        # We'll skip locking employee assignments in boundary weeks to avoid conflicts
        from datetime import timedelta
        dates_list = []
        current = extended_start
        while current <= extended_end:
            dates_list.append(current)
            current += timedelta(days=1)
        
        # Calculate weeks
        weeks_for_boundary = []
        current_week = []
        for d in dates_list:
            if d.weekday() == 6 and current_week:  # Sunday
                weeks_for_boundary.append(current_week)
                current_week = []
            current_week.append(d)
        if current_week:
            weeks_for_boundary.append(current_week)
        
        # Identify boundary weeks (same logic as team lock boundary detection)
        boundary_week_dates = set()
        for week_dates in weeks_for_boundary:
            has_dates_before_month = any(d < start_date for d in week_dates)
            has_dates_in_month = any(start_date <= d <= end_date for d in week_dates)
            has_dates_after_month = any(d > end_date for d in week_dates)
            
            # If week spans the boundary, mark all its dates as boundary dates
            if (has_dates_before_month and has_dates_in_month) or (has_dates_in_month and has_dates_after_month):
                boundary_week_dates.update(week_dates)
                logger.info(f"Boundary week detected: {week_dates[0]} to {week_dates[-1]} - employee locks will be skipped")
        
        # Lock existing employee assignments
        # CRITICAL FIX: Skip locking employee assignments in boundary weeks
        # Boundary weeks span month boundaries and may have assignments that conflict
        # with current shift configuration or team-based rotation requirements
        for emp_id, date_str, shift_code in existing_employee_assignments:
            assignment_date = date.fromisoformat(date_str)
            
            # Skip assignments in boundary weeks - they will be re-planned to match current config
            if assignment_date in boundary_week_dates:
                logger.info(f"Skipping lock for Employee {emp_id}, Date {date_str} (in boundary week)")
                continue
            
            # CRITICAL FIX: Convert emp_id to int to match assignment.employee_id type
            # Database returns TEXT ids as strings, but solver uses integers
            try:
                emp_id_int = int(emp_id)
            except (ValueError, TypeError):
                # If conversion fails, use as-is (for backward compatibility with non-numeric IDs)
                emp_id_int = emp_id
            locked_employee_shift[(emp_id_int, assignment_date)] = shift_code
            logger.info(f"Locked: Employee {emp_id_int}, Date {date_str} -> {shift_code} (existing assignment)")
        
        if extended_end > end_date or extended_start < start_date:
            # Query existing shift assignments for extended dates ONLY (not the main month)
            # Join ShiftAssignments with Employees (for TeamId) and ShiftTypes (for Code)
            # Logic: Get assignments within extended range that are OUTSIDE main month range
            # This ensures we only lock assignments from adjacent months, not current month
            cursor.execute("""
                SELECT e.TeamId, sa.Date, st.Code
                FROM ShiftAssignments sa
                INNER JOIN Employees e ON sa.EmployeeId = e.Id
                INNER JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
                WHERE sa.Date >= ? AND sa.Date <= ?
                AND (sa.Date < ? OR sa.Date > ?)
                AND e.TeamId IS NOT NULL
            """, (extended_start.isoformat(), extended_end.isoformat(),
                  start_date.isoformat(), end_date.isoformat()))
            
            existing_team_assignments = cursor.fetchall()
            
            # Build locked constraints from existing assignments
            # We need to map dates to week indices
            from datetime import timedelta
            dates_list = []
            current = extended_start
            while current <= extended_end:
                dates_list.append(current)
                current += timedelta(days=1)
            
            # Calculate weeks
            weeks = []
            current_week = []
            for d in dates_list:
                if d.weekday() == 6 and current_week:  # Sunday
                    weeks.append(current_week)
                    current_week = []
                current_week.append(d)
            if current_week:
                weeks.append(current_week)
            
            # Map dates to week indices
            date_to_week = {}
            for week_idx, week_dates in enumerate(weeks):
                for d in week_dates:
                    date_to_week[d] = week_idx
            
            # Lock existing team assignments
            # CRITICAL FIX: Only lock team shifts for weeks entirely in adjacent months (not current month)
            # Weeks that span the boundary between adjacent and current months should NOT be locked
            # because they may have conflicting shifts (already-planned days vs. to-be-planned days)
            
            # Identify weeks that cross the month boundary
            boundary_weeks = set()
            for week_idx, week_dates in enumerate(weeks):
                # Check if this week contains dates both inside AND outside the main planning month
                has_dates_before_month = any(d < start_date for d in week_dates)
                has_dates_in_month = any(start_date <= d <= end_date for d in week_dates)
                has_dates_after_month = any(d > end_date for d in week_dates)
                
                # If week spans the boundary, don't lock it
                if (has_dates_before_month and has_dates_in_month) or (has_dates_in_month and has_dates_after_month):
                    boundary_weeks.add(week_idx)
                    logger.info(f"Week {week_idx} spans month boundary - will NOT be locked (dates: {week_dates[0]} to {week_dates[-1]})")
            
            # First pass: identify conflicts and boundary weeks
            conflicting_team_weeks = set()  # Track (team_id, week_idx) pairs with conflicts
            for team_id, date_str, shift_code in existing_team_assignments:
                assignment_date = date.fromisoformat(date_str)
                if assignment_date in date_to_week:
                    week_idx = date_to_week[assignment_date]
                    
                    # Skip weeks that cross the month boundary
                    if week_idx in boundary_weeks:
                        continue
                    
                    # Check for conflicts
                    if (team_id, week_idx) in locked_team_shift:
                        existing_shift = locked_team_shift[(team_id, week_idx)]
                        if existing_shift != shift_code:
                            # Conflict detected: different shift codes for same team/week
                            logger.warning(f"CONFLICT: Team {team_id}, Week {week_idx} has conflicting shifts: {existing_shift} vs {shift_code}")
                            conflicting_team_weeks.add((team_id, week_idx))
                    else:
                        # No conflict yet - tentatively add this lock
                        locked_team_shift[(team_id, week_idx)] = shift_code
            
            # Second pass: remove all conflicting locks
            for team_id, week_idx in conflicting_team_weeks:
                if (team_id, week_idx) in locked_team_shift:
                    logger.warning(f"  Removing team lock for Team {team_id}, Week {week_idx} to avoid INFEASIBLE")
                    del locked_team_shift[(team_id, week_idx)]
            
            # Log remaining locks
            for (team_id, week_idx), shift_code in locked_team_shift.items():
                logger.info(f"Locked: Team {team_id}, Week {week_idx} -> {shift_code} (from existing assignments)")
        
        conn.close()
        
        # Load previous shifts for cross-month consecutive days checking
        # CRITICAL FIX: Extended lookback to capture full consecutive chains
        max_consecutive_limit = max((st.max_consecutive_days for st in shift_types), default=7)
        
        # Maximum lookback period to prevent excessive database queries
        max_lookback_days = 60
        
        previous_employee_shifts = {}
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # First pass: Load initial lookback period (same as before)
        initial_lookback_start = extended_start - timedelta(days=max_consecutive_limit)
        initial_lookback_end = extended_start - timedelta(days=1)
        
        cursor.execute("""
            SELECT sa.EmployeeId, sa.Date, st.Code
            FROM ShiftAssignments sa
            INNER JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
            WHERE sa.Date >= ? AND sa.Date <= ?
            ORDER BY sa.Date
        """, (initial_lookback_start.isoformat(), initial_lookback_end.isoformat()))
        
        initial_shifts = cursor.fetchall()
        
        # Group shifts by employee for analysis
        employee_shift_dates = {}
        for emp_id, date_str, shift_code in initial_shifts:
            shift_date = date.fromisoformat(date_str)
            try:
                emp_id_int = int(emp_id)
            except (ValueError, TypeError):
                emp_id_int = emp_id
                
            if emp_id_int not in employee_shift_dates:
                employee_shift_dates[emp_id_int] = []
            employee_shift_dates[emp_id_int].append((shift_date, shift_code))
            previous_employee_shifts[(emp_id_int, shift_date)] = shift_code
        
        # Second pass: For each employee with shifts at the start of lookback period,
        # extend lookback to capture their full consecutive chain
        employees_to_extend = []
        for emp_id, shifts in employee_shift_dates.items():
            if not shifts:
                continue
            
            # Sort by date
            shifts.sort(key=lambda x: x[0])
            
            # Check if employee has shifts at the very beginning of lookback period
            # If so, they might have more consecutive days further back
            earliest_shift_date = shifts[0][0]
            
            # Check if there's a consecutive chain leading up to extended_start
            # Work backwards from extended_start - 1 to find consecutive days
            consecutive_days = 0
            check_date = extended_start - timedelta(days=1)
            # Check max_consecutive_limit days to see if all have shifts
            for _ in range(max_consecutive_limit):
                has_shift = any(shift_date == check_date for shift_date, _ in shifts)
                if has_shift:
                    consecutive_days += 1
                    check_date -= timedelta(days=1)
                else:
                    break
            
            # If we found exactly max_consecutive_limit consecutive days without breaking,
            # the chain might extend further back. We need extended lookback to find out.
            if consecutive_days == max_consecutive_limit:
                employees_to_extend.append(emp_id)
        
        # Extend lookback for employees who need it
        if employees_to_extend:
            extended_lookback_start = extended_start - timedelta(days=max_lookback_days)
            extended_lookback_end = initial_lookback_start - timedelta(days=1)
            
            logger.info(f"Extending lookback for {len(employees_to_extend)} employees with long consecutive chains")
            
            # Query extended period for these employees only
            # Use parameterized query to prevent SQL injection
            placeholders = ','.join('?' * len(employees_to_extend))
            query = f"""
                SELECT sa.EmployeeId, sa.Date, st.Code
                FROM ShiftAssignments sa
                INNER JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
                WHERE sa.Date >= ? AND sa.Date <= ?
                AND sa.EmployeeId IN ({placeholders})
                ORDER BY sa.Date
            """
            params = [extended_lookback_start.isoformat(), extended_lookback_end.isoformat()] + employees_to_extend
            cursor.execute(query, params)
            
            for emp_id, date_str, shift_code in cursor.fetchall():
                shift_date = date.fromisoformat(date_str)
                try:
                    emp_id_int = int(emp_id)
                except (ValueError, TypeError):
                    emp_id_int = emp_id
                previous_employee_shifts[(emp_id_int, shift_date)] = shift_code
        
        conn.close()
        
        logger.info(f"Loaded {len(previous_employee_shifts)} previous shift assignments for consecutive days checking")
        if previous_employee_shifts:
            # Find actual date range
            all_dates = [d for (_, d) in previous_employee_shifts.keys()]
            if all_dates:
                actual_lookback_start = min(all_dates)
                actual_lookback_end = max(all_dates)
                logger.info(f"  Previous shifts date range: {actual_lookback_start} to {actual_lookback_end}")
                if employees_to_extend:
                    logger.info(f"  Extended lookback for {len(employees_to_extend)} employees to capture full consecutive chains")
        
        # Create model with extended dates and locked constraints
        _update('running', 'Planungsmodell wird erstellt…', step=2)
        from model import create_shift_planning_model
        from solver import solve_shift_planning, get_infeasibility_diagnostics
        planning_model = create_shift_planning_model(
            employees, teams, extended_start, extended_end, absences, 
            shift_types=shift_types,
            locked_team_shift=locked_team_shift if locked_team_shift else None,
            locked_employee_shift=locked_employee_shift if locked_employee_shift else None,
            previous_employee_shifts=previous_employee_shifts if previous_employee_shifts else None
        )
        
        # Check if cancelled before starting the solve
        row = _get_job(db, job_id)
        if row and row['status'] == 'cancelled':
            return

        # Load the previous month's completed shift assignments as warmstart hints.
        # These are passed to the solver via warm_start_shifts so that CP-SAT starts
        # near a known-good solution, reducing time-to-first-feasible by 20–40 %.
        # Only the month directly before start_date is used; older history is ignored.
        warm_start_shifts = {}
        try:
            prev_month_end = start_date - timedelta(days=1)
            prev_month_start = prev_month_end.replace(day=1)
            conn_ws = db.get_connection()
            cursor_ws = conn_ws.cursor()
            cursor_ws.execute(
                """
                SELECT sa.EmployeeId, sa.Date, st.Code
                FROM ShiftAssignments sa
                INNER JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
                WHERE sa.Date >= ? AND sa.Date <= ?
                """,
                (prev_month_start.isoformat(), prev_month_end.isoformat()),
            )
            for emp_id, date_str, shift_code in cursor_ws.fetchall():
                try:
                    emp_id_int = int(emp_id)
                except (ValueError, TypeError):
                    emp_id_int = emp_id
                warm_start_shifts[(emp_id_int, date.fromisoformat(date_str))] = shift_code
            conn_ws.close()
            if warm_start_shifts:
                logger.info(
                    f"Warmstart: loaded {len(warm_start_shifts)} previous-month assignments "
                    f"({prev_month_start} – {prev_month_end}) as solver hints"
                )
        except Exception as _ws_err:
            logger.warning(f"Warmstart hint loading failed (non-critical): {_ws_err}")
            warm_start_shifts = {}

        # Solve
        # SOLVER_TIME_LIMIT_SECONDS can be set in Flask config for test environments.
        # Production leaves it unset (None = unlimited).
        _update('running', 'Optimierung läuft… (dies kann mehrere Minuten dauern)', step=3)
        solver_time_limit = None  # use default
        result = solve_shift_planning(
            planning_model,
            global_settings=global_settings,
            db_path=db.db_path,
            time_limit_seconds=solver_time_limit,
            warm_start_shifts=warm_start_shifts if warm_start_shifts else None,
        )
        
        if not result:
            # Get diagnostic information to help user understand the issue
            diagnostics = get_infeasibility_diagnostics(planning_model)
            
            # Build helpful error message with root cause analysis
            error_details = []
            error_details.append(f"Planung für {start_date.strftime('%d.%m.%Y')} bis {end_date.strftime('%d.%m.%Y')} nicht möglich.")
            error_details.append("")
            error_details.append("GRUNDINFORMATIONEN:")
            error_details.append(f"• Mitarbeiter gesamt: {diagnostics['total_employees']}")
            error_details.append(f"• Teams: {diagnostics['total_teams']}")
            error_details.append(f"• Planungszeitraum: {diagnostics['planning_days']} Tage ({diagnostics['planning_weeks']:.1f} Wochen)")
            
            if diagnostics['employees_with_absences'] > 0:
                error_details.append(f"• Mitarbeiter mit Abwesenheiten: {diagnostics['employees_with_absences']}")
                error_details.append(f"• Abwesenheitstage gesamt: {diagnostics['total_absence_days']} von {diagnostics['total_employees'] * diagnostics['planning_days']} ({diagnostics['absence_ratio']*100:.1f}%)")
            
            # Add specific issues - these are the root causes
            if diagnostics['potential_issues']:
                error_details.append("")
                error_details.append("URSACHEN (Warum die Planung nicht möglich ist):")
                for i, issue in enumerate(diagnostics['potential_issues'], 1):
                    error_details.append(f"{i}. {issue}")
            else:
                error_details.append("")
                error_details.append("URSACHE:")
                error_details.append("Die genaue Ursache konnte nicht automatisch ermittelt werden.")
                error_details.append("Mögliche Gründe:")
                error_details.append("• Zu viele Abwesenheiten im Planungszeitraum")
                error_details.append("• Zu wenige Mitarbeiter für die erforderliche Schichtbesetzung")
                error_details.append("• Konflikte zwischen Ruhezeiten und Schichtzuweisungen")
                error_details.append("• Teams sind zu klein für die Rotationsanforderungen")
            
            # Add staffing analysis for shifts with issues
            problem_shifts = [shift for shift, data in diagnostics['shift_analysis'].items() 
                             if not data['is_feasible']]
            if problem_shifts:
                error_details.append("")
                error_details.append("SCHICHTBESETZUNGSPROBLEME:")
                for shift in problem_shifts:
                    data = diagnostics['shift_analysis'][shift]
                    error_details.append(f"• Schicht {shift}: Nur {data['eligible_employees']} Mitarbeiter verfügbar, aber {data['min_required']} erforderlich")
            
            error_message = "\n".join(error_details)
            
            _update('error',
                    'Planung fehlgeschlagen',
                    details=error_message,
                    diagnostics={
                        'total_employees': diagnostics['total_employees'],
                        'available_employees': diagnostics['available_employees'],
                        'employees_with_absences': diagnostics['employees_with_absences'],
                        'absent_employees': diagnostics['absent_employees'],
                        'potential_issues': diagnostics['potential_issues'],
                        'shift_analysis': diagnostics.get('shift_analysis', {})
                    })
            return
        
        assignments, complete_schedule, planning_report = result
        
        # Filter assignments to include:
        # 1. All days in the requested month (start_date to end_date)
        # 2. Extended days into NEXT month (end_date < date <= extended_end) - to maintain rotation continuity
        # 3. EXCLUDE extended days from PREVIOUS month (extended_start <= date < start_date) - these should already exist
        filtered_assignments = [a for a in assignments if start_date <= a.date <= extended_end]
        
        # Count assignments by category for logging
        current_month_count = len([a for a in filtered_assignments if start_date <= a.date <= end_date])
        future_extended_count = len([a for a in filtered_assignments if a.date > end_date])
        past_excluded_count = len([a for a in assignments if a.date < start_date])
        
        logger.info(f"Total assignments generated: {len(assignments)}")
        logger.info(f"  - Current month ({start_date} to {end_date}): {current_month_count}")
        if future_extended_count > 0:
            logger.info(f"  - Extended into next month ({end_date + timedelta(days=1)} to {extended_end}): {future_extended_count}")
        if past_excluded_count > 0:
            logger.info(f"  - Excluded from previous month (already planned): {past_excluded_count}")
        
        # Save to database
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Delete existing non-fixed assignments for current month AND future extended days
        # (but NOT for past extended days - those were planned by previous month)
        if force:
            cursor.execute("""
                DELETE FROM ShiftAssignments 
                WHERE Date >= ? AND Date <= ? AND IsFixed = 0
            """, (start_date.isoformat(), extended_end.isoformat()))
        
        # Insert new assignments (current month + future extended days)
        # CRITICAL FIX: Skip assignments that are locked (already exist from previous planning)
        # This prevents duplicate shifts when planning months that overlap with previously planned weeks
        skipped_locked = 0
        inserted = 0
        for assignment in filtered_assignments:
            # Check if this assignment was locked (already exists from previous month)
            if (assignment.employee_id, assignment.date) in locked_employee_shift:
                # Skip inserting - this assignment already exists in the database
                # It was loaded as a locked constraint and should not be duplicated
                skipped_locked += 1
                continue
            
            # CRITICAL: Check if assignment already exists (safety against double shifts)
            # With unique constraint on (EmployeeId, Date), this prevents database errors
            cursor.execute("""
                SELECT Id FROM ShiftAssignments 
                WHERE EmployeeId = ? AND Date = ?
            """, (assignment.employee_id, assignment.date.isoformat()))
            
            if cursor.fetchone():
                # Assignment already exists - skip to prevent duplicate
                skipped_locked += 1
                continue
            
            cursor.execute("""
                INSERT INTO ShiftAssignments 
                (EmployeeId, ShiftTypeId, Date, IsManual, IsFixed, CreatedAt, CreatedBy)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                assignment.employee_id,
                assignment.shift_type_id,
                assignment.date.isoformat(),
                0,
                0,
                datetime.utcnow().isoformat(),
                "Python-OR-Tools"
            ))
            inserted += 1
        
        logger.info(f"Inserted {inserted} new assignments, skipped {skipped_locked} locked assignments")
        
        # TD (Tag Dienst / Day Duty) assignments have been removed from the system
        # This section is no longer used
        
        # Create or update approval record for this month (not approved by default)
        cursor.execute("""
            INSERT INTO ShiftPlanApprovals (Year, Month, IsApproved, CreatedAt)
            VALUES (?, ?, 0, ?)
            ON CONFLICT(Year, Month) DO UPDATE SET
                IsApproved = 0,
                ApprovedAt = NULL,
                ApprovedBy = NULL,
                ApprovedByName = NULL
        """, (start_date.year, start_date.month, datetime.utcnow().isoformat()))
        
        conn.commit()
        conn.close()

        # Serialize and persist the PlanningReport so it can be retrieved later
        _update('running', 'Schichten werden gespeichert…', step=4)
        _save_planning_report(db, start_date.year, start_date.month, planning_report)

        report_url = f"/api/planning/report/{start_date.year}/{start_date.month}"

        _update('success',
                f'Erfolgreich! {len(filtered_assignments)} Schichten wurden geplant.',
                assignmentsCount=len(filtered_assignments),
                year=start_date.year,
                month=start_date.month,
                report_url=report_url,
                extendedPlanning={
                    'extendedEnd': extended_end.isoformat() if extended_end > end_date else None,
                    'daysExtended': (extended_end - end_date).days if extended_end > end_date else 0
                })

    except Exception as exc:
        _logger.exception(f"Planning job {job_id} failed")
        _update('error', 'Unbekannter Fehler', details=str(exc))


@router.post('/api/shifts/plan', dependencies=[Depends(require_role('Admin', 'Disponent')), Depends(check_csrf)])
def plan_shifts(request: Request):
    """
    Start asynchronous shift planning using OR-Tools.

    Returns a job_id immediately; the caller should poll
    GET /api/shifts/plan/status/{job_id} for progress and result.
    """
    start_date_str = request.query_params.get('startDate')
    end_date_str = request.query_params.get('endDate')
    force = request.query_params.get('force', 'false').lower() == 'true'
    if not start_date_str or not end_date_str:
        return JSONResponse(content={'error': 'startDate and endDate are required'}, status_code=400)

    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)

        # Validate that planning is for a complete single month
        is_valid, error_msg = validate_monthly_date_range(start_date, end_date)
        if not is_valid:
            return JSONResponse(content={'error': error_msg}, status_code=400)

        # Enforce max concurrent job limit
        with _futures_lock:
            running = sum(1 for f in _active_futures.values() if not f.done())
            if running >= MAX_CONCURRENT_JOBS:
                return JSONResponse(
                    content={'error': f'Maximale Anzahl gleichzeitiger Planungsjobs ({MAX_CONCURRENT_JOBS}) erreicht. Bitte warten Sie, bis ein Job abgeschlossen ist.'},
                    status_code=503
                )

        # Create job entry and submit to process pool
        job_id = str(uuid.uuid4())
        db = get_db()
        _create_job(db, job_id)

        future = _solver_pool.submit(
            _run_planning_job,
            job_id, start_date, end_date, force, db.db_path
        )
        with _futures_lock:
            _active_futures[job_id] = future


        return JSONResponse(content={'jobId': job_id, 'status': 'running'}, status_code=202)

    except Exception as e:
        return JSONResponse(content={'error': str(e)}, status_code=500)


@router.get('/api/shifts/plan/status/{job_id}', dependencies=[Depends(require_role('Admin', 'Disponent'))])
def get_plan_status(request: Request, job_id: str):
    """
    Poll the status of a background planning job.

    Returns:
        status: 'running' | 'success' | 'error'
        message: human-readable status text
        (on success) assignmentsCount, year, month, extendedPlanning
        (on error)   details, diagnostics
    """
    db = get_db()
    job = _get_job(db, job_id)
    if job is None:
        return JSONResponse(content={'error': 'Job not found'}, status_code=404)

    result = {
        'status': job['status'],
        'message': job['message'],
    }
    if job['result_json']:
        try:
            result.update(json.loads(job['result_json']))
        except Exception:
            pass

    if job['started_at']:
        try:
            started = datetime.fromisoformat(job['started_at'])
            elapsed = int((datetime.utcnow() - started).total_seconds())
            result['elapsedSeconds'] = elapsed
        except Exception:
            result['elapsedSeconds'] = 0

    return result


@router.delete('/api/shifts/plan/{job_id}', dependencies=[Depends(require_role('Admin', 'Disponent')), Depends(check_csrf)])

def cancel_plan_job(request: Request, job_id):
    """
    Request cancellation of a background planning job.

    The job is marked as cancelled. Since OR-Tools may not immediately stop,
    this sets the job status to 'cancelled' in the job store.
    """
    db = get_db()
    job = _get_job(db, job_id)
    if job is None:
        return JSONResponse(content={'error': 'Job not found'}, status_code=404)
    if job['status'] != 'running':
        return JSONResponse(content={'error': 'Job is not running'}, status_code=400)
    _update_job(db, job_id, 'cancelled', 'Planung wurde abgebrochen.')
    return {'success': True, 'message': 'Planung wird abgebrochen.'}


@router.get('/api/shifts/plan/approvals', dependencies=[Depends(require_role('Admin'))])

def get_plan_approvals(request: Request):
    """Get all shift plan approvals (Admin only)"""
    try:
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT * FROM ShiftPlanApprovals
            ORDER BY Year DESC, Month DESC
        """)
        
        approvals = []
        for row in cursor.fetchall():
            approvals.append({
                'id': row['Id'],
                'year': row['Year'],
                'month': row['Month'],
                'isApproved': bool(row['IsApproved']),
                'approvedAt': row['ApprovedAt'],
                'approvedBy': row['ApprovedBy'],
                'approvedByName': row['ApprovedByName'],
                'notes': row['Notes'],
                'createdAt': row['CreatedAt']
            })
        
        conn.close()
        return approvals
        
    except Exception as e:
        return JSONResponse(content={'error': str(e)}, status_code=500)


@router.get('/api/shifts/plan/approvals/{year:int}/{month:int}')

def get_plan_approval_status(request: Request, year, month):
    """Get approval status for a specific month"""
    try:
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT * FROM ShiftPlanApprovals
            WHERE Year = ? AND Month = ?
        """, (year, month))
        
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            return {
                'year': year,
                'month': month,
                'isApproved': False,
                'exists': False
            }
        
        return {
            'id': row['Id'],
            'year': row['Year'],
            'month': row['Month'],
            'isApproved': bool(row['IsApproved']),
            'approvedAt': row['ApprovedAt'],
            'approvedBy': row['ApprovedBy'],
            'approvedByName': row['ApprovedByName'],
            'notes': row['Notes'],
            'createdAt': row['CreatedAt'],
            'exists': True
        }
        
    except Exception as e:
        return JSONResponse(content={'error': str(e)}, status_code=500)


@router.put('/api/shifts/plan/approvals/{year:int}/{month:int}', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def approve_plan(request: Request, year, month):
    """Approve or unapprove a shift plan for a specific month (Admin only)"""
    try:
        is_approved = data.get('isApproved', True)
        notes = data.get('notes', '')
        
        # Get current user info
        user_id = request.session.get('user_id')
        user_name = request.session.get('user_fullname', 'Unknown Admin')
        
        if not user_id:
            return JSONResponse(content={'error': 'User not authenticated'}, status_code=401)
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Update or insert approval record
        if is_approved:
            cursor.execute("""
                INSERT INTO ShiftPlanApprovals (Year, Month, IsApproved, ApprovedAt, ApprovedBy, ApprovedByName, Notes, CreatedAt)
                VALUES (?, ?, 1, ?, ?, ?, ?, ?)
                ON CONFLICT(Year, Month) DO UPDATE SET
                    IsApproved = 1,
                    ApprovedAt = ?,
                    ApprovedBy = ?,
                    ApprovedByName = ?,
                    Notes = ?
            """, (year, month, datetime.utcnow().isoformat(), user_id, user_name, notes,
                  datetime.utcnow().isoformat(),
                  datetime.utcnow().isoformat(), user_id, user_name, notes))
        else:
            cursor.execute("""
                INSERT INTO ShiftPlanApprovals (Year, Month, IsApproved, CreatedAt)
                VALUES (?, ?, 0, ?)
                ON CONFLICT(Year, Month) DO UPDATE SET
                    IsApproved = 0,
                    ApprovedAt = NULL,
                    ApprovedBy = NULL,
                    ApprovedByName = NULL,
                    Notes = ?
            """, (year, month, datetime.utcnow().isoformat(), notes))
        
        conn.commit()
        conn.close()
        
        action = 'freigegeben' if is_approved else 'Freigabe aufgehoben'
        return {
            'success': True,
            'message': f'Dienstplan für {month:02d}/{year} wurde {action}.'
        }
        
    except Exception as e:
        return JSONResponse(content={'error': str(e)}, status_code=500)


@router.put('/api/shifts/assignments/{id:int}', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def update_shift_assignment(request: Request, id):
    """Update a shift assignment (manual edit)"""
    try:
        
        # Validate data types
        try:
            employee_id = int(data.get('employeeId'))
            shift_type_id = int(data.get('shiftTypeId'))
            assignment_date = date.fromisoformat(data.get('date'))
        except (ValueError, TypeError) as e:
            return JSONResponse(content={'error': f'Ungültige Daten: {str(e)}'}, status_code=400)
        
        conn = None
        try:
            db = get_db()
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if assignment exists and get old values for audit
            cursor.execute("""
                SELECT EmployeeId, ShiftTypeId, Date, IsFixed, Notes 
                FROM ShiftAssignments WHERE Id = ?
            """, (id,))
            old_row = cursor.fetchone()
            if not old_row:
                return JSONResponse(content={'error': 'Schichtzuweisung nicht gefunden'}, status_code=404)
            
            # Update assignment
            cursor.execute("""
                UPDATE ShiftAssignments 
                SET EmployeeId = ?, ShiftTypeId = ?, Date = ?, 
                    IsManual = 1, IsFixed = ?, Notes = ?,
                    ModifiedAt = ?, ModifiedBy = ?
                WHERE Id = ?
            """, (
                employee_id,
                shift_type_id,
                assignment_date.isoformat(),
                1 if data.get('isFixed') else 0,
                data.get('notes'),
                datetime.utcnow().isoformat(),
                request.session.get('user_email'),
                id
            ))
            
            # Log audit entry with changes
            changes_dict = {}
            if old_row['EmployeeId'] != employee_id:
                changes_dict['employeeId'] = {'old': old_row['EmployeeId'], 'new': employee_id}
            if old_row['ShiftTypeId'] != shift_type_id:
                changes_dict['shiftTypeId'] = {'old': old_row['ShiftTypeId'], 'new': shift_type_id}
            if old_row['Date'] != assignment_date.isoformat():
                changes_dict['date'] = {'old': old_row['Date'], 'new': assignment_date.isoformat()}
            new_is_fixed = 1 if data.get('isFixed') else 0
            if old_row['IsFixed'] != new_is_fixed:
                changes_dict['isFixed'] = {'old': bool(old_row['IsFixed']), 'new': bool(new_is_fixed)}
            if old_row['Notes'] != data.get('notes'):
                changes_dict['notes'] = {'old': old_row['Notes'], 'new': data.get('notes')}
            
            if changes_dict:
                changes = json.dumps(changes_dict, ensure_ascii=False)
                log_audit(conn, 'ShiftAssignment', id, 'Updated', changes)
            
            conn.commit()
            
            return {'success': True}
            
        finally:
            if conn:
                conn.close()
        
    except Exception as e:
        logger.error(f"Update shift assignment error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Aktualisieren: {str(e)}'}, status_code=500)


@router.post('/api/shifts/assignments', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def create_shift_assignment(request: Request):
    """Create a shift assignment manually"""
    try:
        
        # Validate required fields
        if not data.get('employeeId') or not data.get('shiftTypeId') or not data.get('date'):
            return JSONResponse(content={'error': 'EmployeeId, ShiftTypeId und Date sind erforderlich'}, status_code=400)
        
        # Validate data types
        try:
            employee_id = int(data.get('employeeId'))
            shift_type_id = int(data.get('shiftTypeId'))
            assignment_date = date.fromisoformat(data.get('date'))
        except (ValueError, TypeError) as e:
            return JSONResponse(content={'error': f'Ungültige Daten: {str(e)}'}, status_code=400)
        
        conn = None
        try:
            db = get_db()
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check for existing assignment (same employee, date, shift type)
            cursor.execute("""
                SELECT Id FROM ShiftAssignments 
                WHERE EmployeeId = ? AND Date = ? AND ShiftTypeId = ?
            """, (employee_id, assignment_date.isoformat(), shift_type_id))
            
            if cursor.fetchone():
                return JSONResponse(content={'error': 'Diese Schichtzuweisung existiert bereits'}, status_code=400)
            
            # Create assignment
            cursor.execute("""
                INSERT INTO ShiftAssignments 
                (EmployeeId, ShiftTypeId, Date, IsManual, IsFixed, Notes, CreatedAt, CreatedBy)
                VALUES (?, ?, ?, 1, ?, ?, ?, ?)
            """, (
                employee_id,
                shift_type_id,
                assignment_date.isoformat(),
                1 if data.get('isFixed') else 0,
                data.get('notes'),
                datetime.utcnow().isoformat(),
                request.session.get('user_email')
            ))
            
            assignment_id = cursor.lastrowid
            
            # Log audit entry
            changes = json.dumps({
                'employeeId': employee_id,
                'shiftTypeId': shift_type_id,
                'date': assignment_date.isoformat(),
                'isFixed': data.get('isFixed'),
                'notes': data.get('notes')
            }, ensure_ascii=False)
            log_audit(conn, 'ShiftAssignment', assignment_id, 'Created', changes)
            
            conn.commit()
            
            return JSONResponse(content={'success': True, 'id': assignment_id}, status_code=201)
            
        finally:
            if conn:
                conn.close()
        
    except Exception as e:
        logger.error(f"Create shift assignment error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Erstellen: {str(e)}'}, status_code=500)


@router.delete('/api/shifts/assignments/{id:int}', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def delete_shift_assignment(request: Request, id):
    """Delete a shift assignment"""
    try:
        conn = None
        try:
            db = get_db()
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if assignment exists and get info for audit
            cursor.execute("""
                SELECT EmployeeId, ShiftTypeId, Date, IsFixed 
                FROM ShiftAssignments WHERE Id = ?
            """, (id,))
            row = cursor.fetchone()
            
            if not row:
                return JSONResponse(content={'error': 'Schichtzuweisung nicht gefunden'}, status_code=404)
            
            # Warn if trying to delete a fixed assignment
            if row['IsFixed']:
                return JSONResponse(content={'error': 'Fixierte Schichtzuweisungen können nicht gelöscht werden. Bitte erst entsperren.'}, status_code=400)
            
            # Delete assignment
            cursor.execute("DELETE FROM ShiftAssignments WHERE Id = ?", (id,))
            
            # Log audit entry
            changes = json.dumps({
                'employeeId': row['EmployeeId'],
                'shiftTypeId': row['ShiftTypeId'],
                'date': row['Date']
            }, ensure_ascii=False)
            log_audit(conn, 'ShiftAssignment', id, 'Deleted', changes)
            
            conn.commit()
            
            return {'success': True}
            
        finally:
            if conn:
                conn.close()
        
    except Exception as e:
        logger.error(f"Delete shift assignment error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Löschen: {str(e)}'}, status_code=500)


@router.put('/api/shifts/assignments/bulk', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def bulk_update_shift_assignments(request: Request):
    """Bulk update multiple shift assignments"""
    try:
        
        # Validate required fields
        if not data.get('shiftIds') or not isinstance(data.get('shiftIds'), list):
            return JSONResponse(content={'error': 'ShiftIds array ist erforderlich'}, status_code=400)
        
        if not data.get('changes') or not isinstance(data.get('changes'), dict):
            return JSONResponse(content={'error': 'Changes object ist erforderlich'}, status_code=400)
        
        shift_ids = data['shiftIds']
        changes = data['changes']
        
        if len(shift_ids) == 0:
            return JSONResponse(content={'error': 'Keine Schichten zum Aktualisieren ausgewählt'}, status_code=400)
        
        # Validate that at least one field is being changed
        allowed_fields = {'employeeId', 'shiftTypeId', 'isFixed', 'notes'}
        if not any(key in changes for key in allowed_fields):
            return JSONResponse(content={'error': 'Mindestens ein Feld muss geändert werden'}, status_code=400)
        
        # Validate that only allowed fields are present
        invalid_fields = set(changes.keys()) - allowed_fields
        if invalid_fields:
            return JSONResponse(content={'error': f'Ungültige Felder: {", ".join(invalid_fields)}'}, status_code=400)
        
        conn = None
        updated_count = 0
        try:
            db = get_db()
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Whitelist for allowed column names to prevent SQL injection
            ALLOWED_COLUMNS = {
                'employeeId': 'EmployeeId',
                'shiftTypeId': 'ShiftTypeId',
                'isFixed': 'IsFixed',
                'notes': 'Notes'
            }
            
            # Process each shift
            for shift_id in shift_ids:
                # Verify shift exists
                cursor.execute("SELECT Id FROM ShiftAssignments WHERE Id = ?", (shift_id,))
                if not cursor.fetchone():
                    logger.warning(f"Shift {shift_id} not found, skipping")
                    continue
                
                # Build UPDATE query dynamically based on changes
                update_fields = []
                update_values = []
                
                if 'employeeId' in changes:
                    update_fields.append(f"{ALLOWED_COLUMNS['employeeId']} = ?")
                    update_values.append(changes['employeeId'])
                
                if 'shiftTypeId' in changes:
                    update_fields.append(f"{ALLOWED_COLUMNS['shiftTypeId']} = ?")
                    update_values.append(changes['shiftTypeId'])
                
                if 'isFixed' in changes:
                    update_fields.append(f"{ALLOWED_COLUMNS['isFixed']} = ?")
                    update_values.append(1 if changes['isFixed'] else 0)
                
                if 'notes' in changes:
                    # Append notes instead of replacing
                    cursor.execute("SELECT Notes FROM ShiftAssignments WHERE Id = ?", (shift_id,))
                    row = cursor.fetchone()
                    existing_notes = row['Notes'] if row and row['Notes'] else ''
                    new_notes = existing_notes
                    if existing_notes:
                        new_notes += '\n' + changes['notes']
                    else:
                        new_notes = changes['notes']
                    update_fields.append(f"{ALLOWED_COLUMNS['notes']} = ?")
                    update_values.append(new_notes)
                
                # Always update modification timestamp and user
                update_fields.append("ModifiedAt = ?")
                update_fields.append("ModifiedBy = ?")
                update_values.append(datetime.utcnow().isoformat())
                update_values.append(request.session.get('user_email'))
                
                # Add shift ID as last parameter
                update_values.append(shift_id)
                
                # Execute update - fields are from whitelist, safe to use in f-string
                update_query = f"""
                    UPDATE ShiftAssignments 
                    SET {', '.join(update_fields)}
                    WHERE Id = ?
                """
                
                cursor.execute(update_query, update_values)
                
                # Log audit entry
                changes_json = json.dumps(changes, ensure_ascii=False)
                log_audit(conn, 'ShiftAssignment', shift_id, 'BulkUpdate', changes_json)
                
                updated_count += 1
            
            conn.commit()
            
            return {
                'success': True,
                'updated': updated_count,
                'total': len(shift_ids)
            }
            
        finally:
            if conn:
                conn.close()
        
    except ValueError as e:
        logger.error(f"Bulk update validation error: {str(e)}")
        return JSONResponse(content={'error': f'Validierungsfehler: {str(e)}'}, status_code=400)
    except Exception as e:
        logger.error(f"Bulk update shift assignments error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Aktualisieren: {str(e)}'}, status_code=500)


@router.put('/api/shifts/assignments/{id:int}/toggle-fixed', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def toggle_fixed_assignment(request: Request, id):
    """Toggle the IsFixed flag on an assignment (lock/unlock)"""
    try:
        conn = None
        try:
            db = get_db()
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if assignment exists
            cursor.execute("SELECT Id, IsFixed FROM ShiftAssignments WHERE Id = ?", (id,))
            row = cursor.fetchone()
            
            if not row:
                return JSONResponse(content={'error': 'Schichtzuweisung nicht gefunden'}, status_code=404)
            
            # Toggle fixed status
            new_fixed_status = 0 if row['IsFixed'] else 1
            
            cursor.execute("""
                UPDATE ShiftAssignments 
                SET IsFixed = ?, ModifiedAt = ?, ModifiedBy = ?
                WHERE Id = ?
            """, (
                new_fixed_status,
                datetime.utcnow().isoformat(),
                request.session.get('user_email'),
                id
            ))
            
            conn.commit()
            
            return {
                'success': True,
                'isFixed': bool(new_fixed_status)
            }
            
        finally:
            if conn:
                conn.close()
        
    except Exception as e:
        logger.error(f"Toggle fixed assignment error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Sperren/Entsperren: {str(e)}'}, status_code=500)


# ============================================================================
# EXPORT ROUTES
# ============================================================================

@router.get('/api/shifts/export/csv')

def export_schedule_csv(request: Request):
    """Export schedule to CSV format"""
    start_date_str = request.query_params.get('startDate')
    end_date_str = request.query_params.get('endDate')
    
    if not start_date_str or not end_date_str:
        return JSONResponse(content={'error': 'startDate and endDate are required'}, status_code=400)
    
    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Get assignments
        cursor.execute("""
            SELECT sa.Date, e.Vorname, e.Name, e.Personalnummer, 
                   t.Name as TeamName, st.Code, st.Name as ShiftName
            FROM ShiftAssignments sa
            JOIN Employees e ON sa.EmployeeId = e.Id
            LEFT JOIN Teams t ON e.TeamId = t.Id
            JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
            WHERE sa.Date >= ? AND sa.Date <= ?
            ORDER BY sa.Date, t.Name, e.Name, e.Vorname
        """, (start_date.isoformat(), end_date.isoformat()))
        
        # Build CSV
        import io
        output = io.StringIO()
        output.write("Datum,Team,Mitarbeiter,Personalnummer,Schichttyp,Schichtname\n")
        
        for row in cursor.fetchall():
            team_name = row['TeamName'] or 'Ohne Team'
            output.write(f"{row['Date']},{team_name},{row['Vorname']} {row['Name']},{row['Personalnummer']},{row['Code']},{row['ShiftName']}\n")
        
        conn.close()
        
        # Return as downloadable file
        csv_data = output.getvalue()
        output.close()
        
        return Response(content=csv_data, media_type='text/csv; charset=utf-8', headers={'Content-Disposition': f'attachment; filename=Dienstplan_{start_date_str}_bis_{end_date_str}.csv'})
        
    except Exception as e:
        logger.error(f"CSV export error: {str(e)}")
        return JSONResponse(content={'error': f'Export-Fehler: {str(e)}'}, status_code=500)


def _get_shift_color(shift_code: str) -> tuple:
    """
    Get background color and text color for a shift type.
    Returns (bg_color_hex, text_color_hex)
    Matches the colors from the UI (wwwroot/css/styles.css and database)
    """
    colors_map = {
        'F': ('#4CAF50', '#000000'),   # Früh - green with black text
        'S': ('#FF9800', '#FFFFFF'),   # Spät - orange with white text
        'N': ('#2196F3', '#FFFFFF'),   # Nacht - blue with white text
        'Z': ('#9C27B0', '#FFFFFF'),   # Zwischendienst - purple with white text
        'TD': ('#673AB7', '#FFFFFF'),  # Tagdienst - deep purple with white text
        'BMT': ('#F44336', '#FFFFFF'), # Brandmeldetechniker - red with white text
        'BSB': ('#E91E63', '#FFFFFF'), # Brandschutzbeauftragter - pink with white text
        'U': ('#64748b', '#FFFFFF'),   # Urlaub - gray with white text
        'AU': ('#dc2626', '#FFFFFF'),  # Krank - dark red with white text
        'L': ('#3b82f6', '#FFFFFF'),   # Lehrgang - blue with white text
    }
    return colors_map.get(shift_code, ('#E0E0E0', '#000000'))  # Default gray


def _group_data_by_team_and_employee(conn, start_date: date, end_date: date, view_type: str = 'week'):
    """
    Group shift assignments by team and employee, mirroring the UI's groupByTeamAndEmployee logic.
    Returns: (team_groups, dates, absences_by_employee)
    """
    cursor = conn.cursor()
    
    # Get all employees with their team info and special functions
    cursor.execute("""
        SELECT e.Id, e.Vorname, e.Name, e.Personalnummer, e.TeamId, 
               t.Name as TeamName,
               e.IsBrandmeldetechniker, e.IsBrandschutzbeauftragter
        FROM Employees e
        LEFT JOIN Teams t ON e.TeamId = t.Id
        ORDER BY t.Name NULLS LAST, e.Name, e.Vorname
    """)
    employees = cursor.fetchall()
    
    # Get all shift assignments in the date range
    cursor.execute("""
        SELECT sa.Date, sa.EmployeeId, st.Code, st.Name as ShiftName, st.ColorCode
        FROM ShiftAssignments sa
        JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
        WHERE sa.Date >= ? AND sa.Date <= ?
    """, (start_date.isoformat(), end_date.isoformat()))
    assignments = cursor.fetchall()
    
    # Get all absences in the date range
    cursor.execute("""
        SELECT a.EmployeeId, a.StartDate, a.EndDate, a.Type, a.Notes
        FROM Absences a
        WHERE (a.StartDate <= ? AND a.EndDate >= ?)
           OR (a.StartDate >= ? AND a.StartDate <= ?)
    """, (end_date.isoformat(), start_date.isoformat(), 
          start_date.isoformat(), end_date.isoformat()))
    absences = cursor.fetchall()
    
    # Generate date range
    dates = []
    current = start_date
    while current <= end_date:
        dates.append(current.isoformat())
        current += timedelta(days=1)
    
    # Build absences lookup
    absences_by_employee = {}
    for absence in absences:
        emp_id = absence['EmployeeId']
        if emp_id not in absences_by_employee:
            absences_by_employee[emp_id] = []
        absences_by_employee[emp_id].append(absence)
    
    # Build assignments lookup
    assignments_by_emp_date = {}
    for assignment in assignments:
        key = (assignment['EmployeeId'], assignment['Date'])
        if key not in assignments_by_emp_date:
            assignments_by_emp_date[key] = []
        assignments_by_emp_date[key].append(assignment)
    
    # Group by team
    UNASSIGNED_TEAM_ID = -1
    
    teams = {}
    for emp in employees:
        # Format employee name once for reuse
        emp_name = f"{emp['Vorname']} {emp['Name']}"
        if emp['Personalnummer']:
            emp_name = f"{emp_name} ({emp['Personalnummer']})"
        
        # Add to regular team
        team_id = emp['TeamId'] if emp['TeamId'] else UNASSIGNED_TEAM_ID
        team_name = emp['TeamName'] if emp['TeamName'] else 'Ohne Team'
        
        if team_id not in teams:
            teams[team_id] = {
                'teamId': team_id,
                'teamName': team_name,
                'employees': {}
            }
        
        teams[team_id]['employees'][emp['Id']] = {
            'id': emp['Id'],
            'name': emp_name,
            'shifts': {}
        }
    
    # Populate shifts for each employee
    for team in teams.values():
        for emp_id, emp_data in team['employees'].items():
            for date_str in dates:
                key = (emp_id, date_str)
                shifts = assignments_by_emp_date.get(key, [])
                emp_data['shifts'][date_str] = shifts
    
    # Sort teams (regular -> Ohne Team)
    sorted_teams = []
    for team_id in sorted(teams.keys()):
        if team_id == UNASSIGNED_TEAM_ID:
            continue
        sorted_teams.append(teams[team_id])
    
    if UNASSIGNED_TEAM_ID in teams:
        sorted_teams.append(teams[UNASSIGNED_TEAM_ID])
    
    # Sort employees within each team by name
    for team in sorted_teams:
        team['employees'] = dict(sorted(
            team['employees'].items(),
            key=lambda x: x[1]['name']
        ))
    
    return sorted_teams, dates, absences_by_employee


def _get_absence_for_date(absences: list, date_str: str) -> Optional[dict]:
    """Check if an employee has an absence on a specific date"""
    target_date = date.fromisoformat(date_str)
    for absence in absences:
        start = date.fromisoformat(absence['StartDate'])
        end = date.fromisoformat(absence['EndDate'])
        if start <= target_date <= end:
            return absence
    return None


def _get_absence_code(absence_type: int) -> str:
    """Convert absence type to code (U, AU, L)"""
    # From entities.py: U=1 (Urlaub), AU=2 (Krank), L=3 (Lehrgang/Fortbildung)
    codes = {1: 'U', 2: 'AU', 3: 'L'}
    return codes.get(absence_type, 'U')


@router.get('/api/shifts/export/pdf')

def export_schedule_pdf(request: Request):
    """Export schedule to PDF format matching the UI view structure"""
    start_date_str = request.query_params.get('startDate')
    end_date_str = request.query_params.get('endDate')
    view_type = request.query_params.get('view', 'week')  # week, month, or year
    
    if not start_date_str or not end_date_str:
        return JSONResponse(content={'error': 'startDate and endDate are required'}, status_code=400)
    
    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        
        db = get_db()
        conn = db.get_connection()
        
        # Get grouped data matching UI structure
        team_groups, dates, absences_by_employee = _group_data_by_team_and_employee(conn, start_date, end_date, view_type)
        
        # Create PDF
        import io
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import A4, A3, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        buffer = io.BytesIO()
        
        # Determine appropriate page size based on view type and number of columns
        num_columns = len(dates) + 1  # +1 for employee name column
        
        # Calculate required table width
        employee_col_width = 5*cm
        if view_type == 'year':
            date_col_width = 0.8*cm  # Smaller for year view (365 days)
        elif view_type == 'month' and len(dates) > 28:
            date_col_width = 1.2*cm  # Compressed for month view
        else:
            date_col_width = 1.8*cm  # Normal for week view
        
        required_width = employee_col_width + (len(dates) * date_col_width)
        
        # Standard landscape A4 width for comparison
        landscape_a4_width = landscape(A4)[0]
        
        # Determine page size - use A3 for large tables
        if required_width > landscape_a4_width - 2*cm:
            # Use landscape A3 for larger tables
            pagesize = landscape(A3)
        else:
            pagesize = landscape(A4)
        
        # Set margins to maximize usable space
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=pagesize,
            leftMargin=0.5*cm,
            rightMargin=0.5*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        if view_type == 'week':
            # Get week number
            first_date_obj = datetime.fromisoformat(dates[0])
            week_num = first_date_obj.isocalendar()[1]
            year = first_date_obj.year
            title_text = f"Dienstplan - Woche: KW {week_num} {year}"
        elif view_type == 'month':
            month_name = start_date.strftime('%B %Y')
            title_text = f"Dienstplan - Monat: {month_name}"
        else:  # year
            year = datetime.fromisoformat(dates[0]).year
            title_text = f"Dienstplan - Jahr: {year}"
        
        title = Paragraph(title_text, styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*cm))
        
        # Build table data matching UI structure
        table_data = []
        
        # Header row
        header_row = ['Team / Mitarbeiter']
        for date_str in dates:
            date_obj = datetime.fromisoformat(date_str)
            if view_type == 'year':
                # For year view, show only date number
                header_row.append(date_obj.strftime('%d.%m'))
            else:
                # For week/month view, show day name and date
                day_name = date_obj.strftime('%a')
                day_num = date_obj.strftime('%d.%m')
                header_row.append(f"{day_name}\n{day_num}")
        table_data.append(header_row)
        
        # Data rows - grouped by team
        for team in team_groups:
            # Team header row
            team_row = [team['teamName']] + [''] * len(dates)
            table_data.append(team_row)
            
            # Employee rows
            for emp_id, emp_data in team['employees'].items():
                emp_row = [f"  - {emp_data['name']}"]
                
                for date_str in dates:
                    # Check for absence first
                    abs_list = absences_by_employee.get(emp_id, [])
                    absence = _get_absence_for_date(abs_list, date_str)
                    
                    if absence:
                        absence_code = _get_absence_code(absence['Type'])
                        emp_row.append(absence_code)
                    else:
                        # Get shifts for this date
                        shifts = emp_data['shifts'].get(date_str, [])
                        if shifts:
                            shift_codes = ' '.join([s['Code'] for s in shifts])
                            emp_row.append(shift_codes)
                        else:
                            emp_row.append('-')
                
                table_data.append(emp_row)
        
        conn.close()
        
        # Create table with styling
        # Use the dynamically calculated column widths
        col_widths = [employee_col_width] + [date_col_width] * len(dates)
        
        table = Table(table_data, colWidths=col_widths)
        
        # Apply styling
        # Adjust font sizes based on view type
        if view_type == 'year':
            header_font_size = 6
            data_font_size = 5
        elif view_type == 'month':
            header_font_size = 7
            data_font_size = 6
        else:  # week
            header_font_size = 9
            data_font_size = 8
        
        style_commands = [
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
            # First column (employee names) - left aligned
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTSIZE', (0, 1), (-1, -1), data_font_size),
            ('LEFTPADDING', (0, 1), (0, -1), 3),
            ('RIGHTPADDING', (0, 1), (0, -1), 3),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ]
        
        # Add team header row styling
        row_idx = 1
        for team in team_groups:
            # Team header background with gradient-like color
            style_commands.append(
                ('BACKGROUND', (0, row_idx), (-1, row_idx), HexColor('#2563eb'))
            )
            style_commands.append(
                ('TEXTCOLOR', (0, row_idx), (-1, row_idx), rl_colors.white)
            )
            style_commands.append(
                ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold')
            )
            style_commands.append(
                ('ALIGN', (0, row_idx), (0, row_idx), 'LEFT')
            )
            row_idx += 1 + len(team['employees'])
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        doc.build(elements)
        
        # Return PDF
        buffer.seek(0)
        pdf_bytes = buffer.getvalue()
        return Response(content=pdf_bytes, media_type='application/pdf', headers={'Content-Disposition': f'attachment; filename=Dienstplan_{start_date_str}_bis_{end_date_str}.pdf'})
        
    except Exception as e:
        logger.error(f"PDF export error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JSONResponse(content={'error': f'PDF-Export-Fehler: {str(e)}'}, status_code=500)


@router.get('/api/shifts/export/excel')

def export_schedule_excel(request: Request):
    """Export schedule to Excel format matching the UI view structure"""
    start_date_str = request.query_params.get('startDate')
    end_date_str = request.query_params.get('endDate')
    view_type = request.query_params.get('view', 'week')  # week, month, or year
    
    if not start_date_str or not end_date_str:
        return JSONResponse(content={'error': 'startDate and endDate are required'}, status_code=400)
    
    try:
        # Import Excel library
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return JSONResponse(
                content={'error': 'Excel-Export erfordert openpyxl. Bitte installieren Sie es mit: pip install openpyxl'},
                status_code=501
            )
        
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        
        db = get_db()
        conn = db.get_connection()
        
        # Get grouped data matching UI structure
        team_groups, dates, absences_by_employee = _group_data_by_team_and_employee(conn, start_date, end_date, view_type)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set title based on view type
        if view_type == 'week':
            first_date_obj = datetime.fromisoformat(dates[0])
            week_num = first_date_obj.isocalendar()[1]
            year = first_date_obj.year
            ws.title = f"KW{week_num} {year}"
        elif view_type == 'month':
            ws.title = start_date.strftime('%B %Y')
        else:  # year
            year = datetime.fromisoformat(dates[0]).year
            ws.title = f"Jahr {year}"
        
        # Header row
        header_row = ['Team / Mitarbeiter']
        for date_str in dates:
            date_obj = datetime.fromisoformat(date_str)
            if view_type == 'year':
                header_row.append(date_obj.strftime('%d.%m'))
            else:
                day_name = date_obj.strftime('%a')
                day_num = date_obj.strftime('%d.%m')
                header_row.append(f"{day_name}\n{day_num}")
        ws.append(header_row)
        
        # Style header row
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Data rows - grouped by team
        current_row = 2
        for team in team_groups:
            # Team header row
            team_row = [team['teamName']] + [''] * len(dates)
            ws.append(team_row)
            
            # Style team header
            team_font = Font(bold=True, color="FFFFFF", size=10)
            team_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            team_alignment = Alignment(horizontal="left", vertical="center")
            
            for col_idx in range(1, len(dates) + 2):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = team_font
                cell.fill = team_fill
                cell.alignment = team_alignment
                cell.border = border
            
            current_row += 1
            
            # Employee rows
            for emp_id, emp_data in team['employees'].items():
                emp_row = [f"  - {emp_data['name']}"]
                
                for date_str in dates:
                    # Check for absence first
                    abs_list = absences_by_employee.get(emp_id, [])
                    absence = _get_absence_for_date(abs_list, date_str)
                    
                    if absence:
                        absence_code = _get_absence_code(absence['Type'])
                        emp_row.append(absence_code)
                    else:
                        # Get shifts for this date
                        shifts = emp_data['shifts'].get(date_str, [])
                        if shifts:
                            shift_codes = ' '.join([s['Code'] for s in shifts])
                            emp_row.append(shift_codes)
                        else:
                            emp_row.append('-')
                
                ws.append(emp_row)
                
                # Style employee row
                emp_font = Font(size=9)
                emp_alignment_left = Alignment(horizontal="left", vertical="center")
                emp_alignment_center = Alignment(horizontal="center", vertical="center")
                
                # First cell (employee name) - left aligned
                cell = ws.cell(row=current_row, column=1)
                cell.font = emp_font
                cell.alignment = emp_alignment_left
                cell.border = border
                
                # Shift cells - with color coding
                for col_idx in range(2, len(dates) + 2):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = Font(size=8, bold=True)
                    cell.alignment = emp_alignment_center
                    cell.border = border
                    
                    # Get the shift code to apply color
                    cell_value = str(cell.value) if cell.value else ''
                    if cell_value and cell_value != '-':
                        # Split multiple shifts and use first one for color
                        first_shift = cell_value.split()[0]
                        bg_color, text_color = _get_shift_color(first_shift)
                        # Remove # from hex colors for openpyxl
                        bg_hex = bg_color.replace('#', '')
                        text_hex = text_color.replace('#', '')
                        cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                        cell.font = Font(size=8, bold=True, color=text_hex)
                
                current_row += 1
        
        conn.close()
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30  # Employee names column
        for col_idx in range(2, len(dates) + 2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if view_type == 'year':
                ws.column_dimensions[col_letter].width = 6
            else:
                ws.column_dimensions[col_letter].width = 8
        
        # Save to BytesIO
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return Excel file
        xl_bytes = output.getvalue()
        return Response(content=xl_bytes, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename=Dienstplan_{start_date_str}_bis_{end_date_str}.xlsx'})
        
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JSONResponse(content={'error': f'Excel-Export-Fehler: {str(e)}'}, status_code=500)


# ============================================================================
# SHIFT EXCHANGE ENDPOINTS
# ============================================================================

@router.get('/api/shiftexchanges/available')

def get_available_shift_exchanges(request: Request):
    """Get available shift exchanges"""
    db = get_db()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT se.*, 
               sa.Date, sa.ShiftTypeId,
               st.Code as ShiftCode, st.Name as ShiftName,
               e.Vorname, e.Name, e.TeamId
        FROM ShiftExchanges se
        JOIN ShiftAssignments sa ON se.ShiftAssignmentId = sa.Id
        JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
        JOIN Employees e ON se.OfferingEmployeeId = e.Id
        WHERE se.Status = 'Angeboten'
        ORDER BY sa.Date
    """)
    
    exchanges = []
    for row in cursor.fetchall():
        exchanges.append({
            'id': row['Id'],
            'shiftAssignmentId': row['ShiftAssignmentId'],
            'offeringEmployeeId': row['OfferingEmployeeId'],
            'offeringEmployeeName': f"{row['Vorname']} {row['Name']}",
            'teamId': row['TeamId'],
            'date': row['Date'],
            'shiftCode': row['ShiftCode'],
            'shiftName': row['ShiftName'],
            'status': row['Status'],
            'offeringReason': row['OfferingReason'],
            'createdAt': row['CreatedAt']
        })
    
    conn.close()
    return exchanges


@router.get('/api/shiftexchanges/pending', dependencies=[Depends(require_role('Admin'))])

def get_pending_shift_exchanges(request: Request):
    """Get pending shift exchanges (Admin only)"""
    db = get_db()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT se.*, 
               sa.Date, sa.ShiftTypeId,
               st.Code as ShiftCode, st.Name as ShiftName,
               e1.Vorname as OfferingVorname, e1.Name as OfferingName,
               e2.Vorname as RequestingVorname, e2.Name as RequestingName
        FROM ShiftExchanges se
        JOIN ShiftAssignments sa ON se.ShiftAssignmentId = sa.Id
        JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
        JOIN Employees e1 ON se.OfferingEmployeeId = e1.Id
        LEFT JOIN Employees e2 ON se.RequestingEmployeeId = e2.Id
        WHERE se.Status = 'Angefragt'
        ORDER BY sa.Date
    """)
    
    exchanges = []
    for row in cursor.fetchall():
        exchanges.append({
            'id': row['Id'],
            'shiftAssignmentId': row['ShiftAssignmentId'],
            'offeringEmployeeId': row['OfferingEmployeeId'],
            'offeringEmployeeName': f"{row['OfferingVorname']} {row['OfferingName']}",
            'requestingEmployeeId': row['RequestingEmployeeId'],
            'requestingEmployeeName': f"{row['RequestingVorname']} {row['RequestingName']}" if row['RequestingEmployeeId'] else None,
            'date': row['Date'],
            'shiftCode': row['ShiftCode'],
            'shiftName': row['ShiftName'],
            'status': row['Status'],
            'offeringReason': row['OfferingReason'],
            'createdAt': row['CreatedAt']
        })
    
    conn.close()
    return exchanges


@router.post('/api/shiftexchanges', dependencies=[Depends(require_auth), Depends(check_csrf)])

def create_shift_exchange(request: Request):
    """Create new shift exchange offer"""
    try:
        
        if not data.get('shiftAssignmentId') or not data.get('offeringEmployeeId'):
            return JSONResponse(content={'error': 'ShiftAssignmentId und OfferingEmployeeId sind erforderlich'}, status_code=400)
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO ShiftExchanges 
            (ShiftAssignmentId, OfferingEmployeeId, Status, OfferingReason, CreatedAt)
            VALUES (?, ?, 'Angeboten', ?, ?)
        """, (
            data.get('shiftAssignmentId'),
            data.get('offeringEmployeeId'),
            data.get('offeringReason'),
            datetime.utcnow().isoformat()
        ))
        
        exchange_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return JSONResponse(content={'success': True, 'id': exchange_id}, status_code=201)
        
    except Exception as e:
        logger.error(f"Create shift exchange error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Erstellen: {str(e)}'}, status_code=500)


@router.post('/api/shiftexchanges/{id:int}/request', dependencies=[Depends(require_auth), Depends(check_csrf)])

def request_shift_exchange(request: Request, id):
    """Request a shift exchange"""
    try:
        requesting_employee_id = data.get('requestingEmployeeId')
        
        if not requesting_employee_id:
            return JSONResponse(content={'error': 'RequestingEmployeeId ist erforderlich'}, status_code=400)
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE ShiftExchanges 
            SET RequestingEmployeeId = ?, Status = 'Angefragt'
            WHERE Id = ? AND Status = 'Angeboten'
        """, (requesting_employee_id, id))
        
        if cursor.rowcount == 0:
            conn.close()
            return JSONResponse(content={'error': 'Tauschangebot nicht verfügbar'}, status_code=404)
        
        conn.commit()
        conn.close()
        
        return {'success': True}
        
    except Exception as e:
        logger.error(f"Request shift exchange error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Anfragen: {str(e)}'}, status_code=500)


@router.put('/api/shiftexchanges/{id:int}/process', dependencies=[Depends(require_role('Admin')), Depends(check_csrf)])

def process_shift_exchange(request: Request, id):
    """Process shift exchange (approve/reject)"""
    try:
        status = data.get('status')
        notes = data.get('notes')
        
        if status not in ['Genehmigt', 'Abgelehnt']:
            return JSONResponse(content={'error': 'Status muss Genehmigt oder Abgelehnt sein'}, status_code=400)
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Get exchange details
        cursor.execute("""
            SELECT ShiftAssignmentId, OfferingEmployeeId, RequestingEmployeeId
            FROM ShiftExchanges
            WHERE Id = ? AND Status = 'Angefragt'
        """, (id,))
        
        row = cursor.fetchone()
        if not row:
            conn.close()
            return JSONResponse(content={'error': 'Tauschangebot nicht gefunden oder bereits bearbeitet'}, status_code=404)
        
        shift_assignment_id = row['ShiftAssignmentId']
        requesting_employee_id = row['RequestingEmployeeId']
        
        # Update exchange status
        cursor.execute("""
            UPDATE ShiftExchanges 
            SET Status = ?, DisponentNotes = ?, ProcessedAt = ?, ProcessedBy = ?
            WHERE Id = ?
        """, (
            status,
            notes,
            datetime.utcnow().isoformat(),
            request.session.get('user_email'),
            id
        ))
        
        # If approved, update the shift assignment
        if status == 'Genehmigt' and requesting_employee_id:
            cursor.execute("""
                UPDATE ShiftAssignments
                SET EmployeeId = ?, ModifiedAt = ?, ModifiedBy = ?
                WHERE Id = ?
            """, (
                requesting_employee_id,
                datetime.utcnow().isoformat(),
                request.session.get('user_email'),
                shift_assignment_id
            ))
        
        conn.commit()
        conn.close()
        
        return {'success': True}
        
    except Exception as e:
        logger.error(f"Process shift exchange error: {str(e)}")
        return JSONResponse(content={'error': f'Fehler beim Bearbeiten: {str(e)}'}, status_code=500)
