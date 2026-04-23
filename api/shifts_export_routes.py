"""Shift schedule export (CSV, PDF, Excel) API routes."""

import io
import logging
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Request
from fastapi.responses import JSONResponse, Response

from .shared import get_db

logger = logging.getLogger(__name__)
router = APIRouter()

@router.get('/api/shifts/export/csv')

def export_schedule_csv(request: Request):
    """Export schedule to CSV format"""
    start_date_str = request.query_params.get('startDate')
    end_date_str = request.query_params.get('endDate')
    
    if not start_date_str or not end_date_str:
        return JSONResponse(content={'error': 'startDate and endDate are required'}, status_code=400)
    
    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        
        db = get_db()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Get assignments
        cursor.execute("""
            SELECT sa.Date, e.Vorname, e.Name, e.Personalnummer, 
                   t.Name as TeamName, st.Code, st.Name as ShiftName
            FROM ShiftAssignments sa
            JOIN Employees e ON sa.EmployeeId = e.Id
            LEFT JOIN Teams t ON e.TeamId = t.Id
            JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
            WHERE sa.Date >= ? AND sa.Date <= ?
            ORDER BY sa.Date, t.Name, e.Name, e.Vorname
        """, (start_date.isoformat(), end_date.isoformat()))
        
        # Build CSV
        import io
        output = io.StringIO()
        output.write("Datum,Team,Mitarbeiter,Personalnummer,Schichttyp,Schichtname\n")
        
        for row in cursor.fetchall():
            team_name = row['TeamName'] or 'Ohne Team'
            output.write(f"{row['Date']},{team_name},{row['Vorname']} {row['Name']},{row['Personalnummer']},{row['Code']},{row['ShiftName']}\n")
        
        conn.close()
        
        # Return as downloadable file
        csv_data = output.getvalue()
        output.close()
        
        return Response(content=csv_data, media_type='text/csv; charset=utf-8', headers={'Content-Disposition': f'attachment; filename=Dienstplan_{start_date_str}_bis_{end_date_str}.csv'})
        
    except Exception as e:
        logger.error(f"CSV export error: {str(e)}")
        return JSONResponse(content={'error': f'Export-Fehler: {str(e)}'}, status_code=500)


def _get_shift_color(shift_code: str) -> tuple:
    """
    Get background color and text color for a shift type.
    Returns (bg_color_hex, text_color_hex)
    Matches the colors from the UI (wwwroot/css/styles.css and database)
    """
    colors_map = {
        'F': ('#4CAF50', '#000000'),   # Früh - green with black text
        'S': ('#FF9800', '#FFFFFF'),   # Spät - orange with white text
        'N': ('#2196F3', '#FFFFFF'),   # Nacht - blue with white text
        'Z': ('#9C27B0', '#FFFFFF'),   # Zwischendienst - purple with white text
        'TD': ('#673AB7', '#FFFFFF'),  # Tagdienst - deep purple with white text
        'BMT': ('#F44336', '#FFFFFF'), # Brandmeldetechniker - red with white text
        'BSB': ('#E91E63', '#FFFFFF'), # Brandschutzbeauftragter - pink with white text
        'U': ('#64748b', '#FFFFFF'),   # Urlaub - gray with white text
        'AU': ('#dc2626', '#FFFFFF'),  # Krank - dark red with white text
        'L': ('#3b82f6', '#FFFFFF'),   # Lehrgang - blue with white text
    }
    return colors_map.get(shift_code, ('#E0E0E0', '#000000'))  # Default gray


def _group_data_by_team_and_employee(conn, start_date: date, end_date: date, view_type: str = 'week'):
    """
    Group shift assignments by team and employee, mirroring the UI's groupByTeamAndEmployee logic.
    Returns: (team_groups, dates, absences_by_employee)
    """
    cursor = conn.cursor()
    
    # Get all employees with their team info and special functions
    cursor.execute("""
        SELECT e.Id, e.Vorname, e.Name, e.Personalnummer, e.TeamId, 
               t.Name as TeamName,
               e.IsBrandmeldetechniker, e.IsBrandschutzbeauftragter
        FROM Employees e
        LEFT JOIN Teams t ON e.TeamId = t.Id
        ORDER BY t.Name NULLS LAST, e.Name, e.Vorname
    """)
    employees = cursor.fetchall()
    
    # Get all shift assignments in the date range
    cursor.execute("""
        SELECT sa.Date, sa.EmployeeId, st.Code, st.Name as ShiftName, st.ColorCode
        FROM ShiftAssignments sa
        JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
        WHERE sa.Date >= ? AND sa.Date <= ?
    """, (start_date.isoformat(), end_date.isoformat()))
    assignments = cursor.fetchall()
    
    # Get all absences in the date range
    cursor.execute("""
        SELECT a.EmployeeId, a.StartDate, a.EndDate, a.Type, a.Notes
        FROM Absences a
        WHERE (a.StartDate <= ? AND a.EndDate >= ?)
           OR (a.StartDate >= ? AND a.StartDate <= ?)
    """, (end_date.isoformat(), start_date.isoformat(), 
          start_date.isoformat(), end_date.isoformat()))
    absences = cursor.fetchall()
    
    # Generate date range
    dates = []
    current = start_date
    while current <= end_date:
        dates.append(current.isoformat())
        current += timedelta(days=1)
    
    # Build absences lookup
    absences_by_employee = {}
    for absence in absences:
        emp_id = absence['EmployeeId']
        if emp_id not in absences_by_employee:
            absences_by_employee[emp_id] = []
        absences_by_employee[emp_id].append(absence)
    
    # Build assignments lookup
    assignments_by_emp_date = {}
    for assignment in assignments:
        key = (assignment['EmployeeId'], assignment['Date'])
        if key not in assignments_by_emp_date:
            assignments_by_emp_date[key] = []
        assignments_by_emp_date[key].append(assignment)
    
    # Group by team
    UNASSIGNED_TEAM_ID = -1
    
    teams = {}
    for emp in employees:
        # Format employee name once for reuse
        emp_name = f"{emp['Vorname']} {emp['Name']}"
        if emp['Personalnummer']:
            emp_name = f"{emp_name} ({emp['Personalnummer']})"
        
        # Add to regular team
        team_id = emp['TeamId'] if emp['TeamId'] else UNASSIGNED_TEAM_ID
        team_name = emp['TeamName'] if emp['TeamName'] else 'Ohne Team'
        
        if team_id not in teams:
            teams[team_id] = {
                'teamId': team_id,
                'teamName': team_name,
                'employees': {}
            }
        
        teams[team_id]['employees'][emp['Id']] = {
            'id': emp['Id'],
            'name': emp_name,
            'shifts': {}
        }
    
    # Populate shifts for each employee
    for team in teams.values():
        for emp_id, emp_data in team['employees'].items():
            for date_str in dates:
                key = (emp_id, date_str)
                shifts = assignments_by_emp_date.get(key, [])
                emp_data['shifts'][date_str] = shifts
    
    # Sort teams (regular -> Ohne Team)
    sorted_teams = []
    for team_id in sorted(teams.keys()):
        if team_id == UNASSIGNED_TEAM_ID:
            continue
        sorted_teams.append(teams[team_id])
    
    if UNASSIGNED_TEAM_ID in teams:
        sorted_teams.append(teams[UNASSIGNED_TEAM_ID])
    
    # Sort employees within each team by name
    for team in sorted_teams:
        team['employees'] = dict(sorted(
            team['employees'].items(),
            key=lambda x: x[1]['name']
        ))
    
    return sorted_teams, dates, absences_by_employee


def _get_absence_for_date(absences: list, date_str: str) -> Optional[dict]:
    """Check if an employee has an absence on a specific date"""
    target_date = date.fromisoformat(date_str)
    for absence in absences:
        start = date.fromisoformat(absence['StartDate'])
        end = date.fromisoformat(absence['EndDate'])
        if start <= target_date <= end:
            return absence
    return None


def _get_absence_code(absence_type: int) -> str:
    """Convert absence type to code (U, AU, L)"""
    # From entities.py: U=1 (Urlaub), AU=2 (Krank), L=3 (Lehrgang/Fortbildung)
    codes = {1: 'U', 2: 'AU', 3: 'L'}
    return codes.get(absence_type, 'U')


@router.get('/api/shifts/export/pdf')

def export_schedule_pdf(request: Request):
    """Export schedule to PDF format matching the UI view structure"""
    start_date_str = request.query_params.get('startDate')
    end_date_str = request.query_params.get('endDate')
    view_type = request.query_params.get('view', 'week')  # week, month, or year
    
    if not start_date_str or not end_date_str:
        return JSONResponse(content={'error': 'startDate and endDate are required'}, status_code=400)
    
    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        
        db = get_db()
        conn = db.get_connection()
        
        # Get grouped data matching UI structure
        team_groups, dates, absences_by_employee = _group_data_by_team_and_employee(conn, start_date, end_date, view_type)
        
        # Create PDF
        import io
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import A4, A3, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        buffer = io.BytesIO()
        
        # Determine appropriate page size based on view type and number of columns
        num_columns = len(dates) + 1  # +1 for employee name column
        
        # Calculate required table width
        employee_col_width = 5*cm
        if view_type == 'year':
            date_col_width = 0.8*cm  # Smaller for year view (365 days)
        elif view_type == 'month' and len(dates) > 28:
            date_col_width = 1.2*cm  # Compressed for month view
        else:
            date_col_width = 1.8*cm  # Normal for week view
        
        required_width = employee_col_width + (len(dates) * date_col_width)
        
        # Standard landscape A4 width for comparison
        landscape_a4_width = landscape(A4)[0]
        
        # Determine page size - use A3 for large tables
        if required_width > landscape_a4_width - 2*cm:
            # Use landscape A3 for larger tables
            pagesize = landscape(A3)
        else:
            pagesize = landscape(A4)
        
        # Set margins to maximize usable space
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=pagesize,
            leftMargin=0.5*cm,
            rightMargin=0.5*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        if view_type == 'week':
            # Get week number
            first_date_obj = datetime.fromisoformat(dates[0])
            week_num = first_date_obj.isocalendar()[1]
            year = first_date_obj.year
            title_text = f"Dienstplan - Woche: KW {week_num} {year}"
        elif view_type == 'month':
            month_name = start_date.strftime('%B %Y')
            title_text = f"Dienstplan - Monat: {month_name}"
        else:  # year
            year = datetime.fromisoformat(dates[0]).year
            title_text = f"Dienstplan - Jahr: {year}"
        
        title = Paragraph(title_text, styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*cm))
        
        # Build table data matching UI structure
        table_data = []
        
        # Header row
        header_row = ['Team / Mitarbeiter']
        for date_str in dates:
            date_obj = datetime.fromisoformat(date_str)
            if view_type == 'year':
                # For year view, show only date number
                header_row.append(date_obj.strftime('%d.%m'))
            else:
                # For week/month view, show day name and date
                day_name = date_obj.strftime('%a')
                day_num = date_obj.strftime('%d.%m')
                header_row.append(f"{day_name}\n{day_num}")
        table_data.append(header_row)
        
        # Data rows - grouped by team
        for team in team_groups:
            # Team header row
            team_row = [team['teamName']] + [''] * len(dates)
            table_data.append(team_row)
            
            # Employee rows
            for emp_id, emp_data in team['employees'].items():
                emp_row = [f"  - {emp_data['name']}"]
                
                for date_str in dates:
                    # Check for absence first
                    abs_list = absences_by_employee.get(emp_id, [])
                    absence = _get_absence_for_date(abs_list, date_str)
                    
                    if absence:
                        absence_code = _get_absence_code(absence['Type'])
                        emp_row.append(absence_code)
                    else:
                        # Get shifts for this date
                        shifts = emp_data['shifts'].get(date_str, [])
                        if shifts:
                            shift_codes = ' '.join([s['Code'] for s in shifts])
                            emp_row.append(shift_codes)
                        else:
                            emp_row.append('-')
                
                table_data.append(emp_row)
        
        conn.close()
        
        # Create table with styling
        # Use the dynamically calculated column widths
        col_widths = [employee_col_width] + [date_col_width] * len(dates)
        
        table = Table(table_data, colWidths=col_widths)
        
        # Apply styling
        # Adjust font sizes based on view type
        if view_type == 'year':
            header_font_size = 6
            data_font_size = 5
        elif view_type == 'month':
            header_font_size = 7
            data_font_size = 6
        else:  # week
            header_font_size = 9
            data_font_size = 8
        
        style_commands = [
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
            # First column (employee names) - left aligned
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTSIZE', (0, 1), (-1, -1), data_font_size),
            ('LEFTPADDING', (0, 1), (0, -1), 3),
            ('RIGHTPADDING', (0, 1), (0, -1), 3),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ]
        
        # Add team header row styling
        row_idx = 1
        for team in team_groups:
            # Team header background with gradient-like color
            style_commands.append(
                ('BACKGROUND', (0, row_idx), (-1, row_idx), HexColor('#2563eb'))
            )
            style_commands.append(
                ('TEXTCOLOR', (0, row_idx), (-1, row_idx), rl_colors.white)
            )
            style_commands.append(
                ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold')
            )
            style_commands.append(
                ('ALIGN', (0, row_idx), (0, row_idx), 'LEFT')
            )
            row_idx += 1 + len(team['employees'])
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        doc.build(elements)
        
        # Return PDF
        buffer.seek(0)
        pdf_bytes = buffer.getvalue()
        return Response(content=pdf_bytes, media_type='application/pdf', headers={'Content-Disposition': f'attachment; filename=Dienstplan_{start_date_str}_bis_{end_date_str}.pdf'})
        
    except Exception as e:
        logger.error(f"PDF export error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JSONResponse(content={'error': f'PDF-Export-Fehler: {str(e)}'}, status_code=500)


@router.get('/api/shifts/export/excel')

def export_schedule_excel(request: Request):
    """Export schedule to Excel format matching the UI view structure"""
    start_date_str = request.query_params.get('startDate')
    end_date_str = request.query_params.get('endDate')
    view_type = request.query_params.get('view', 'week')  # week, month, or year
    
    if not start_date_str or not end_date_str:
        return JSONResponse(content={'error': 'startDate and endDate are required'}, status_code=400)
    
    try:
        # Import Excel library
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return JSONResponse(
                content={'error': 'Excel-Export erfordert openpyxl. Bitte installieren Sie es mit: pip install openpyxl'},
                status_code=501
            )
        
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        
        db = get_db()
        conn = db.get_connection()
        
        # Get grouped data matching UI structure
        team_groups, dates, absences_by_employee = _group_data_by_team_and_employee(conn, start_date, end_date, view_type)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set title based on view type
        if view_type == 'week':
            first_date_obj = datetime.fromisoformat(dates[0])
            week_num = first_date_obj.isocalendar()[1]
            year = first_date_obj.year
            ws.title = f"KW{week_num} {year}"
        elif view_type == 'month':
            ws.title = start_date.strftime('%B %Y')
        else:  # year
            year = datetime.fromisoformat(dates[0]).year
            ws.title = f"Jahr {year}"
        
        # Header row
        header_row = ['Team / Mitarbeiter']
        for date_str in dates:
            date_obj = datetime.fromisoformat(date_str)
            if view_type == 'year':
                header_row.append(date_obj.strftime('%d.%m'))
            else:
                day_name = date_obj.strftime('%a')
                day_num = date_obj.strftime('%d.%m')
                header_row.append(f"{day_name}\n{day_num}")
        ws.append(header_row)
        
        # Style header row
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Data rows - grouped by team
        current_row = 2
        for team in team_groups:
            # Team header row
            team_row = [team['teamName']] + [''] * len(dates)
            ws.append(team_row)
            
            # Style team header
            team_font = Font(bold=True, color="FFFFFF", size=10)
            team_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            team_alignment = Alignment(horizontal="left", vertical="center")
            
            for col_idx in range(1, len(dates) + 2):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = team_font
                cell.fill = team_fill
                cell.alignment = team_alignment
                cell.border = border
            
            current_row += 1
            
            # Employee rows
            for emp_id, emp_data in team['employees'].items():
                emp_row = [f"  - {emp_data['name']}"]
                
                for date_str in dates:
                    # Check for absence first
                    abs_list = absences_by_employee.get(emp_id, [])
                    absence = _get_absence_for_date(abs_list, date_str)
                    
                    if absence:
                        absence_code = _get_absence_code(absence['Type'])
                        emp_row.append(absence_code)
                    else:
                        # Get shifts for this date
                        shifts = emp_data['shifts'].get(date_str, [])
                        if shifts:
                            shift_codes = ' '.join([s['Code'] for s in shifts])
                            emp_row.append(shift_codes)
                        else:
                            emp_row.append('-')
                
                ws.append(emp_row)
                
                # Style employee row
                emp_font = Font(size=9)
                emp_alignment_left = Alignment(horizontal="left", vertical="center")
                emp_alignment_center = Alignment(horizontal="center", vertical="center")
                
                # First cell (employee name) - left aligned
                cell = ws.cell(row=current_row, column=1)
                cell.font = emp_font
                cell.alignment = emp_alignment_left
                cell.border = border
                
                # Shift cells - with color coding
                for col_idx in range(2, len(dates) + 2):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = Font(size=8, bold=True)
                    cell.alignment = emp_alignment_center
                    cell.border = border
                    
                    # Get the shift code to apply color
                    cell_value = str(cell.value) if cell.value else ''
                    if cell_value and cell_value != '-':
                        # Split multiple shifts and use first one for color
                        first_shift = cell_value.split()[0]
                        bg_color, text_color = _get_shift_color(first_shift)
                        # Remove # from hex colors for openpyxl
                        bg_hex = bg_color.replace('#', '')
                        text_hex = text_color.replace('#', '')
                        cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                        cell.font = Font(size=8, bold=True, color=text_hex)
                
                current_row += 1
        
        conn.close()
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30  # Employee names column
        for col_idx in range(2, len(dates) + 2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if view_type == 'year':
                ws.column_dimensions[col_letter].width = 6
            else:
                ws.column_dimensions[col_letter].width = 8
        
        # Save to BytesIO
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return Excel file
        xl_bytes = output.getvalue()
        return Response(content=xl_bytes, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename=Dienstplan_{start_date_str}_bis_{end_date_str}.xlsx'})
        
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JSONResponse(content={'error': f'Excel-Export-Fehler: {str(e)}'}, status_code=500)

