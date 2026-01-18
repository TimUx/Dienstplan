"""
Flask Web API for shift planning system.
Provides REST API endpoints compatible with the existing .NET Web UI.
"""

from flask import Flask, jsonify, request, send_file, session
from flask_cors import CORS
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict
import sqlite3
import json
import hashlib
import secrets
import sys
from functools import wraps

# PDF export dependencies
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

from data_loader import load_from_database, get_existing_assignments
from model import create_shift_planning_model
from solver import solve_shift_planning, get_infeasibility_diagnostics
from entities import Employee, Team, Absence, AbsenceType, ShiftAssignment, VacationPeriod
from notification_manager import (
    process_absence_for_notifications,
    get_unread_notifications,
    mark_notification_as_read,
    get_notification_count
)
from springer_replacement import process_absence_with_springer_assignment


def get_row_value(row: sqlite3.Row, key: str, default):
    """
    Helper to safely get value from sqlite3.Row with default.
    
    sqlite3.Row objects don't have a .get() method like dictionaries.
    This helper provides similar functionality with proper error handling.
    
    Args:
        row: sqlite3.Row object
        key: Column name
        default: Default value if key doesn't exist or value is None
        
    Returns:
        Value from row or default
    """
    try:
        val = row[key]
        return val if val is not None else default
    except (KeyError, IndexError):
        return default


class Database:
    """Database connection helper"""
    
    def __init__(self, db_path: str):
        self.db_path = db_path
    
    def get_connection(self):
        """Get database connection"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn


def hash_password(password: str) -> str:
    """
    Hash password using SHA256.
    
    Note: This is a simple implementation for development/migration from .NET.
    For production, consider using bcrypt, scrypt, or Argon2 for better security.
    
    TODO: Upgrade to bcrypt/Argon2 with proper salting for production use.
    SHA256 alone is vulnerable to rainbow table attacks without salting.
    
    Recommended migration path:
    1. Install bcrypt: pip install bcrypt
    2. Replace with: bcrypt.hashpw(password.encode(), bcrypt.gensalt())
    3. Update verify_password to use bcrypt.checkpw()
    """
    return hashlib.sha256(password.encode()).hexdigest()


def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against hash"""
    return hash_password(password) == password_hash


def get_employee_by_email(db, email: str) -> Optional[Dict]:
    """Get employee by email (employees now include authentication data)"""
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT e.*, GROUP_CONCAT(r.Name) as roles
        FROM Employees e
        LEFT JOIN AspNetUserRoles ur ON CAST(e.Id AS TEXT) = ur.UserId
        LEFT JOIN AspNetRoles r ON ur.RoleId = r.Id
        WHERE e.Email = ?
        GROUP BY e.Id
    """, (email,))
    
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return None
    
    full_name = f"{row['Vorname']} {row['Name']}"
    
    return {
        'id': row['Id'],
        'email': row['Email'],
        'passwordHash': row['PasswordHash'],
        'fullName': full_name,
        'vorname': row['Vorname'],
        'name': row['Name'],
        'personalnummer': row['Personalnummer'],
        'lockoutEnd': row['LockoutEnd'],
        'accessFailedCount': row['AccessFailedCount'],
        'roles': row['roles'].split(',') if row['roles'] else []
    }


def require_auth(f):
    """Decorator to require authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function


def require_role(*required_roles):
    """Decorator to require specific role(s)"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({'error': 'Authentication required'}), 401
            
            user_roles = session.get('user_roles', [])
            if not any(role in user_roles for role in required_roles):
                return jsonify({'error': 'Insufficient permissions'}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def log_audit(conn, entity_name: str, entity_id: str, action: str, changes: Optional[str] = None, 
              user_id: Optional[str] = None, user_name: Optional[str] = None):
    """
    Log an audit entry to the AuditLogs table.
    
    Args:
        conn: Database connection (must be already opened)
        entity_name: Name of the entity (e.g., 'Employee', 'Team', 'ShiftAssignment', 'Absence')
        entity_id: ID of the entity being modified
        action: Action performed (e.g., 'Create', 'Update', 'Delete')
        changes: Optional JSON string with details of changes
        user_id: Optional user ID (will try to get from session if not provided)
        user_name: Optional user name (will try to get from session if not provided)
    
    Note: Audit logging failures are logged but do not prevent the main operation from succeeding.
    """
    try:
        cursor = conn.cursor()
        
        # Get user info from session if not provided
        if user_id is None:
            user_id = session.get('user_id')
        if user_name is None:
            user_name = session.get('user_email')
        
        cursor.execute("""
            INSERT INTO AuditLogs (Timestamp, UserId, UserName, EntityName, EntityId, Action, Changes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            datetime.utcnow().isoformat(),
            user_id,
            user_name,
            entity_name,
            str(entity_id),
            action,
            changes
        ))
    except Exception as e:
        # Log the audit failure but don't raise - we don't want audit logging to break business operations
        print(f"Warning: Failed to log audit entry: {e}", file=sys.stderr)


def get_virtual_team_employee_count(cursor, team_id: int) -> int:
    """
    Get the actual employee count for a virtual team based on employee qualifications.
    
    Args:
        cursor: Database cursor
        team_id: Team ID
        
    Returns:
        Number of employees with the qualifications for this virtual team
    """
    # For virtual teams, return 0 (no special counting logic needed now)
    return 0


def validate_monthly_date_range(start_date: date, end_date: date) -> tuple[bool, str]:
    """
    Validate that the date range covers exactly one complete month.
    
    Args:
        start_date: Start date of the range
        end_date: End date of the range
    
    Returns:
        Tuple of (is_valid, error_message). If valid, error_message is empty string.
    """
    # Validate that the date range is within a single month
    if start_date.year != end_date.year or start_date.month != end_date.month:
        return False, 'Shift planning is only allowed for a single month. Year-based planning has been removed.'
    
    # Validate that the date range covers the entire month
    # First day of month
    first_day = date(start_date.year, start_date.month, 1)
    # Last day of month
    if start_date.month == 12:
        last_day = date(start_date.year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(start_date.year, start_date.month + 1, 1) - timedelta(days=1)
    
    if start_date != first_day or end_date != last_day:
        return False, f'Planning must cover the entire month. Expected: {first_day.isoformat()} to {last_day.isoformat()}'
    
    return True, ''


def create_app(db_path: str = "dienstplan.db") -> Flask:
    """
    Create and configure Flask application.
    
    Args:
        db_path: Path to SQLite database
        
    Returns:
        Configured Flask app
    """
    app = Flask(__name__, static_folder='wwwroot', static_url_path='')
    
    # Configure session
    # Use a consistent secret key (in production, load from environment variable)
    import os
    app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY', 'dev-secret-key-change-in-production')
    app.config['SESSION_COOKIE_NAME'] = 'dienstplan_session'
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    
    CORS(app, supports_credentials=True)  # Enable CORS with credentials
    
    db = Database(db_path)
    
    # ============================================================================
    # AUTHENTICATION ENDPOINTS
    # ============================================================================
    
    @app.route('/api/auth/login', methods=['POST'])
    def login():
        """Authenticate employee and create session"""
        try:
            data = request.get_json()
            email = data.get('email')
            password = data.get('password')
            remember_me = data.get('rememberMe', False)
            
            if not email or not password:
                return jsonify({'error': 'Email und Passwort sind erforderlich'}), 400
            
            # Get employee from database (employees now have auth data)
            employee = get_employee_by_email(db, email)
            
            if not employee:
                return jsonify({'error': 'Ungültige Anmeldedaten'}), 401
            
            # Check if password is set
            if not employee.get('passwordHash'):
                return jsonify({'error': 'Kein Passwort gesetzt. Bitte Administrator kontaktieren.'}), 401
            
            # Check if account is locked
            if employee['lockoutEnd']:
                lockout_end = datetime.fromisoformat(employee['lockoutEnd'])
                if lockout_end > datetime.utcnow():
                    return jsonify({'error': 'Konto ist gesperrt'}), 403
            
            # Verify password
            if not verify_password(password, employee['passwordHash']):
                # Increment failed attempts
                conn = db.get_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE Employees 
                    SET AccessFailedCount = AccessFailedCount + 1
                    WHERE Id = ?
                """, (employee['id'],))
                conn.commit()
                conn.close()
                
                return jsonify({'error': 'Ungültige Anmeldedaten'}), 401
            
            # Reset failed attempts on successful login
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE Employees 
                SET AccessFailedCount = 0
                WHERE Id = ?
            """, (employee['id'],))
            conn.commit()
            conn.close()
            
            # Create session
            session['user_id'] = employee['id']
            session['user_email'] = employee['email']
            session['user_fullname'] = employee['fullName']
            session['user_roles'] = employee['roles']
            
            if remember_me:
                session.permanent = True
            
            return jsonify({
                'success': True,
                'user': {
                    'email': employee['email'],
                    'fullName': employee['fullName'],
                    'roles': employee['roles']
                }
            })
            
        except Exception as e:
            app.logger.error(f"Login error: {str(e)}")
            return jsonify({'error': 'Anmeldefehler aufgetreten'}), 500
    
    @app.route('/api/auth/logout', methods=['POST'])
    def logout():
        """Logout user and clear session"""
        session.clear()
        return jsonify({'success': True})
    
    @app.route('/api/auth/current-user', methods=['GET'])
    def get_current_user():
        """Get currently authenticated user"""
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        
        return jsonify({
            'email': session.get('user_email'),
            'fullName': session.get('user_fullname'),
            'roles': session.get('user_roles', [])
        })
    
    # ============================================================================
    # EMPLOYEE/USER MANAGEMENT ENDPOINTS (Unified)
    # Note: Employees now include authentication data - no separate users table
    # ============================================================================
    
    @app.route('/api/users', methods=['GET'])
    @require_role('Admin')
    def get_all_users():
        """Get all employees with authentication/roles (Admin only)"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT e.Id, e.Vorname, e.Name, e.Personalnummer, e.Email, e.NormalizedEmail,
                   e.Funktion, e.TeamId, e.LockoutEnd, e.AccessFailedCount, e.IsActive,
                   e.IsFerienjobber, e.IsBrandmeldetechniker, e.IsBrandschutzbeauftragter,
                   e.IsTdQualified, e.IsTeamLeader,
                   t.Name as TeamName,
                   GROUP_CONCAT(r.Name) as roles
            FROM Employees e
            LEFT JOIN Teams t ON e.TeamId = t.Id
            LEFT JOIN AspNetUserRoles ur ON CAST(e.Id AS TEXT) = ur.UserId
            LEFT JOIN AspNetRoles r ON ur.RoleId = r.Id
            GROUP BY e.Id
            ORDER BY e.Name, e.Vorname
        """)
        
        users = []
        for row in cursor.fetchall():
            users.append({
                'id': row['Id'],
                'email': row['Email'],
                'vorname': row['Vorname'],
                'name': row['Name'],
                'fullName': f"{row['Vorname']} {row['Name']}",
                'personalnummer': row['Personalnummer'],
                'funktion': row['Funktion'],
                'teamId': row['TeamId'],
                'teamName': row['TeamName'],
                'lockoutEnd': row['LockoutEnd'],
                'accessFailedCount': row['AccessFailedCount'],
                'isActive': bool(row['IsActive']),
                'hasPassword': bool(row['Email']),  # Has auth if email is set
                'roles': row['roles'].split(',') if row['roles'] else [],
                # Employee data
                'isFerienjobber': bool(row['IsFerienjobber']),
                'isBrandmeldetechniker': bool(row['IsBrandmeldetechniker']),
                'isBrandschutzbeauftragter': bool(row['IsBrandschutzbeauftragter']),
                'isTdQualified': bool(row['IsTdQualified']),
                'isTeamLeader': bool(row['IsTeamLeader'])
            })
        
        conn.close()
        return jsonify(users)
    
    @app.route('/api/users/<int:user_id>', methods=['GET'])
    @require_role('Admin')
    def get_user(user_id):
        """Get single employee/user by ID (Admin only)"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT e.*, t.Name as TeamName,
                   GROUP_CONCAT(r.Name) as roles
            FROM Employees e
            LEFT JOIN Teams t ON e.TeamId = t.Id
            LEFT JOIN AspNetUserRoles ur ON CAST(e.Id AS TEXT) = ur.UserId
            LEFT JOIN AspNetRoles r ON ur.RoleId = r.Id
            WHERE e.Id = ?
            GROUP BY e.Id
        """, (user_id,))
        
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            return jsonify({'error': 'Mitarbeiter/Benutzer nicht gefunden'}), 404
        
        return jsonify({
            'id': row['Id'],
            'email': row['Email'],
            'vorname': row['Vorname'],
            'name': row['Name'],
            'fullName': f"{row['Vorname']} {row['Name']}",
            'personalnummer': row['Personalnummer'],
            'funktion': row['Funktion'],
            'teamId': row['TeamId'],
            'teamName': row['TeamName'],
            'lockoutEnd': row['LockoutEnd'],
            'accessFailedCount': row['AccessFailedCount'],
            'isActive': bool(row['IsActive']),
            'roles': row['roles'].split(',') if row['roles'] else [],
            # Employee data
            'isFerienjobber': bool(row['IsFerienjobber']),
            'isBrandmeldetechniker': bool(row['IsBrandmeldetechniker']),
            'isBrandschutzbeauftragter': bool(row['IsBrandschutzbeauftragter']),
            'isTdQualified': bool(row['IsTdQualified']),
            'isTeamLeader': bool(row['IsTeamLeader'])
        })
    
    @app.route('/api/users', methods=['POST'])
    @require_role('Admin')
    def create_user():
        """Create new employee with authentication credentials (Admin only)"""
        try:
            data = request.get_json()
            
            # Required fields
            vorname = data.get('vorname')
            name = data.get('name')
            personalnummer = data.get('personalnummer')
            email = data.get('email')
            password = data.get('password')
            roles = data.get('roles', ['Mitarbeiter'])
            
            # Optional fields
            funktion = data.get('funktion')
            team_id = data.get('teamId')
            geburtsdatum = data.get('geburtsdatum')
            
            # Qualifications
            is_ferienjobber = data.get('isFerienjobber', False)
            is_bmt = data.get('isBrandmeldetechniker', False)
            is_bsb = data.get('isBrandschutzbeauftragter', False)
            is_td_qualified = data.get('isTdQualified', is_bmt or is_bsb)  # Auto-set if BMT or BSB
            is_team_leader = data.get('isTeamLeader', False)
            
            # Validation
            if not vorname or not name or not personalnummer:
                return jsonify({'error': 'Vorname, Name und Personalnummer sind erforderlich'}), 400
            
            if email and not password:
                return jsonify({'error': 'Passwort ist erforderlich wenn E-Mail angegeben wird'}), 400
            
            # Validate roles
            valid_roles = ['Admin', 'Mitarbeiter']
            if not isinstance(roles, list):
                roles = [roles]
            for role in roles:
                if role not in valid_roles:
                    return jsonify({'error': f'Ungültige Rolle: {role}. Erlaubt: {", ".join(valid_roles)}'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if personalnummer already exists
            cursor.execute("SELECT Id FROM Employees WHERE Personalnummer = ?", (personalnummer,))
            if cursor.fetchone():
                conn.close()
                return jsonify({'error': 'Personalnummer bereits vorhanden'}), 400
            
            # Check if email already exists
            if email:
                cursor.execute("SELECT Id FROM Employees WHERE Email = ?", (email,))
                if cursor.fetchone():
                    conn.close()
                    return jsonify({'error': 'E-Mail wird bereits verwendet'}), 400
            
            # Create employee with authentication data
            password_hash = hash_password(password) if password else None
            security_stamp = secrets.token_hex(16) if password else None
            
            cursor.execute("""
                INSERT INTO Employees 
                (Vorname, Name, Personalnummer, Email, NormalizedEmail, PasswordHash, SecurityStamp,
                 Geburtsdatum, Funktion, TeamId, AccessFailedCount, IsActive,
                 IsFerienjobber, IsBrandmeldetechniker, IsBrandschutzbeauftragter, 
                 IsTdQualified, IsTeamLeader)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 1, ?, ?, ?, ?, ?)
            """, (
                vorname, name, personalnummer,
                email, email.upper() if email else None, password_hash, security_stamp,
                geburtsdatum, funktion, team_id,
                1 if is_ferienjobber else 0,
                1 if is_bmt else 0,
                1 if is_bsb else 0,
                1 if is_td_qualified else 0,
                1 if is_team_leader else 0
            ))
            
            employee_id = cursor.lastrowid
            
            # Assign roles (UserId in AspNetUserRoles now contains EmployeeId)
            for role in roles:
                cursor.execute("SELECT Id FROM AspNetRoles WHERE Name = ?", (role,))
                role_row = cursor.fetchone()
                if role_row:
                    cursor.execute("""
                        INSERT INTO AspNetUserRoles (UserId, RoleId)
                        VALUES (?, ?)
                    """, (str(employee_id), role_row['Id']))
            
            # Log audit entry
            changes = json.dumps({
                'vorname': vorname,
                'name': name,
                'personalnummer': personalnummer,
                'email': email,
                'funktion': funktion,
                'teamId': team_id,
                'roles': roles
            }, ensure_ascii=False)
            log_audit(conn, 'Employee', employee_id, 'Created', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'userId': employee_id, 'employeeId': employee_id}), 201
            
        except Exception as e:
            app.logger.error(f"Create employee/user error: {str(e)}")
            return jsonify({'error': f'Fehler beim Erstellen: {str(e)}'}), 500
    
    @app.route('/api/users/<int:user_id>', methods=['PUT'])
    @require_role('Admin')
    def update_user(user_id):
        """Update employee with authentication data (Admin only)"""
        try:
            data = request.get_json()
            
            # Employee data
            vorname = data.get('vorname')
            name = data.get('name')
            personalnummer = data.get('personalnummer')
            email = data.get('email')
            password = data.get('password')  # Optional password change
            funktion = data.get('funktion')
            team_id = data.get('teamId')
            geburtsdatum = data.get('geburtsdatum')
            roles = data.get('roles', [])
            
            # Qualifications
            is_ferienjobber = data.get('isFerienjobber', False)
            is_bmt = data.get('isBrandmeldetechniker', False)
            is_bsb = data.get('isBrandschutzbeauftragter', False)
            is_td_qualified = data.get('isTdQualified', is_bmt or is_bsb)
            is_team_leader = data.get('isTeamLeader', False)
            
            # Validation
            if not vorname or not name or not personalnummer:
                return jsonify({'error': 'Vorname, Name und Personalnummer sind erforderlich'}), 400
            
            # Validate roles
            valid_roles = ['Admin', 'Mitarbeiter']
            if not isinstance(roles, list):
                roles = [roles]
            for role in roles:
                if role not in valid_roles:
                    return jsonify({'error': f'Ungültige Rolle: {role}. Erlaubt: {", ".join(valid_roles)}'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if employee exists
            cursor.execute("""
                SELECT Vorname, Name, Personalnummer, Email, Funktion, TeamId
                FROM Employees WHERE Id = ?
            """, (user_id,))
            old_row = cursor.fetchone()
            if not old_row:
                conn.close()
                return jsonify({'error': 'Mitarbeiter/Benutzer nicht gefunden'}), 404
            
            # Check if personalnummer is taken by another employee
            cursor.execute("SELECT Id FROM Employees WHERE Personalnummer = ? AND Id != ?", 
                          (personalnummer, user_id))
            if cursor.fetchone():
                conn.close()
                return jsonify({'error': 'Personalnummer bereits von anderem Mitarbeiter verwendet'}), 400
            
            # Check if email is taken by another employee
            if email:
                cursor.execute("SELECT Id FROM Employees WHERE Email = ? AND Id != ?", (email, user_id))
                if cursor.fetchone():
                    conn.close()
                    return jsonify({'error': 'E-Mail wird bereits verwendet'}), 400
            
            # Update employee
            if password:
                password_hash = hash_password(password)
                security_stamp = secrets.token_hex(16)
                cursor.execute("""
                    UPDATE Employees 
                    SET Vorname = ?, Name = ?, Personalnummer = ?, Email = ?, NormalizedEmail = ?,
                        PasswordHash = ?, SecurityStamp = ?, Geburtsdatum = ?, Funktion = ?, TeamId = ?,
                        IsFerienjobber = ?, IsBrandmeldetechniker = ?, IsBrandschutzbeauftragter = ?,
                        IsTdQualified = ?, IsTeamLeader = ?
                    WHERE Id = ?
                """, (
                    vorname, name, personalnummer, email, email.upper() if email else None,
                    password_hash, security_stamp, geburtsdatum, funktion, team_id,
                    1 if is_ferienjobber else 0, 1 if is_bmt else 0, 1 if is_bsb else 0,
                    1 if is_td_qualified else 0, 1 if is_team_leader else 0,
                    user_id
                ))
            else:
                cursor.execute("""
                    UPDATE Employees 
                    SET Vorname = ?, Name = ?, Personalnummer = ?, Email = ?, NormalizedEmail = ?,
                        Geburtsdatum = ?, Funktion = ?, TeamId = ?,
                        IsFerienjobber = ?, IsBrandmeldetechniker = ?, IsBrandschutzbeauftragter = ?,
                        IsTdQualified = ?, IsTeamLeader = ?
                    WHERE Id = ?
                """, (
                    vorname, name, personalnummer, email, email.upper() if email else None,
                    geburtsdatum, funktion, team_id,
                    1 if is_ferienjobber else 0, 1 if is_bmt else 0, 1 if is_bsb else 0,
                    1 if is_td_qualified else 0, 1 if is_team_leader else 0,
                    user_id
                ))
            
            # Update roles - delete old roles and insert new ones
            cursor.execute("DELETE FROM AspNetUserRoles WHERE UserId = ?", (str(user_id),))
            for role in roles:
                cursor.execute("SELECT Id FROM AspNetRoles WHERE Name = ?", (role,))
                role_row = cursor.fetchone()
                if role_row:
                    cursor.execute("""
                        INSERT INTO AspNetUserRoles (UserId, RoleId)
                        VALUES (?, ?)
                    """, (str(user_id), role_row['Id']))
            
            # Log audit entry with changes
            changes_dict = {}
            if old_row['Vorname'] != vorname:
                changes_dict['vorname'] = {'old': old_row['Vorname'], 'new': vorname}
            if old_row['Name'] != name:
                changes_dict['name'] = {'old': old_row['Name'], 'new': name}
            if old_row['Personalnummer'] != personalnummer:
                changes_dict['personalnummer'] = {'old': old_row['Personalnummer'], 'new': personalnummer}
            if old_row['Email'] != email:
                changes_dict['email'] = {'old': old_row['Email'], 'new': email}
            changes_dict['roles'] = roles
            if password:
                changes_dict['passwordChanged'] = True
            
            if changes_dict:
                changes = json.dumps(changes_dict, ensure_ascii=False)
                log_audit(conn, 'Employee', user_id, 'Updated', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update employee/user error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    @app.route('/api/users/<int:user_id>', methods=['DELETE'])
    @require_role('Admin')
    def delete_user(user_id):
        """Delete employee/user (Admin only)"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if employee exists and get info for audit
            cursor.execute("SELECT Vorname, Name, Email FROM Employees WHERE Id = ?", (user_id,))
            employee_row = cursor.fetchone()
            if not employee_row:
                conn.close()
                return jsonify({'error': 'Mitarbeiter/Benutzer nicht gefunden'}), 404
            
            # Prevent deleting the last admin
            cursor.execute("""
                SELECT COUNT(*) as count
                FROM Employees e
                JOIN AspNetUserRoles ur ON CAST(e.Id AS TEXT) = ur.UserId
                JOIN AspNetRoles r ON ur.RoleId = r.Id
                WHERE r.Name = 'Admin'
            """)
            admin_count = cursor.fetchone()['count']
            
            cursor.execute("""
                SELECT COUNT(*) as is_admin
                FROM AspNetUserRoles ur
                JOIN AspNetRoles r ON ur.RoleId = r.Id
                WHERE ur.UserId = ? AND r.Name = 'Admin'
            """, (str(user_id),))
            is_admin = cursor.fetchone()['is_admin'] > 0
            
            if is_admin and admin_count <= 1:
                conn.close()
                return jsonify({'error': 'Der letzte Administrator kann nicht gelöscht werden'}), 400
            
            # Delete roles first (foreign key constraint)
            cursor.execute("DELETE FROM AspNetUserRoles WHERE UserId = ?", (str(user_id),))
            
            # Delete employee (cascade will handle related data)
            cursor.execute("DELETE FROM Employees WHERE Id = ?", (user_id,))
            
            # Log audit entry
            changes = json.dumps({
                'vorname': employee_row['Vorname'],
                'name': employee_row['Name'],
                'email': employee_row['Email']
            }, ensure_ascii=False)
            log_audit(conn, 'Employee', user_id, 'Deleted', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Delete employee/user error: {str(e)}")
            return jsonify({'error': f'Fehler beim Löschen: {str(e)}'}), 500
    
    @app.route('/api/roles', methods=['GET'])
    @require_role('Admin')
    def get_roles():
        """Get all available roles (Admin only)"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT Id, Name FROM AspNetRoles ORDER BY Name")
        
        roles = []
        for row in cursor.fetchall():
            roles.append({
                'id': row['Id'],
                'name': row['Name']
            })
        
        conn.close()
        return jsonify(roles)
    
    # ============================================================================
    # EMPLOYEE ENDPOINTS
    # ============================================================================
    
    @app.route('/api/employees', methods=['GET'])
    def get_employees():
        """Get all employees (now includes authentication data)"""
        conn = None
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT e.*, t.Name as TeamName,
                       GROUP_CONCAT(r.Name) as roles
                FROM Employees e
                LEFT JOIN Teams t ON e.TeamId = t.Id
                LEFT JOIN AspNetUserRoles ur ON CAST(e.Id AS TEXT) = ur.UserId
                LEFT JOIN AspNetRoles r ON ur.RoleId = r.Id
                GROUP BY e.Id
                ORDER BY e.Name, e.Vorname
            """)
            
            employees = []
            for row in cursor.fetchall():
                # Handle fields which may not exist in older databases
                try:
                    is_td_qualified = bool(row['IsTdQualified'])
                except (KeyError, IndexError):
                    is_td_qualified = False
                
                try:
                    is_team_leader = bool(row['IsTeamLeader'])
                except (KeyError, IndexError):
                    is_team_leader = False
                
                try:
                    is_active = bool(row['IsActive'])
                except (KeyError, IndexError):
                    is_active = True
                
                employees.append({
                    'id': row['Id'],
                    'vorname': row['Vorname'],
                    'name': row['Name'],
                    'personalnummer': row['Personalnummer'],
                    'email': row['Email'],
                    'geburtsdatum': row['Geburtsdatum'],
                    'funktion': row['Funktion'],
                    'isFerienjobber': bool(row['IsFerienjobber']),
                    'isBrandmeldetechniker': bool(row['IsBrandmeldetechniker']),
                    'isBrandschutzbeauftragter': bool(row['IsBrandschutzbeauftragter']),
                    'isTdQualified': is_td_qualified,
                    'isTeamLeader': is_team_leader,
                    'isActive': is_active,
                    'teamId': row['TeamId'],
                    'teamName': row['TeamName'],
                    'fullName': f"{row['Vorname']} {row['Name']}",
                    'hasPassword': bool(row['Email']),  # Has auth if email is set
                    'roles': row['roles'].split(',') if row['roles'] else []
                })
            
            return jsonify(employees)
        except Exception as e:
            return jsonify({'error': f'Database error: {str(e)}'}), 500
        finally:
            if conn:
                conn.close()
    
    @app.route('/api/employees/<int:id>', methods=['GET'])
    def get_employee(id):
        """Get employee by ID"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT e.*, t.Name as TeamName,
                   GROUP_CONCAT(r.Name) as roles
            FROM Employees e
            LEFT JOIN Teams t ON e.TeamId = t.Id
            LEFT JOIN AspNetUserRoles ur ON CAST(e.Id AS TEXT) = ur.UserId
            LEFT JOIN AspNetRoles r ON ur.RoleId = r.Id
            WHERE e.Id = ?
            GROUP BY e.Id
        """, (id,))
        
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            return jsonify({'error': 'Employee not found'}), 404
        
        # Handle IsTdQualified field which may not exist in older databases
        try:
            is_td_qualified = bool(row['IsTdQualified'])
        except (KeyError, IndexError):
            is_td_qualified = False
        
        # Handle IsTeamLeader field which may not exist in older databases
        try:
            is_team_leader = bool(row['IsTeamLeader'])
        except (KeyError, IndexError):
            is_team_leader = False
        
        return jsonify({
            'id': row['Id'],
            'vorname': row['Vorname'],
            'name': row['Name'],
            'personalnummer': row['Personalnummer'],
            'email': row['Email'],
            'geburtsdatum': row['Geburtsdatum'],
            'funktion': row['Funktion'],
            'isSpringer': bool(row['IsSpringer']),
            'isFerienjobber': bool(row['IsFerienjobber']),
            'isBrandmeldetechniker': bool(row['IsBrandmeldetechniker']),
            'isBrandschutzbeauftragter': bool(row['IsBrandschutzbeauftragter']),
            'isTdQualified': is_td_qualified,
            'isTeamLeader': is_team_leader,
            'teamId': row['TeamId'],
            'teamName': row['TeamName'],
            'fullName': f"{row['Vorname']} {row['Name']}",
            'roles': row['roles'] if row['roles'] else ''
        })
    
    @app.route('/api/employees/springers', methods=['GET'])
    def get_springers():
        """Get all springers"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT e.*, t.Name as TeamName
            FROM Employees e
            LEFT JOIN Teams t ON e.TeamId = t.Id
            WHERE e.IsSpringer = 1
            ORDER BY e.Name, e.Vorname
        """)
        
        springers = []
        for row in cursor.fetchall():
            springers.append({
                'id': row['Id'],
                'vorname': row['Vorname'],
                'name': row['Name'],
                'personalnummer': row['Personalnummer'],
                'email': row['Email'],
                'isSpringer': True,
                'teamId': row['TeamId'],
                'teamName': row['TeamName'],
                'fullName': f"{row['Vorname']} {row['Name']}"
            })
        
        conn.close()
        return jsonify(springers)
    
    @app.route('/api/employees', methods=['POST'])
    @require_role('Admin')
    def create_employee():
        """Create new employee"""
        try:
            data = request.get_json()
            
            # Validate required fields
            if not data.get('vorname') or not data.get('name') or not data.get('personalnummer'):
                return jsonify({'error': 'Vorname, Name und Personalnummer sind Pflichtfelder'}), 400
            
            # Validate password - required for new employees
            password = data.get('password')
            if not password:
                return jsonify({'error': 'Passwort ist erforderlich'}), 400
            
            if len(password) < 8:
                return jsonify({'error': 'Passwort muss mindestens 8 Zeichen lang sein'}), 400
            
            # Validate Funktion field - only allow specific values
            funktion = data.get('funktion')
            if funktion and funktion not in ['Brandmeldetechniker', 'Brandschutzbeauftragter', 'Techniker', 'Springer']:
                return jsonify({'error': 'Ungültige Funktion. Erlaubt: Brandmeldetechniker, Brandschutzbeauftragter, Techniker, Springer'}), 400
            
            # Use checkbox values directly from frontend for BMT/BSB flags
            is_bmt = 1 if data.get('isBrandmeldetechniker') else 0
            is_bsb = 1 if data.get('isBrandschutzbeauftragter') else 0
            # TD qualification is automatically set if BMT or BSB is true
            is_td = 1 if (is_bmt or is_bsb) else 0
            is_team_leader = 1 if data.get('isTeamLeader') else 0
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if Personalnummer already exists
            cursor.execute("SELECT Id FROM Employees WHERE Personalnummer = ?", (data.get('personalnummer'),))
            if cursor.fetchone():
                conn.close()
                return jsonify({'error': 'Personalnummer bereits vorhanden'}), 400
            
            # Check if email already exists
            email = data.get('email')
            if email:
                cursor.execute("SELECT Id FROM Employees WHERE Email = ?", (email,))
                if cursor.fetchone():
                    conn.close()
                    return jsonify({'error': 'E-Mail wird bereits verwendet'}), 400
            
            # Hash password
            password_hash = hash_password(password)
            security_stamp = secrets.token_hex(16)
            
            cursor.execute("""
                INSERT INTO Employees 
                (Vorname, Name, Personalnummer, Email, NormalizedEmail, PasswordHash, SecurityStamp,
                 Geburtsdatum, Funktion, 
                 IsSpringer, IsFerienjobber, IsBrandmeldetechniker, IsBrandschutzbeauftragter, IsTdQualified, IsTeamLeader, TeamId)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('vorname'),
                data.get('name'),
                data.get('personalnummer'),
                email,
                email.upper() if email else None,
                password_hash,
                security_stamp,
                data.get('geburtsdatum'),
                funktion,
                1 if data.get('isSpringer') else 0,
                1 if data.get('isFerienjobber') else 0,
                is_bmt,
                is_bsb,
                is_td,
                is_team_leader,
                data.get('teamId')
            ))
            
            employee_id = cursor.lastrowid
            
            # Assign roles: All employees get "Mitarbeiter" role by default
            cursor.execute("SELECT Id FROM AspNetRoles WHERE Name = ?", ('Mitarbeiter',))
            mitarbeiter_role = cursor.fetchone()
            if not mitarbeiter_role:
                conn.close()
                app.logger.error("Mitarbeiter role not found in database")
                return jsonify({'error': 'System-Fehler: Mitarbeiter-Rolle nicht gefunden'}), 500
            
            cursor.execute("""
                INSERT INTO AspNetUserRoles (UserId, RoleId)
                VALUES (?, ?)
            """, (str(employee_id), mitarbeiter_role['Id']))
            
            # If isAdmin flag is set, also assign "Admin" role
            is_admin = data.get('isAdmin', False)
            if is_admin:
                cursor.execute("SELECT Id FROM AspNetRoles WHERE Name = ?", ('Admin',))
                admin_role = cursor.fetchone()
                if not admin_role:
                    conn.close()
                    app.logger.error("Admin role not found in database")
                    return jsonify({'error': 'System-Fehler: Admin-Rolle nicht gefunden'}), 500
                
                cursor.execute("""
                    INSERT INTO AspNetUserRoles (UserId, RoleId)
                    VALUES (?, ?)
                """, (str(employee_id), admin_role['Id']))
            
            # Log audit entry
            changes = json.dumps({
                'vorname': data.get('vorname'),
                'name': data.get('name'),
                'personalnummer': data.get('personalnummer'),
                'email': data.get('email'),
                'funktion': funktion,
                'teamId': data.get('teamId'),
                'roles': ['Mitarbeiter', 'Admin'] if is_admin else ['Mitarbeiter']
            }, ensure_ascii=False)
            log_audit(conn, 'Employee', employee_id, 'Created', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'id': employee_id}), 201
            
        except Exception as e:
            app.logger.error(f"Create employee error: {str(e)}")
            return jsonify({'error': f'Fehler beim Erstellen: {str(e)}'}), 500
    
    @app.route('/api/employees/<int:id>', methods=['PUT'])
    @require_role('Admin')
    def update_employee(id):
        """Update employee"""
        try:
            data = request.get_json()
            
            # Validate required fields
            if not data.get('vorname') or not data.get('name') or not data.get('personalnummer'):
                return jsonify({'error': 'Vorname, Name und Personalnummer sind Pflichtfelder'}), 400
            
            # Validate password if provided
            password = data.get('password')
            if password and len(password) < 8:
                return jsonify({'error': 'Passwort muss mindestens 8 Zeichen lang sein'}), 400
            
            # Validate Funktion field
            funktion = data.get('funktion')
            if funktion and funktion not in ['Brandmeldetechniker', 'Brandschutzbeauftragter', 'Techniker', 'Springer']:
                return jsonify({'error': 'Ungültige Funktion. Erlaubt: Brandmeldetechniker, Brandschutzbeauftragter, Techniker, Springer'}), 400
            
            # Use checkbox values directly from frontend for BMT/BSB flags
            is_bmt = 1 if data.get('isBrandmeldetechniker') else 0
            is_bsb = 1 if data.get('isBrandschutzbeauftragter') else 0
            # TD qualification is automatically set if BMT or BSB is true
            is_td = 1 if (is_bmt or is_bsb) else 0
            is_team_leader = 1 if data.get('isTeamLeader') else 0
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if employee exists and get old values for audit
            cursor.execute("""
                SELECT Vorname, Name, Personalnummer, Email, Geburtsdatum, Funktion, 
                       IsSpringer, IsFerienjobber, IsBrandmeldetechniker, 
                       IsBrandschutzbeauftragter, IsTdQualified, TeamId 
                FROM Employees WHERE Id = ?
            """, (id,))
            old_row = cursor.fetchone()
            if not old_row:
                conn.close()
                return jsonify({'error': 'Mitarbeiter nicht gefunden'}), 404
            
            # Check if Personalnummer is taken by another employee
            cursor.execute("SELECT Id FROM Employees WHERE Personalnummer = ? AND Id != ?", 
                          (data.get('personalnummer'), id))
            if cursor.fetchone():
                conn.close()
                return jsonify({'error': 'Personalnummer bereits von anderem Mitarbeiter verwendet'}), 400
            
            # Check if email is taken by another employee
            email = data.get('email')
            if email:
                cursor.execute("SELECT Id FROM Employees WHERE Email = ? AND Id != ?", (email, id))
                if cursor.fetchone():
                    conn.close()
                    return jsonify({'error': 'E-Mail wird bereits verwendet'}), 400
            
            # Update employee with or without password
            if password:
                password_hash = hash_password(password)
                security_stamp = secrets.token_hex(16)
                cursor.execute("""
                    UPDATE Employees 
                    SET Vorname = ?, Name = ?, Personalnummer = ?, Email = ?, NormalizedEmail = ?,
                        PasswordHash = ?, SecurityStamp = ?, Geburtsdatum = ?, 
                        Funktion = ?, IsSpringer = ?, IsFerienjobber = ?, 
                        IsBrandmeldetechniker = ?, IsBrandschutzbeauftragter = ?, IsTdQualified = ?, IsTeamLeader = ?, TeamId = ?
                    WHERE Id = ?
                """, (
                    data.get('vorname'),
                    data.get('name'),
                    data.get('personalnummer'),
                    email,
                    email.upper() if email else None,
                    password_hash,
                    security_stamp,
                    data.get('geburtsdatum'),
                    funktion,
                    1 if data.get('isSpringer') else 0,
                    1 if data.get('isFerienjobber') else 0,
                    is_bmt,
                    is_bsb,
                    is_td,
                    is_team_leader,
                    data.get('teamId'),
                    id
                ))
            else:
                cursor.execute("""
                    UPDATE Employees 
                    SET Vorname = ?, Name = ?, Personalnummer = ?, Email = ?, NormalizedEmail = ?, Geburtsdatum = ?, 
                        Funktion = ?, IsSpringer = ?, IsFerienjobber = ?, 
                        IsBrandmeldetechniker = ?, IsBrandschutzbeauftragter = ?, IsTdQualified = ?, IsTeamLeader = ?, TeamId = ?
                    WHERE Id = ?
                """, (
                    data.get('vorname'),
                    data.get('name'),
                    data.get('personalnummer'),
                    email,
                    email.upper() if email else None,
                    data.get('geburtsdatum'),
                    funktion,
                    1 if data.get('isSpringer') else 0,
                    1 if data.get('isFerienjobber') else 0,
                    is_bmt,
                    is_bsb,
                    is_td,
                    is_team_leader,
                    data.get('teamId'),
                    id
                ))
            
            # Update roles: Ensure employee always has "Mitarbeiter" role
            cursor.execute("SELECT Id FROM AspNetRoles WHERE Name = ?", ('Mitarbeiter',))
            mitarbeiter_role = cursor.fetchone()
            if not mitarbeiter_role:
                conn.close()
                app.logger.error("Mitarbeiter role not found in database")
                return jsonify({'error': 'System-Fehler: Mitarbeiter-Rolle nicht gefunden'}), 500
            
            # Check if Mitarbeiter role exists, add if not
            cursor.execute("""
                SELECT 1 FROM AspNetUserRoles 
                WHERE UserId = ? AND RoleId = ?
            """, (str(id), mitarbeiter_role['Id']))
            if not cursor.fetchone():
                cursor.execute("""
                    INSERT INTO AspNetUserRoles (UserId, RoleId)
                    VALUES (?, ?)
                """, (str(id), mitarbeiter_role['Id']))
            
            # Handle Admin role based on isAdmin flag
            is_admin = data.get('isAdmin', False)
            cursor.execute("SELECT Id FROM AspNetRoles WHERE Name = ?", ('Admin',))
            admin_role = cursor.fetchone()
            
            if not admin_role:
                conn.close()
                app.logger.error("Admin role not found in database")
                return jsonify({'error': 'System-Fehler: Admin-Rolle nicht gefunden'}), 500
            
            # Check if Admin role currently exists for this employee
            cursor.execute("""
                SELECT 1 FROM AspNetUserRoles 
                WHERE UserId = ? AND RoleId = ?
            """, (str(id), admin_role['Id']))
            has_admin_role = cursor.fetchone() is not None
            
            if is_admin and not has_admin_role:
                # Add Admin role
                cursor.execute("""
                    INSERT INTO AspNetUserRoles (UserId, RoleId)
                    VALUES (?, ?)
                """, (str(id), admin_role['Id']))
            elif not is_admin and has_admin_role:
                # Remove Admin role
                cursor.execute("""
                    DELETE FROM AspNetUserRoles 
                    WHERE UserId = ? AND RoleId = ?
                """, (str(id), admin_role['Id']))
            
            # Log audit entry with changes
            changes_dict = {}
            if old_row['Vorname'] != data.get('vorname'):
                changes_dict['vorname'] = {'old': old_row['Vorname'], 'new': data.get('vorname')}
            if old_row['Name'] != data.get('name'):
                changes_dict['name'] = {'old': old_row['Name'], 'new': data.get('name')}
            if old_row['Personalnummer'] != data.get('personalnummer'):
                changes_dict['personalnummer'] = {'old': old_row['Personalnummer'], 'new': data.get('personalnummer')}
            if old_row['Email'] != data.get('email'):
                changes_dict['email'] = {'old': old_row['Email'], 'new': data.get('email')}
            if old_row['Funktion'] != funktion:
                changes_dict['funktion'] = {'old': old_row['Funktion'], 'new': funktion}
            if old_row['TeamId'] != data.get('teamId'):
                changes_dict['teamId'] = {'old': old_row['TeamId'], 'new': data.get('teamId')}
            if password:
                changes_dict['password'] = 'changed'
            
            # Track role changes in audit log
            if 'isAdmin' in data:
                # Get old roles for comparison
                cursor.execute("""
                    SELECT r.Name FROM AspNetUserRoles ur
                    JOIN AspNetRoles r ON ur.RoleId = r.Id
                    WHERE ur.UserId = ?
                """, (str(id),))
                old_roles = [row['Name'] for row in cursor.fetchall()]
                new_roles = ['Mitarbeiter']
                if is_admin:
                    new_roles.append('Admin')
                
                if set(old_roles) != set(new_roles):
                    changes_dict['roles'] = {'old': old_roles, 'new': new_roles}
            
            if changes_dict:
                changes = json.dumps(changes_dict, ensure_ascii=False)
                log_audit(conn, 'Employee', id, 'Updated', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update employee error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    @app.route('/api/employees/<int:id>', methods=['DELETE'])
    @require_role('Admin')
    def delete_employee(id):
        """Delete employee (Admin only)"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if employee exists and get info for audit
            cursor.execute("SELECT Vorname, Name, Personalnummer FROM Employees WHERE Id = ?", (id,))
            emp_row = cursor.fetchone()
            if not emp_row:
                conn.close()
                return jsonify({'error': 'Mitarbeiter nicht gefunden'}), 404
            
            # Check if employee has assignments
            cursor.execute("SELECT COUNT(*) as count FROM ShiftAssignments WHERE EmployeeId = ?", (id,))
            assignment_count = cursor.fetchone()['count']
            
            if assignment_count > 0:
                conn.close()
                return jsonify({'error': f'Mitarbeiter hat {assignment_count} Schichtzuweisungen und kann nicht gelöscht werden'}), 400
            
            # Delete employee
            cursor.execute("DELETE FROM Employees WHERE Id = ?", (id,))
            
            # Log audit entry
            changes = json.dumps({
                'vorname': emp_row['Vorname'],
                'name': emp_row['Name'],
                'personalnummer': emp_row['Personalnummer']
            }, ensure_ascii=False)
            log_audit(conn, 'Employee', id, 'Deleted', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Delete employee error: {str(e)}")
            return jsonify({'error': f'Fehler beim Löschen: {str(e)}'}), 500
    
    # ============================================================================
    # TEAM ENDPOINTS
    # ============================================================================
    
    @app.route('/api/teams', methods=['GET'])
    def get_teams():
        """Get all teams with employee count"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT t.Id, t.Name, t.Description, t.Email, t.IsVirtual,
                   COUNT(e.Id) as EmployeeCount
            FROM Teams t
            LEFT JOIN Employees e ON t.Id = e.TeamId
            GROUP BY t.Id, t.Name, t.Description, t.Email, t.IsVirtual
            ORDER BY t.Name
        """)
        
        teams = []
        for row in cursor.fetchall():
            employee_count = row['EmployeeCount']
            
            # For virtual teams, count employees with special qualifications instead of TeamId
            if bool(row['IsVirtual']):
                employee_count = get_virtual_team_employee_count(cursor, row['Id'])
            
            teams.append({
                'id': row['Id'],
                'name': row['Name'],
                'description': row['Description'],
                'email': row['Email'],
                'isVirtual': bool(row['IsVirtual']),
                'employeeCount': employee_count
            })
        
        conn.close()
        return jsonify(teams)
    
    @app.route('/api/teams/<int:id>', methods=['GET'])
    def get_team(id):
        """Get single team by ID"""
        conn = None
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT t.Id, t.Name, t.Description, t.Email, t.IsVirtual,
                       COUNT(e.Id) as EmployeeCount
                FROM Teams t
                LEFT JOIN Employees e ON t.Id = e.TeamId
                WHERE t.Id = ?
                GROUP BY t.Id, t.Name, t.Description, t.Email, t.IsVirtual
            """, (id,))
            
            row = cursor.fetchone()
            
            if not row:
                return jsonify({'error': 'Team nicht gefunden'}), 404
            
            employee_count = row['EmployeeCount']
            
            # For virtual teams, count employees with special qualifications instead of TeamId
            if bool(row['IsVirtual']):
                employee_count = get_virtual_team_employee_count(cursor, row['Id'])
            
            return jsonify({
                'id': row['Id'],
                'name': row['Name'],
                'description': row['Description'],
                'email': row['Email'],
                'isVirtual': bool(row['IsVirtual']),
                'employeeCount': employee_count
            })
        finally:
            if conn:
                conn.close()
    
    @app.route('/api/teams', methods=['POST'])
    @require_role('Admin')
    def create_team():
        """Create new team"""
        try:
            data = request.get_json()
            
            # Validate required fields
            if not data.get('name'):
                return jsonify({'error': 'Teamname ist Pflichtfeld'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO Teams (Name, Description, Email, IsVirtual)
                VALUES (?, ?, ?, ?)
            """, (
                data.get('name'),
                data.get('description'),
                data.get('email'),
                1 if data.get('isVirtual') else 0
            ))
            
            team_id = cursor.lastrowid
            
            # Log audit entry
            changes = json.dumps({
                'name': data.get('name'),
                'description': data.get('description'),
                'email': data.get('email'),
                'isVirtual': data.get('isVirtual')
            }, ensure_ascii=False)
            log_audit(conn, 'Team', team_id, 'Created', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'id': team_id}), 201
            
        except Exception as e:
            app.logger.error(f"Create team error: {str(e)}")
            return jsonify({'error': f'Fehler beim Erstellen: {str(e)}'}), 500
    
    @app.route('/api/teams/<int:id>', methods=['PUT'])
    @require_role('Admin')
    def update_team(id):
        """Update team"""
        try:
            data = request.get_json()
            
            # Validate required fields
            if not data.get('name'):
                return jsonify({'error': 'Teamname ist Pflichtfeld'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if team exists and get old values for audit
            cursor.execute("SELECT Name, Description, Email, IsVirtual FROM Teams WHERE Id = ?", (id,))
            old_row = cursor.fetchone()
            if not old_row:
                conn.close()
                return jsonify({'error': 'Team nicht gefunden'}), 404
            
            cursor.execute("""
                UPDATE Teams 
                SET Name = ?, Description = ?, Email = ?, IsVirtual = ?
                WHERE Id = ?
            """, (
                data.get('name'),
                data.get('description'),
                data.get('email'),
                1 if data.get('isVirtual') else 0,
                id
            ))
            
            # Log audit entry with changes
            changes_dict = {}
            if old_row['Name'] != data.get('name'):
                changes_dict['name'] = {'old': old_row['Name'], 'new': data.get('name')}
            if old_row['Description'] != data.get('description'):
                changes_dict['description'] = {'old': old_row['Description'], 'new': data.get('description')}
            if old_row['Email'] != data.get('email'):
                changes_dict['email'] = {'old': old_row['Email'], 'new': data.get('email')}
            new_is_virtual = 1 if data.get('isVirtual') else 0
            if old_row['IsVirtual'] != new_is_virtual:
                changes_dict['isVirtual'] = {'old': bool(old_row['IsVirtual']), 'new': bool(new_is_virtual)}
            
            if changes_dict:
                changes = json.dumps(changes_dict, ensure_ascii=False)
                log_audit(conn, 'Team', id, 'Updated', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update team error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    @app.route('/api/teams/<int:id>', methods=['DELETE'])
    @require_role('Admin')
    def delete_team(id):
        """Delete team (Admin only)"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if team exists and get info for audit
            cursor.execute("SELECT Name FROM Teams WHERE Id = ?", (id,))
            team_row = cursor.fetchone()
            if not team_row:
                conn.close()
                return jsonify({'error': 'Team nicht gefunden'}), 404
            
            # Check if team has employees
            cursor.execute("SELECT COUNT(*) as count FROM Employees WHERE TeamId = ?", (id,))
            employee_count = cursor.fetchone()['count']
            
            if employee_count > 0:
                conn.close()
                return jsonify({'error': f'Team hat {employee_count} Mitarbeiter und kann nicht gelöscht werden'}), 400
            
            # Clear TeamId from AdminNotifications to avoid foreign key constraint violations
            # (AdminNotifications don't have CASCADE delete)
            cursor.execute("UPDATE AdminNotifications SET TeamId = NULL WHERE TeamId = ?", (id,))
            
            # Delete team (TeamShiftAssignments will be automatically deleted due to CASCADE)
            cursor.execute("DELETE FROM Teams WHERE Id = ?", (id,))
            
            # Log audit entry
            changes = json.dumps({'name': team_row['Name']}, ensure_ascii=False)
            log_audit(conn, 'Team', id, 'Deleted', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Delete team error: {str(e)}")
            return jsonify({'error': f'Fehler beim Löschen: {str(e)}'}), 500
    
    # ============================================================================
    # VACATION PERIODS ENDPOINTS (Ferienzeiten)
    # ============================================================================
    
    @app.route('/api/vacation-periods', methods=['GET'])
    def get_vacation_periods():
        """Get all vacation periods"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT Id, Name, StartDate, EndDate, ColorCode, 
                       CreatedAt, CreatedBy, ModifiedAt, ModifiedBy
                FROM VacationPeriods
                ORDER BY StartDate
            """)
            
            periods = []
            for row in cursor.fetchall():
                periods.append({
                    'id': row['Id'],
                    'name': row['Name'],
                    'startDate': row['StartDate'],
                    'endDate': row['EndDate'],
                    'colorCode': row['ColorCode'] or '#E8F5E9',
                    'createdAt': row['CreatedAt'],
                    'createdBy': row['CreatedBy'],
                    'modifiedAt': row['ModifiedAt'],
                    'modifiedBy': row['ModifiedBy']
                })
            
            conn.close()
            return jsonify(periods)
            
        except Exception as e:
            app.logger.error(f"Get vacation periods error: {str(e)}")
            return jsonify({'error': f'Fehler beim Laden: {str(e)}'}), 500
    
    @app.route('/api/vacation-periods/<int:id>', methods=['GET'])
    def get_vacation_period(id):
        """Get a specific vacation period"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT Id, Name, StartDate, EndDate, ColorCode, 
                       CreatedAt, CreatedBy, ModifiedAt, ModifiedBy
                FROM VacationPeriods
                WHERE Id = ?
            """, (id,))
            
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                return jsonify({'error': 'Ferienzeit nicht gefunden'}), 404
            
            return jsonify({
                'id': row['Id'],
                'name': row['Name'],
                'startDate': row['StartDate'],
                'endDate': row['EndDate'],
                'colorCode': row['ColorCode'] or '#E8F5E9',
                'createdAt': row['CreatedAt'],
                'createdBy': row['CreatedBy'],
                'modifiedAt': row['ModifiedAt'],
                'modifiedBy': row['ModifiedBy']
            })
            
        except Exception as e:
            app.logger.error(f"Get vacation period error: {str(e)}")
            return jsonify({'error': f'Fehler beim Laden: {str(e)}'}), 500
    
    @app.route('/api/vacation-periods', methods=['POST'])
    @require_role('Admin')
    def create_vacation_period():
        """Create new vacation period"""
        try:
            data = request.get_json()
            
            # Validate required fields
            if not data.get('name'):
                return jsonify({'error': 'Name ist Pflichtfeld'}), 400
            if not data.get('startDate'):
                return jsonify({'error': 'Startdatum ist Pflichtfeld'}), 400
            if not data.get('endDate'):
                return jsonify({'error': 'Enddatum ist Pflichtfeld'}), 400
            
            # Validate dates
            try:
                start_date = date.fromisoformat(data.get('startDate'))
                end_date = date.fromisoformat(data.get('endDate'))
            except (ValueError, TypeError):
                return jsonify({'error': 'Ungültiges Datumsformat'}), 400
            
            if end_date < start_date:
                return jsonify({'error': 'Enddatum muss nach Startdatum liegen'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO VacationPeriods (Name, StartDate, EndDate, ColorCode, CreatedAt, CreatedBy)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                data.get('name'),
                start_date.isoformat(),
                end_date.isoformat(),
                data.get('colorCode', '#E8F5E9'),
                datetime.utcnow().isoformat(),
                session.get('user_email')
            ))
            
            period_id = cursor.lastrowid
            
            # Log audit entry
            changes = json.dumps({
                'name': data.get('name'),
                'startDate': start_date.isoformat(),
                'endDate': end_date.isoformat(),
                'colorCode': data.get('colorCode', '#E8F5E9')
            }, ensure_ascii=False)
            log_audit(conn, 'VacationPeriod', period_id, 'Created', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'id': period_id}), 201
            
        except Exception as e:
            app.logger.error(f"Create vacation period error: {str(e)}")
            return jsonify({'error': f'Fehler beim Erstellen: {str(e)}'}), 500
    
    @app.route('/api/vacation-periods/<int:id>', methods=['PUT'])
    @require_role('Admin')
    def update_vacation_period(id):
        """Update vacation period"""
        try:
            data = request.get_json()
            
            # Validate required fields
            if not data.get('name'):
                return jsonify({'error': 'Name ist Pflichtfeld'}), 400
            if not data.get('startDate'):
                return jsonify({'error': 'Startdatum ist Pflichtfeld'}), 400
            if not data.get('endDate'):
                return jsonify({'error': 'Enddatum ist Pflichtfeld'}), 400
            
            # Validate dates
            try:
                start_date = date.fromisoformat(data.get('startDate'))
                end_date = date.fromisoformat(data.get('endDate'))
            except (ValueError, TypeError):
                return jsonify({'error': 'Ungültiges Datumsformat'}), 400
            
            if end_date < start_date:
                return jsonify({'error': 'Enddatum muss nach Startdatum liegen'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if period exists and get old values for audit
            cursor.execute("""
                SELECT Name, StartDate, EndDate, ColorCode 
                FROM VacationPeriods WHERE Id = ?
            """, (id,))
            old_row = cursor.fetchone()
            if not old_row:
                conn.close()
                return jsonify({'error': 'Ferienzeit nicht gefunden'}), 404
            
            cursor.execute("""
                UPDATE VacationPeriods 
                SET Name = ?, StartDate = ?, EndDate = ?, ColorCode = ?,
                    ModifiedAt = ?, ModifiedBy = ?
                WHERE Id = ?
            """, (
                data.get('name'),
                start_date.isoformat(),
                end_date.isoformat(),
                data.get('colorCode', '#E8F5E9'),
                datetime.utcnow().isoformat(),
                session.get('user_email'),
                id
            ))
            
            # Log audit entry with changes
            changes_dict = {}
            if old_row['Name'] != data.get('name'):
                changes_dict['name'] = {'old': old_row['Name'], 'new': data.get('name')}
            if old_row['StartDate'] != start_date.isoformat():
                changes_dict['startDate'] = {'old': old_row['StartDate'], 'new': start_date.isoformat()}
            if old_row['EndDate'] != end_date.isoformat():
                changes_dict['endDate'] = {'old': old_row['EndDate'], 'new': end_date.isoformat()}
            if old_row['ColorCode'] != data.get('colorCode', '#E8F5E9'):
                changes_dict['colorCode'] = {'old': old_row['ColorCode'], 'new': data.get('colorCode', '#E8F5E9')}
            
            if changes_dict:
                changes = json.dumps(changes_dict, ensure_ascii=False)
                log_audit(conn, 'VacationPeriod', id, 'Updated', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update vacation period error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    @app.route('/api/vacation-periods/<int:id>', methods=['DELETE'])
    @require_role('Admin')
    def delete_vacation_period(id):
        """Delete vacation period (Admin only)"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if period exists and get info for audit
            cursor.execute("SELECT Name FROM VacationPeriods WHERE Id = ?", (id,))
            period_row = cursor.fetchone()
            if not period_row:
                conn.close()
                return jsonify({'error': 'Ferienzeit nicht gefunden'}), 404
            
            # Delete period
            cursor.execute("DELETE FROM VacationPeriods WHERE Id = ?", (id,))
            
            # Log audit entry
            changes = json.dumps({'name': period_row['Name']}, ensure_ascii=False)
            log_audit(conn, 'VacationPeriod', id, 'Deleted', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Delete vacation period error: {str(e)}")
            return jsonify({'error': f'Fehler beim Löschen: {str(e)}'}), 500
    
    # ============================================================================
    # SHIFT TYPE ENDPOINTS
    # ============================================================================
    
    @app.route('/api/shifttypes', methods=['GET'])
    def get_shift_types():
        """Get all shift types"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM ShiftTypes ORDER BY Id")
        
        shift_types = []
        for row in cursor.fetchall():
            shift_types.append({
                'id': row['Id'],
                'code': row['Code'],
                'name': row['Name'],
                'startTime': row['StartTime'],
                'endTime': row['EndTime'],
                'colorCode': row['ColorCode'],
                'durationHours': row['DurationHours'],
                'weeklyWorkingHours': row['WeeklyWorkingHours'],
                'isActive': bool(row['IsActive']),
                'worksMonday': bool(row['WorksMonday']),
                'worksTuesday': bool(row['WorksTuesday']),
                'worksWednesday': bool(row['WorksWednesday']),
                'worksThursday': bool(row['WorksThursday']),
                'worksFriday': bool(row['WorksFriday']),
                'worksSaturday': bool(row['WorksSaturday']),
                'worksSunday': bool(row['WorksSunday']),
                'minStaffWeekday': row['MinStaffWeekday'],
                'maxStaffWeekday': row['MaxStaffWeekday'],
                'minStaffWeekend': row['MinStaffWeekend'],
                'maxStaffWeekend': row['MaxStaffWeekend']
            })
        
        conn.close()
        return jsonify(shift_types)
    
    @app.route('/api/shifttypes', methods=['POST'])
    @require_role('Admin')
    def create_shift_type():
        """Create new shift type (Admin only)"""
        try:
            data = request.get_json()
            
            # Validate required fields
            required_fields = ['code', 'name', 'startTime', 'endTime', 'durationHours']
            for field in required_fields:
                if not data.get(field):
                    return jsonify({'error': f'{field} ist Pflichtfeld'}), 400
            
            # Validate staffing requirements
            min_staff_weekday = data.get('minStaffWeekday', 3)
            max_staff_weekday = data.get('maxStaffWeekday', 5)
            min_staff_weekend = data.get('minStaffWeekend', 2)
            max_staff_weekend = data.get('maxStaffWeekend', 3)
            
            if min_staff_weekday > max_staff_weekday:
                return jsonify({'error': 'Minimale Personalstärke an Wochentagen darf nicht größer sein als die maximale Personalstärke'}), 400
            if min_staff_weekend > max_staff_weekend:
                return jsonify({'error': 'Minimale Personalstärke am Wochenende darf nicht größer sein als die maximale Personalstärke'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if code already exists
            cursor.execute("SELECT Id FROM ShiftTypes WHERE Code = ?", (data.get('code'),))
            if cursor.fetchone():
                conn.close()
                return jsonify({'error': 'Schichtkürzel bereits vorhanden'}), 400
            
            # Insert shift type
            cursor.execute("""
                INSERT INTO ShiftTypes (Code, Name, StartTime, EndTime, DurationHours, ColorCode, IsActive,
                                      WorksMonday, WorksTuesday, WorksWednesday, WorksThursday, WorksFriday,
                                      WorksSaturday, WorksSunday, WeeklyWorkingHours, 
                                      MinStaffWeekday, MaxStaffWeekday, MinStaffWeekend, MaxStaffWeekend, CreatedBy)
                VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('code'),
                data.get('name'),
                data.get('startTime'),
                data.get('endTime'),
                data.get('durationHours'),
                data.get('colorCode', '#808080'),
                1 if data.get('worksMonday', True) else 0,
                1 if data.get('worksTuesday', True) else 0,
                1 if data.get('worksWednesday', True) else 0,
                1 if data.get('worksThursday', True) else 0,
                1 if data.get('worksFriday', True) else 0,
                1 if data.get('worksSaturday', False) else 0,
                1 if data.get('worksSunday', False) else 0,
                data.get('weeklyWorkingHours', 40.0),
                min_staff_weekday,
                max_staff_weekday,
                min_staff_weekend,
                max_staff_weekend,
                session.get('user_email', 'system')
            ))
            
            shift_type_id = cursor.lastrowid
            
            # Log audit entry
            changes = json.dumps(data, ensure_ascii=False)
            log_audit(conn, 'ShiftType', shift_type_id, 'Created', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'id': shift_type_id}), 201
            
        except Exception as e:
            app.logger.error(f"Create shift type error: {str(e)}")
            return jsonify({'error': f'Fehler beim Erstellen: {str(e)}'}), 500
    
    @app.route('/api/shifttypes/<int:id>', methods=['GET'])
    def get_shift_type(id):
        """Get single shift type by ID"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM ShiftTypes WHERE Id = ?", (id,))
        row = cursor.fetchone()
        
        if not row:
            conn.close()
            return jsonify({'error': 'Schichttyp nicht gefunden'}), 404
        
        shift_type = {
            'id': row['Id'],
            'code': row['Code'],
            'name': row['Name'],
            'startTime': row['StartTime'],
            'endTime': row['EndTime'],
            'durationHours': row['DurationHours'],
            'colorCode': row['ColorCode'],
            'isActive': bool(row['IsActive']),
            'worksMonday': bool(row['WorksMonday']),
            'worksTuesday': bool(row['WorksTuesday']),
            'worksWednesday': bool(row['WorksWednesday']),
            'worksThursday': bool(row['WorksThursday']),
            'worksFriday': bool(row['WorksFriday']),
            'worksSaturday': bool(row['WorksSaturday']),
            'worksSunday': bool(row['WorksSunday']),
            'weeklyWorkingHours': row['WeeklyWorkingHours'],
            'minStaffWeekday': row['MinStaffWeekday'],
            'maxStaffWeekday': row['MaxStaffWeekday'],
            'minStaffWeekend': row['MinStaffWeekend'],
            'maxStaffWeekend': row['MaxStaffWeekend']
        }
        
        conn.close()
        return jsonify(shift_type)
    
    @app.route('/api/shifttypes/<int:id>', methods=['PUT'])
    @require_role('Admin')
    def update_shift_type(id):
        """Update shift type (Admin only)"""
        try:
            data = request.get_json()
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if shift type exists
            cursor.execute("SELECT * FROM ShiftTypes WHERE Id = ?", (id,))
            old_row = cursor.fetchone()
            if not old_row:
                conn.close()
                return jsonify({'error': 'Schichttyp nicht gefunden'}), 404
            
            # Check if new code conflicts with existing
            if data.get('code') and data.get('code') != old_row['Code']:
                cursor.execute("SELECT Id FROM ShiftTypes WHERE Code = ? AND Id != ?", 
                             (data.get('code'), id))
                if cursor.fetchone():
                    conn.close()
                    return jsonify({'error': 'Schichtkürzel bereits vorhanden'}), 400
            
            # Validate staffing requirements using the helper function
            min_staff_weekday = data.get('minStaffWeekday', get_row_value(old_row, 'MinStaffWeekday', 3))
            max_staff_weekday = data.get('maxStaffWeekday', get_row_value(old_row, 'MaxStaffWeekday', 5))
            min_staff_weekend = data.get('minStaffWeekend', get_row_value(old_row, 'MinStaffWeekend', 2))
            max_staff_weekend = data.get('maxStaffWeekend', get_row_value(old_row, 'MaxStaffWeekend', 3))
            
            if min_staff_weekday > max_staff_weekday:
                conn.close()
                return jsonify({'error': 'Minimale Personalstärke an Wochentagen darf nicht größer sein als die maximale Personalstärke'}), 400
            if min_staff_weekend > max_staff_weekend:
                conn.close()
                return jsonify({'error': 'Minimale Personalstärke am Wochenende darf nicht größer sein als die maximale Personalstärke'}), 400
            
            # Update shift type
            cursor.execute("""
                UPDATE ShiftTypes 
                SET Code = ?, Name = ?, StartTime = ?, EndTime = ?, 
                    DurationHours = ?, ColorCode = ?, IsActive = ?,
                    WorksMonday = ?, WorksTuesday = ?, WorksWednesday = ?, WorksThursday = ?, WorksFriday = ?,
                    WorksSaturday = ?, WorksSunday = ?, WeeklyWorkingHours = ?,
                    MinStaffWeekday = ?, MaxStaffWeekday = ?, MinStaffWeekend = ?, MaxStaffWeekend = ?,
                    ModifiedAt = ?, ModifiedBy = ?
                WHERE Id = ?
            """, (
                data.get('code', old_row['Code']),
                data.get('name', old_row['Name']),
                data.get('startTime', old_row['StartTime']),
                data.get('endTime', old_row['EndTime']),
                data.get('durationHours', old_row['DurationHours']),
                data.get('colorCode', old_row['ColorCode']),
                1 if data.get('isActive', True) else 0,
                1 if data.get('worksMonday', get_row_value(old_row, 'WorksMonday', True)) else 0,
                1 if data.get('worksTuesday', get_row_value(old_row, 'WorksTuesday', True)) else 0,
                1 if data.get('worksWednesday', get_row_value(old_row, 'WorksWednesday', True)) else 0,
                1 if data.get('worksThursday', get_row_value(old_row, 'WorksThursday', True)) else 0,
                1 if data.get('worksFriday', get_row_value(old_row, 'WorksFriday', True)) else 0,
                1 if data.get('worksSaturday', get_row_value(old_row, 'WorksSaturday', False)) else 0,
                1 if data.get('worksSunday', get_row_value(old_row, 'WorksSunday', False)) else 0,
                data.get('weeklyWorkingHours', get_row_value(old_row, 'WeeklyWorkingHours', 40.0)),
                min_staff_weekday,
                max_staff_weekday,
                min_staff_weekend,
                max_staff_weekend,
                datetime.utcnow().isoformat(),
                session.get('user_email', 'system'),
                id
            ))
            
            # Log audit entry
            changes_dict = {}
            for field in ['code', 'name', 'startTime', 'endTime', 'durationHours', 'colorCode', 'isActive']:
                old_key = field.replace('Time', 'Time').replace('Hours', 'Hours')
                db_field = field[0].upper() + field[1:]
                if field == 'isActive':
                    db_field = 'IsActive'
                elif field == 'startTime':
                    db_field = 'StartTime'
                elif field == 'endTime':
                    db_field = 'EndTime'
                elif field == 'durationHours':
                    db_field = 'DurationHours'
                elif field == 'colorCode':
                    db_field = 'ColorCode'
                else:
                    db_field = field[0].upper() + field[1:]
                
                if field in data:
                    old_val = old_row[db_field]
                    new_val = data[field]
                    if field == 'isActive':
                        old_val = bool(old_val)
                    if old_val != new_val:
                        changes_dict[field] = {'old': old_val, 'new': new_val}
            
            if changes_dict:
                changes = json.dumps(changes_dict, ensure_ascii=False)
                log_audit(conn, 'ShiftType', id, 'Updated', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update shift type error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    @app.route('/api/shifttypes/<int:id>', methods=['DELETE'])
    @require_role('Admin')
    def delete_shift_type(id):
        """Delete shift type (Admin only)"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if shift type exists
            cursor.execute("SELECT Code, Name FROM ShiftTypes WHERE Id = ?", (id,))
            shift_row = cursor.fetchone()
            if not shift_row:
                conn.close()
                return jsonify({'error': 'Schichttyp nicht gefunden'}), 404
            
            # Check if shift type is used in assignments
            cursor.execute("SELECT COUNT(*) as count FROM ShiftAssignments WHERE ShiftTypeId = ?", (id,))
            assignment_count = cursor.fetchone()['count']
            
            if assignment_count > 0:
                conn.close()
                return jsonify({'error': f'Schichttyp wird in {assignment_count} Zuweisungen verwendet und kann nicht gelöscht werden'}), 400
            
            # Delete shift type (cascade will delete relationships and team assignments)
            cursor.execute("DELETE FROM ShiftTypes WHERE Id = ?", (id,))
            
            # Log audit entry
            changes = json.dumps({'code': shift_row['Code'], 'name': shift_row['Name']}, ensure_ascii=False)
            log_audit(conn, 'ShiftType', id, 'Deleted', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Delete shift type error: {str(e)}")
            return jsonify({'error': f'Fehler beim Löschen: {str(e)}'}), 500
    
    # Team-Shift Assignment endpoints
    @app.route('/api/shifttypes/<int:shift_id>/teams', methods=['GET'])
    def get_shift_type_teams(shift_id):
        """Get teams assigned to a shift type"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT t.Id, t.Name
            FROM Teams t
            INNER JOIN TeamShiftAssignments tsa ON t.Id = tsa.TeamId
            WHERE tsa.ShiftTypeId = ? AND t.IsVirtual = 0
            ORDER BY t.Name
        """, (shift_id,))
        
        teams = []
        for row in cursor.fetchall():
            teams.append({
                'id': row['Id'],
                'name': row['Name']
            })
        
        conn.close()
        return jsonify(teams)
    
    @app.route('/api/shifttypes/<int:shift_id>/teams', methods=['PUT'])
    @require_role('Admin')
    def update_shift_type_teams(shift_id):
        """Update teams assigned to a shift type (Admin only)"""
        try:
            data = request.get_json()
            team_ids = data.get('teamIds', [])
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if shift type exists
            cursor.execute("SELECT Id FROM ShiftTypes WHERE Id = ?", (shift_id,))
            if not cursor.fetchone():
                conn.close()
                return jsonify({'error': 'Schichttyp nicht gefunden'}), 404
            
            # Delete all existing assignments for this shift
            cursor.execute("DELETE FROM TeamShiftAssignments WHERE ShiftTypeId = ?", (shift_id,))
            
            # Insert new assignments
            for team_id in team_ids:
                cursor.execute("""
                    INSERT INTO TeamShiftAssignments (TeamId, ShiftTypeId, CreatedBy)
                    VALUES (?, ?, ?)
                """, (team_id, shift_id, session.get('user_email', 'system')))
            
            # Log audit entry
            changes = json.dumps({'shiftTypeId': shift_id, 'teamIds': team_ids}, ensure_ascii=False)
            log_audit(conn, 'TeamShiftAssignment', shift_id, 'Updated', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update shift type teams error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    @app.route('/api/teams/<int:team_id>/shifttypes', methods=['GET'])
    def get_team_shift_types(team_id):
        """Get shift types assigned to a team"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT st.Id, st.Code, st.Name, st.ColorCode
            FROM ShiftTypes st
            INNER JOIN TeamShiftAssignments tsa ON st.Id = tsa.ShiftTypeId
            WHERE tsa.TeamId = ? AND st.IsActive = 1
            ORDER BY st.Code
        """, (team_id,))
        
        shift_types = []
        for row in cursor.fetchall():
            shift_types.append({
                'id': row['Id'],
                'code': row['Code'],
                'name': row['Name'],
                'colorCode': row['ColorCode']
            })
        
        conn.close()
        return jsonify(shift_types)
    
    @app.route('/api/teams/<int:team_id>/shifttypes', methods=['PUT'])
    @require_role('Admin')
    def update_team_shift_types(team_id):
        """Update shift types assigned to a team (Admin only)"""
        try:
            data = request.get_json()
            shift_type_ids = data.get('shiftTypeIds', [])
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if team exists
            cursor.execute("SELECT Id FROM Teams WHERE Id = ?", (team_id,))
            if not cursor.fetchone():
                conn.close()
                return jsonify({'error': 'Team nicht gefunden'}), 404
            
            # Delete all existing assignments for this team
            cursor.execute("DELETE FROM TeamShiftAssignments WHERE TeamId = ?", (team_id,))
            
            # Insert new assignments
            for shift_type_id in shift_type_ids:
                cursor.execute("""
                    INSERT INTO TeamShiftAssignments (TeamId, ShiftTypeId, CreatedBy)
                    VALUES (?, ?, ?)
                """, (team_id, shift_type_id, session.get('user_email', 'system')))
            
            # Log audit entry
            changes = json.dumps({'teamId': team_id, 'shiftTypeIds': shift_type_ids}, ensure_ascii=False)
            log_audit(conn, 'TeamShiftAssignment', team_id, 'Updated', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update team shift types error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    # Shift Type Relationships endpoints
    @app.route('/api/shifttypes/<int:shift_id>/relationships', methods=['GET'])
    def get_shift_type_relationships(shift_id):
        """Get related shift types for a shift"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT st.Id, st.Code, st.Name, st.ColorCode, str.DisplayOrder
            FROM ShiftTypes st
            INNER JOIN ShiftTypeRelationships str ON st.Id = str.RelatedShiftTypeId
            WHERE str.ShiftTypeId = ?
            ORDER BY str.DisplayOrder
        """, (shift_id,))
        
        relationships = []
        for row in cursor.fetchall():
            relationships.append({
                'id': row['Id'],
                'code': row['Code'],
                'name': row['Name'],
                'colorCode': row['ColorCode'],
                'displayOrder': row['DisplayOrder']
            })
        
        conn.close()
        return jsonify(relationships)
    
    @app.route('/api/shifttypes/<int:shift_id>/relationships', methods=['PUT'])
    @require_role('Admin')
    def update_shift_type_relationships(shift_id):
        """Update related shift types (Admin only)"""
        try:
            data = request.get_json()
            relationships = data.get('relationships', [])  # [{shiftTypeId, displayOrder}]
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if shift type exists
            cursor.execute("SELECT Id FROM ShiftTypes WHERE Id = ?", (shift_id,))
            if not cursor.fetchone():
                conn.close()
                return jsonify({'error': 'Schichttyp nicht gefunden'}), 404
            
            # Delete all existing relationships for this shift
            cursor.execute("DELETE FROM ShiftTypeRelationships WHERE ShiftTypeId = ?", (shift_id,))
            
            # Insert new relationships
            for rel in relationships:
                cursor.execute("""
                    INSERT INTO ShiftTypeRelationships 
                    (ShiftTypeId, RelatedShiftTypeId, DisplayOrder, CreatedBy)
                    VALUES (?, ?, ?, ?)
                """, (shift_id, rel['shiftTypeId'], rel['displayOrder'], session.get('user_email', 'system')))
            
            # Log audit entry
            changes = json.dumps({'shiftTypeId': shift_id, 'relationships': relationships}, ensure_ascii=False)
            log_audit(conn, 'ShiftTypeRelationship', shift_id, 'Updated', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update shift type relationships error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    # ============================================================================
    # GLOBAL SETTINGS ENDPOINTS
    # ============================================================================
    
    @app.route('/api/settings/global', methods=['GET'])
    def get_global_settings():
        """Get global shift planning settings"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT * FROM GlobalSettings WHERE Id = 1")
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                # Return defaults if not found
                return jsonify({
                    'maxConsecutiveShifts': 6,
                    'maxConsecutiveNightShifts': 3,
                    'minRestHoursBetweenShifts': 11
                })
            
            return jsonify({
                'maxConsecutiveShifts': row['MaxConsecutiveShifts'],
                'maxConsecutiveNightShifts': row['MaxConsecutiveNightShifts'],
                'minRestHoursBetweenShifts': row['MinRestHoursBetweenShifts'],
                'modifiedAt': row['ModifiedAt'],
                'modifiedBy': row['ModifiedBy']
            })
            
        except Exception as e:
            app.logger.error(f"Get global settings error: {str(e)}")
            return jsonify({'error': f'Fehler beim Laden: {str(e)}'}), 500
    
    @app.route('/api/settings/global', methods=['PUT'])
    @require_role('Admin')
    def update_global_settings():
        """Update global shift planning settings (Admin only)"""
        try:
            data = request.get_json()
            
            max_consecutive_shifts = data.get('maxConsecutiveShifts', 6)
            max_consecutive_night_shifts = data.get('maxConsecutiveNightShifts', 3)
            min_rest_hours = data.get('minRestHoursBetweenShifts', 11)
            
            # Validation
            if max_consecutive_shifts < 1 or max_consecutive_shifts > 10:
                return jsonify({'error': 'Maximale aufeinanderfolgende Schichten muss zwischen 1 und 10 liegen'}), 400
            if max_consecutive_night_shifts < 1 or max_consecutive_night_shifts > max_consecutive_shifts:
                return jsonify({'error': 'Maximale Nachtschichten darf nicht größer sein als maximale Schichten'}), 400
            if min_rest_hours < 8 or min_rest_hours > 24:
                return jsonify({'error': 'Mindest-Ruhezeit muss zwischen 8 und 24 Stunden liegen'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Update or insert settings
            cursor.execute("""
                INSERT INTO GlobalSettings 
                (Id, MaxConsecutiveShifts, MaxConsecutiveNightShifts, MinRestHoursBetweenShifts, ModifiedAt, ModifiedBy)
                VALUES (1, ?, ?, ?, ?, ?)
                ON CONFLICT(Id) DO UPDATE SET
                    MaxConsecutiveShifts = excluded.MaxConsecutiveShifts,
                    MaxConsecutiveNightShifts = excluded.MaxConsecutiveNightShifts,
                    MinRestHoursBetweenShifts = excluded.MinRestHoursBetweenShifts,
                    ModifiedAt = excluded.ModifiedAt,
                    ModifiedBy = excluded.ModifiedBy
            """, (
                max_consecutive_shifts,
                max_consecutive_night_shifts,
                min_rest_hours,
                datetime.utcnow().isoformat(),
                session.get('user_email', 'system')
            ))
            
            # Log audit entry
            changes = json.dumps({
                'maxConsecutiveShifts': max_consecutive_shifts,
                'maxConsecutiveNightShifts': max_consecutive_night_shifts,
                'minRestHoursBetweenShifts': min_rest_hours
            }, ensure_ascii=False)
            log_audit(conn, 'GlobalSettings', 1, 'Updated', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update global settings error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    # ============================================================================
    # SHIFT PLANNING ENDPOINTS
    # ============================================================================
    
    @app.route('/api/shifts/schedule', methods=['GET'])
    def get_schedule():
        """Get shift schedule for a date range"""
        start_date_str = request.args.get('startDate')
        end_date_str = request.args.get('endDate')
        view = request.args.get('view', 'week')
        
        if not start_date_str:
            return jsonify({'error': 'startDate is required'}), 400
        
        # Validate and parse dates
        try:
            start_date = date.fromisoformat(start_date_str)
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid startDate format. Use YYYY-MM-DD'}), 400
        
        # Calculate end date based on view
        if not end_date_str:
            if view == 'week':
                end_date = start_date + timedelta(days=6)
            elif view == 'month':
                # Get last day of month
                if start_date.month == 12:
                    end_date = date(start_date.year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = date(start_date.year, start_date.month + 1, 1) - timedelta(days=1)
            elif view == 'year':
                end_date = date(start_date.year, 12, 31)
            else:
                end_date = start_date + timedelta(days=30)
        else:
            try:
                end_date = date.fromisoformat(end_date_str)
            except (ValueError, TypeError):
                return jsonify({'error': 'Invalid endDate format. Use YYYY-MM-DD'}), 400
        
        conn = None
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if user is admin (can see unapproved plans)
            user_roles = session.get('user_roles', [])
            is_admin = 'Admin' in user_roles
            
            # Get approved months if user is not admin
            approved_months = set()
            if not is_admin:
                cursor.execute("""
                    SELECT Year, Month FROM ShiftPlanApprovals
                    WHERE IsApproved = 1
                """)
                for row in cursor.fetchall():
                    approved_months.add((row['Year'], row['Month']))
            
            # Get assignments
            cursor.execute("""
                SELECT sa.*, e.Vorname, e.Name, e.TeamId, e.IsSpringer,
                       st.Code, st.Name as ShiftName, st.ColorCode
                FROM ShiftAssignments sa
                JOIN Employees e ON sa.EmployeeId = e.Id
                JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
                WHERE sa.Date >= ? AND sa.Date <= ?
                ORDER BY sa.Date, e.TeamId, e.Name, e.Vorname
            """, (start_date.isoformat(), end_date.isoformat()))
            
            assignments = []
            for row in cursor.fetchall():
                assignment_date = date.fromisoformat(row['Date'])
                year_month = (assignment_date.year, assignment_date.month)
                
                # Filter based on approval status for non-admin users
                if not is_admin and year_month not in approved_months:
                    # Skip unapproved plans for regular users
                    continue
                
                assignments.append({
                    'id': row['Id'],
                    'employeeId': row['EmployeeId'],
                    'employeeName': f"{row['Vorname']} {row['Name']}",
                    'teamId': row['TeamId'],
                    'isSpringer': bool(row['IsSpringer']),
                    'shiftTypeId': row['ShiftTypeId'],
                    'shiftCode': row['Code'],
                    'shiftName': row['ShiftName'],
                    'colorCode': row['ColorCode'],
                    'date': row['Date'],
                    'isManual': bool(row['IsManual']),
                    'isSpringerAssignment': bool(row['IsSpringerAssignment']),
                    'isFixed': bool(row['IsFixed']),
                    'notes': row['Notes']
                })
            
            # Get absences from Absences table
            cursor.execute("""
                SELECT a.*, e.Vorname, e.Name, e.TeamId
                FROM Absences a
                JOIN Employees e ON a.EmployeeId = e.Id
                WHERE (a.StartDate <= ? AND a.EndDate >= ?)
                   OR (a.StartDate >= ? AND a.StartDate <= ?)
            """, (end_date.isoformat(), start_date.isoformat(), 
                  start_date.isoformat(), end_date.isoformat()))
            
            absences = []
            for row in cursor.fetchall():
                type_index = row['Type']
                absences.append({
                    'id': row['Id'],
                    'employeeId': row['EmployeeId'],
                    'employeeName': f"{row['Vorname']} {row['Name']}",
                    'teamId': row['TeamId'],
                    'type': ['', 'Krank', 'Urlaub', 'Lehrgang'][type_index],
                    'status': 'Genehmigt' if type_index == 2 else None,  # Only Urlaub type has status
                    'startDate': row['StartDate'],
                    'endDate': row['EndDate'],
                    'notes': row['Notes']
                })
            
            # Also get vacation requests (all statuses) and add them as absences
            cursor.execute("""
                SELECT vr.Id, vr.EmployeeId, vr.StartDate, vr.EndDate, vr.Notes, vr.Status,
                       e.Vorname, e.Name, e.TeamId
                FROM VacationRequests vr
                JOIN Employees e ON vr.EmployeeId = e.Id
                WHERE ((vr.StartDate <= ? AND vr.EndDate >= ?)
                   OR (vr.StartDate >= ? AND vr.StartDate <= ?))
            """, (end_date.isoformat(), start_date.isoformat(),
                  start_date.isoformat(), end_date.isoformat()))
            
            vacation_id_offset = 10000  # Offset to avoid ID conflicts
            for row in cursor.fetchall():
                # Determine the type label based on status
                if row['Status'] == 'Genehmigt':
                    type_label = 'Urlaub'
                    notes = row['Notes'] or 'Genehmigter Urlaub'
                elif row['Status'] == 'InBearbeitung':
                    type_label = 'Urlaub (in Genehmigung)'
                    notes = row['Notes'] or 'Urlaubsantrag in Bearbeitung'
                else:  # Abgelehnt
                    type_label = 'Urlaub (abgelehnt)'
                    notes = row['Notes'] or 'Urlaubsantrag abgelehnt'
                
                absences.append({
                    'id': vacation_id_offset + row['Id'],
                    'employeeId': row['EmployeeId'],
                    'employeeName': f"{row['Vorname']} {row['Name']}",
                    'teamId': row['TeamId'],
                    'type': type_label,
                    'status': row['Status'],  # Include status for color-coding
                    'startDate': row['StartDate'],
                    'endDate': row['EndDate'],
                    'notes': notes
                })
            
            # Get vacation periods (Ferienzeiten) that overlap with the date range
            cursor.execute("""
                SELECT Id, Name, StartDate, EndDate, ColorCode
                FROM VacationPeriods
                WHERE (StartDate <= ? AND EndDate >= ?)
                   OR (StartDate >= ? AND StartDate <= ?)
                ORDER BY StartDate
            """, (end_date.isoformat(), start_date.isoformat(),
                  start_date.isoformat(), end_date.isoformat()))
            
            vacation_periods = []
            for row in cursor.fetchall():
                vacation_periods.append({
                    'id': row['Id'],
                    'name': row['Name'],
                    'startDate': row['StartDate'],
                    'endDate': row['EndDate'],
                    'colorCode': row['ColorCode'] or '#E8F5E9'
                })
            
            return jsonify({
                'startDate': start_date.isoformat(),
                'endDate': end_date.isoformat(),
                'assignments': assignments,
                'absences': absences,
                'vacationPeriods': vacation_periods
            })
            
        except Exception as e:
            return jsonify({'error': f'Database error: {str(e)}'}), 500
        finally:
            if conn:
                conn.close()
    
    @app.route('/api/shifts/plan', methods=['POST'])
    @require_role('Admin')
    def plan_shifts():
        """Automatic shift planning using OR-Tools (Admin only) - Monthly planning only"""
        start_date_str = request.args.get('startDate')
        end_date_str = request.args.get('endDate')
        force = request.args.get('force', 'false').lower() == 'true'
        
        if not start_date_str or not end_date_str:
            return jsonify({'error': 'startDate and endDate are required'}), 400
        
        try:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
            
            # Validate that planning is for a complete single month
            is_valid, error_msg = validate_monthly_date_range(start_date, end_date)
            if not is_valid:
                return jsonify({'error': error_msg}), 400
            
            # Load data
            employees, teams, absences, shift_types = load_from_database(db.db_path)
            
            # Create model
            planning_model = create_shift_planning_model(
                employees, teams, start_date, end_date, absences, shift_types=shift_types
            )
            
            # Solve
            result = solve_shift_planning(planning_model, time_limit_seconds=300)
            
            if not result:
                # Get diagnostic information to help user understand the issue
                diagnostics = get_infeasibility_diagnostics(planning_model)
                
                # Build helpful error message
                error_details = []
                error_details.append(f"Planning für {start_date.strftime('%d.%m.%Y')} bis {end_date.strftime('%d.%m.%Y')} nicht möglich.")
                error_details.append(f"Mitarbeiter: {diagnostics['available_employees']} verfügbar / {diagnostics['total_employees']} gesamt")
                
                if diagnostics['absent_employees'] > 0:
                    error_details.append(f"Abwesend: {diagnostics['absent_employees']} Mitarbeiter")
                
                # Add specific issues
                if diagnostics['potential_issues']:
                    error_details.append("")
                    error_details.append("Mögliche Probleme:")
                    for issue in diagnostics['potential_issues'][:3]:  # Show max 3 issues
                        error_details.append(f"• {issue}")
                
                # Add staffing analysis for shifts with issues
                problem_shifts = [shift for shift, data in diagnostics['shift_analysis'].items() 
                                 if not data['is_feasible']]
                if problem_shifts:
                    error_details.append("")
                    error_details.append("Schichtbesetzung:")
                    for shift in problem_shifts[:3]:  # Show max 3 problematic shifts
                        data = diagnostics['shift_analysis'][shift]
                        error_details.append(f"• {shift}: {data['eligible_employees']} verfügbar, {data['min_required']} benötigt")
                
                error_message = "\n".join(error_details)
                
                return jsonify({
                    'error': 'No solution found',
                    'details': error_message,
                    'diagnostics': {
                        'total_employees': diagnostics['total_employees'],
                        'available_employees': diagnostics['available_employees'],
                        'absent_employees': diagnostics['absent_employees'],
                        'potential_issues': diagnostics['potential_issues']
                    }
                }), 500
            
            assignments, special_functions, complete_schedule = result
            
            # Save to database
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Delete existing non-fixed assignments if force
            if force:
                cursor.execute("""
                    DELETE FROM ShiftAssignments 
                    WHERE Date >= ? AND Date <= ? AND IsFixed = 0
                """, (start_date.isoformat(), end_date.isoformat()))
            
            # Insert new assignments
            for assignment in assignments:
                cursor.execute("""
                    INSERT INTO ShiftAssignments 
                    (EmployeeId, ShiftTypeId, Date, IsManual, IsFixed, CreatedAt, CreatedBy)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    assignment.employee_id,
                    assignment.shift_type_id,
                    assignment.date.isoformat(),
                    0,
                    0,
                    datetime.utcnow().isoformat(),
                    "Python-OR-Tools"
                ))
            
            # Insert special functions (TD assignments)
            # Get TD shift type ID
            cursor.execute("SELECT Id FROM ShiftTypes WHERE Code = 'TD'")
            td_row = cursor.fetchone()
            if td_row:
                td_shift_type_id = td_row[0]
                for (emp_id, date_obj), function_code in special_functions.items():
                    if function_code == "TD":
                        cursor.execute("""
                            INSERT INTO ShiftAssignments 
                            (EmployeeId, ShiftTypeId, Date, IsManual, IsSpringerAssignment, IsFixed, CreatedAt, CreatedBy)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            emp_id,
                            td_shift_type_id,
                            date_obj.isoformat(),
                            0,
                            0,
                            0,
                            datetime.utcnow().isoformat(),
                            "Python-OR-Tools"
                        ))
            
            # Create or update approval record for this month (not approved by default)
            cursor.execute("""
                INSERT INTO ShiftPlanApprovals (Year, Month, IsApproved, CreatedAt)
                VALUES (?, ?, 0, ?)
                ON CONFLICT(Year, Month) DO UPDATE SET
                    IsApproved = 0,
                    ApprovedAt = NULL,
                    ApprovedBy = NULL,
                    ApprovedByName = NULL
            """, (start_date.year, start_date.month, datetime.utcnow().isoformat()))
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': f'Successfully planned {len(assignments)} shifts for {start_date.strftime("%B %Y")}. Plan must be approved before visible to regular users.',
                'assignmentsCount': len(assignments),
                'specialFunctionsCount': len(special_functions),
                'year': start_date.year,
                'month': start_date.month
            })
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/shifts/plan/approvals', methods=['GET'])
    @require_role('Admin')
    def get_plan_approvals():
        """Get all shift plan approvals (Admin only)"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT * FROM ShiftPlanApprovals
                ORDER BY Year DESC, Month DESC
            """)
            
            approvals = []
            for row in cursor.fetchall():
                approvals.append({
                    'id': row['Id'],
                    'year': row['Year'],
                    'month': row['Month'],
                    'isApproved': bool(row['IsApproved']),
                    'approvedAt': row['ApprovedAt'],
                    'approvedBy': row['ApprovedBy'],
                    'approvedByName': row['ApprovedByName'],
                    'notes': row['Notes'],
                    'createdAt': row['CreatedAt']
                })
            
            conn.close()
            return jsonify(approvals)
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/shifts/plan/approvals/<int:year>/<int:month>', methods=['GET'])
    def get_plan_approval_status(year, month):
        """Get approval status for a specific month"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT * FROM ShiftPlanApprovals
                WHERE Year = ? AND Month = ?
            """, (year, month))
            
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                return jsonify({
                    'year': year,
                    'month': month,
                    'isApproved': False,
                    'exists': False
                })
            
            return jsonify({
                'id': row['Id'],
                'year': row['Year'],
                'month': row['Month'],
                'isApproved': bool(row['IsApproved']),
                'approvedAt': row['ApprovedAt'],
                'approvedBy': row['ApprovedBy'],
                'approvedByName': row['ApprovedByName'],
                'notes': row['Notes'],
                'createdAt': row['CreatedAt'],
                'exists': True
            })
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/shifts/plan/approvals/<int:year>/<int:month>', methods=['PUT'])
    @require_role('Admin')
    def approve_plan(year, month):
        """Approve or unapprove a shift plan for a specific month (Admin only)"""
        try:
            data = request.get_json()
            is_approved = data.get('isApproved', True)
            notes = data.get('notes', '')
            
            # Get current user info
            user_id = session.get('user_id')
            user_name = session.get('user_fullname', 'Unknown Admin')
            
            if not user_id:
                return jsonify({'error': 'User not authenticated'}), 401
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Update or insert approval record
            if is_approved:
                cursor.execute("""
                    INSERT INTO ShiftPlanApprovals (Year, Month, IsApproved, ApprovedAt, ApprovedBy, ApprovedByName, Notes, CreatedAt)
                    VALUES (?, ?, 1, ?, ?, ?, ?, ?)
                    ON CONFLICT(Year, Month) DO UPDATE SET
                        IsApproved = 1,
                        ApprovedAt = ?,
                        ApprovedBy = ?,
                        ApprovedByName = ?,
                        Notes = ?
                """, (year, month, datetime.utcnow().isoformat(), user_id, user_name, notes,
                      datetime.utcnow().isoformat(),
                      datetime.utcnow().isoformat(), user_id, user_name, notes))
            else:
                cursor.execute("""
                    INSERT INTO ShiftPlanApprovals (Year, Month, IsApproved, CreatedAt)
                    VALUES (?, ?, 0, ?)
                    ON CONFLICT(Year, Month) DO UPDATE SET
                        IsApproved = 0,
                        ApprovedAt = NULL,
                        ApprovedBy = NULL,
                        ApprovedByName = NULL,
                        Notes = ?
                """, (year, month, datetime.utcnow().isoformat(), notes))
            
            conn.commit()
            conn.close()
            
            action = 'freigegeben' if is_approved else 'Freigabe aufgehoben'
            return jsonify({
                'success': True,
                'message': f'Dienstplan für {month:02d}/{year} wurde {action}.'
            })
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/shifts/assignments/<int:id>', methods=['PUT'])
    @require_role('Admin')
    def update_shift_assignment(id):
        """Update a shift assignment (manual edit)"""
        try:
            data = request.get_json()
            
            # Validate data types
            try:
                employee_id = int(data.get('employeeId'))
                shift_type_id = int(data.get('shiftTypeId'))
                assignment_date = date.fromisoformat(data.get('date'))
            except (ValueError, TypeError) as e:
                return jsonify({'error': f'Ungültige Daten: {str(e)}'}), 400
            
            conn = None
            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                
                # Check if assignment exists and get old values for audit
                cursor.execute("""
                    SELECT EmployeeId, ShiftTypeId, Date, IsFixed, Notes 
                    FROM ShiftAssignments WHERE Id = ?
                """, (id,))
                old_row = cursor.fetchone()
                if not old_row:
                    return jsonify({'error': 'Schichtzuweisung nicht gefunden'}), 404
                
                # Update assignment
                cursor.execute("""
                    UPDATE ShiftAssignments 
                    SET EmployeeId = ?, ShiftTypeId = ?, Date = ?, 
                        IsManual = 1, IsFixed = ?, Notes = ?,
                        ModifiedAt = ?, ModifiedBy = ?
                    WHERE Id = ?
                """, (
                    employee_id,
                    shift_type_id,
                    assignment_date.isoformat(),
                    1 if data.get('isFixed') else 0,
                    data.get('notes'),
                    datetime.utcnow().isoformat(),
                    session.get('user_email'),
                    id
                ))
                
                # Log audit entry with changes
                changes_dict = {}
                if old_row['EmployeeId'] != employee_id:
                    changes_dict['employeeId'] = {'old': old_row['EmployeeId'], 'new': employee_id}
                if old_row['ShiftTypeId'] != shift_type_id:
                    changes_dict['shiftTypeId'] = {'old': old_row['ShiftTypeId'], 'new': shift_type_id}
                if old_row['Date'] != assignment_date.isoformat():
                    changes_dict['date'] = {'old': old_row['Date'], 'new': assignment_date.isoformat()}
                new_is_fixed = 1 if data.get('isFixed') else 0
                if old_row['IsFixed'] != new_is_fixed:
                    changes_dict['isFixed'] = {'old': bool(old_row['IsFixed']), 'new': bool(new_is_fixed)}
                if old_row['Notes'] != data.get('notes'):
                    changes_dict['notes'] = {'old': old_row['Notes'], 'new': data.get('notes')}
                
                if changes_dict:
                    changes = json.dumps(changes_dict, ensure_ascii=False)
                    log_audit(conn, 'ShiftAssignment', id, 'Updated', changes)
                
                conn.commit()
                
                return jsonify({'success': True})
                
            finally:
                if conn:
                    conn.close()
            
        except Exception as e:
            app.logger.error(f"Update shift assignment error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    @app.route('/api/shifts/assignments', methods=['POST'])
    @require_role('Admin')
    def create_shift_assignment():
        """Create a shift assignment manually"""
        try:
            data = request.get_json()
            
            # Validate required fields
            if not data.get('employeeId') or not data.get('shiftTypeId') or not data.get('date'):
                return jsonify({'error': 'EmployeeId, ShiftTypeId und Date sind erforderlich'}), 400
            
            # Validate data types
            try:
                employee_id = int(data.get('employeeId'))
                shift_type_id = int(data.get('shiftTypeId'))
                assignment_date = date.fromisoformat(data.get('date'))
            except (ValueError, TypeError) as e:
                return jsonify({'error': f'Ungültige Daten: {str(e)}'}), 400
            
            conn = None
            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                
                # Check for existing assignment (same employee, date, shift type)
                cursor.execute("""
                    SELECT Id FROM ShiftAssignments 
                    WHERE EmployeeId = ? AND Date = ? AND ShiftTypeId = ?
                """, (employee_id, assignment_date.isoformat(), shift_type_id))
                
                if cursor.fetchone():
                    return jsonify({'error': 'Diese Schichtzuweisung existiert bereits'}), 400
                
                # Create assignment
                cursor.execute("""
                    INSERT INTO ShiftAssignments 
                    (EmployeeId, ShiftTypeId, Date, IsManual, IsFixed, Notes, CreatedAt, CreatedBy)
                    VALUES (?, ?, ?, 1, ?, ?, ?, ?)
                """, (
                    employee_id,
                    shift_type_id,
                    assignment_date.isoformat(),
                    1 if data.get('isFixed') else 0,
                    data.get('notes'),
                    datetime.utcnow().isoformat(),
                    session.get('user_email')
                ))
                
                assignment_id = cursor.lastrowid
                
                # Log audit entry
                changes = json.dumps({
                    'employeeId': employee_id,
                    'shiftTypeId': shift_type_id,
                    'date': assignment_date.isoformat(),
                    'isFixed': data.get('isFixed'),
                    'notes': data.get('notes')
                }, ensure_ascii=False)
                log_audit(conn, 'ShiftAssignment', assignment_id, 'Created', changes)
                
                conn.commit()
                
                return jsonify({'success': True, 'id': assignment_id}), 201
                
            finally:
                if conn:
                    conn.close()
            
        except Exception as e:
            app.logger.error(f"Create shift assignment error: {str(e)}")
            return jsonify({'error': f'Fehler beim Erstellen: {str(e)}'}), 500
    
    @app.route('/api/shifts/assignments/<int:id>', methods=['DELETE'])
    @require_role('Admin')
    def delete_shift_assignment(id):
        """Delete a shift assignment"""
        try:
            conn = None
            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                
                # Check if assignment exists and get info for audit
                cursor.execute("""
                    SELECT EmployeeId, ShiftTypeId, Date, IsFixed 
                    FROM ShiftAssignments WHERE Id = ?
                """, (id,))
                row = cursor.fetchone()
                
                if not row:
                    return jsonify({'error': 'Schichtzuweisung nicht gefunden'}), 404
                
                # Warn if trying to delete a fixed assignment
                if row['IsFixed']:
                    return jsonify({'error': 'Fixierte Schichtzuweisungen können nicht gelöscht werden. Bitte erst entsperren.'}), 400
                
                # Delete assignment
                cursor.execute("DELETE FROM ShiftAssignments WHERE Id = ?", (id,))
                
                # Log audit entry
                changes = json.dumps({
                    'employeeId': row['EmployeeId'],
                    'shiftTypeId': row['ShiftTypeId'],
                    'date': row['Date']
                }, ensure_ascii=False)
                log_audit(conn, 'ShiftAssignment', id, 'Deleted', changes)
                
                conn.commit()
                
                return jsonify({'success': True})
                
            finally:
                if conn:
                    conn.close()
            
        except Exception as e:
            app.logger.error(f"Delete shift assignment error: {str(e)}")
            return jsonify({'error': f'Fehler beim Löschen: {str(e)}'}), 500
    
    @app.route('/api/shifts/assignments/bulk', methods=['PUT'])
    @require_role('Admin')
    def bulk_update_shift_assignments():
        """Bulk update multiple shift assignments"""
        try:
            data = request.get_json()
            
            # Validate required fields
            if not data.get('shiftIds') or not isinstance(data.get('shiftIds'), list):
                return jsonify({'error': 'ShiftIds array ist erforderlich'}), 400
            
            if not data.get('changes') or not isinstance(data.get('changes'), dict):
                return jsonify({'error': 'Changes object ist erforderlich'}), 400
            
            shift_ids = data['shiftIds']
            changes = data['changes']
            
            if len(shift_ids) == 0:
                return jsonify({'error': 'Keine Schichten zum Aktualisieren ausgewählt'}), 400
            
            # Validate that at least one field is being changed
            allowed_fields = {'employeeId', 'shiftTypeId', 'isFixed', 'notes'}
            if not any(key in changes for key in allowed_fields):
                return jsonify({'error': 'Mindestens ein Feld muss geändert werden'}), 400
            
            # Validate that only allowed fields are present
            invalid_fields = set(changes.keys()) - allowed_fields
            if invalid_fields:
                return jsonify({'error': f'Ungültige Felder: {", ".join(invalid_fields)}'}), 400
            
            conn = None
            updated_count = 0
            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                
                # Whitelist for allowed column names to prevent SQL injection
                ALLOWED_COLUMNS = {
                    'employeeId': 'EmployeeId',
                    'shiftTypeId': 'ShiftTypeId',
                    'isFixed': 'IsFixed',
                    'notes': 'Notes'
                }
                
                # Process each shift
                for shift_id in shift_ids:
                    # Verify shift exists
                    cursor.execute("SELECT Id FROM ShiftAssignments WHERE Id = ?", (shift_id,))
                    if not cursor.fetchone():
                        app.logger.warning(f"Shift {shift_id} not found, skipping")
                        continue
                    
                    # Build UPDATE query dynamically based on changes
                    update_fields = []
                    update_values = []
                    
                    if 'employeeId' in changes:
                        update_fields.append(f"{ALLOWED_COLUMNS['employeeId']} = ?")
                        update_values.append(changes['employeeId'])
                    
                    if 'shiftTypeId' in changes:
                        update_fields.append(f"{ALLOWED_COLUMNS['shiftTypeId']} = ?")
                        update_values.append(changes['shiftTypeId'])
                    
                    if 'isFixed' in changes:
                        update_fields.append(f"{ALLOWED_COLUMNS['isFixed']} = ?")
                        update_values.append(1 if changes['isFixed'] else 0)
                    
                    if 'notes' in changes:
                        # Append notes instead of replacing
                        cursor.execute("SELECT Notes FROM ShiftAssignments WHERE Id = ?", (shift_id,))
                        row = cursor.fetchone()
                        existing_notes = row['Notes'] if row and row['Notes'] else ''
                        new_notes = existing_notes
                        if existing_notes:
                            new_notes += '\n' + changes['notes']
                        else:
                            new_notes = changes['notes']
                        update_fields.append(f"{ALLOWED_COLUMNS['notes']} = ?")
                        update_values.append(new_notes)
                    
                    # Always update modification timestamp and user
                    update_fields.append("ModifiedAt = ?")
                    update_fields.append("ModifiedBy = ?")
                    update_values.append(datetime.utcnow().isoformat())
                    update_values.append(session.get('user_email'))
                    
                    # Add shift ID as last parameter
                    update_values.append(shift_id)
                    
                    # Execute update - fields are from whitelist, safe to use in f-string
                    update_query = f"""
                        UPDATE ShiftAssignments 
                        SET {', '.join(update_fields)}
                        WHERE Id = ?
                    """
                    
                    cursor.execute(update_query, update_values)
                    
                    # Log audit entry
                    changes_json = json.dumps(changes, ensure_ascii=False)
                    log_audit(conn, 'ShiftAssignment', shift_id, 'BulkUpdate', changes_json)
                    
                    updated_count += 1
                
                conn.commit()
                
                return jsonify({
                    'success': True,
                    'updated': updated_count,
                    'total': len(shift_ids)
                })
                
            finally:
                if conn:
                    conn.close()
            
        except ValueError as e:
            app.logger.error(f"Bulk update validation error: {str(e)}")
            return jsonify({'error': f'Validierungsfehler: {str(e)}'}), 400
        except Exception as e:
            app.logger.error(f"Bulk update shift assignments error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    @app.route('/api/shifts/assignments/<int:id>/toggle-fixed', methods=['PUT'])
    @require_role('Admin')
    def toggle_fixed_assignment(id):
        """Toggle the IsFixed flag on an assignment (lock/unlock)"""
        try:
            conn = None
            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                
                # Check if assignment exists
                cursor.execute("SELECT Id, IsFixed FROM ShiftAssignments WHERE Id = ?", (id,))
                row = cursor.fetchone()
                
                if not row:
                    return jsonify({'error': 'Schichtzuweisung nicht gefunden'}), 404
                
                # Toggle fixed status
                new_fixed_status = 0 if row['IsFixed'] else 1
                
                cursor.execute("""
                    UPDATE ShiftAssignments 
                    SET IsFixed = ?, ModifiedAt = ?, ModifiedBy = ?
                    WHERE Id = ?
                """, (
                    new_fixed_status,
                    datetime.utcnow().isoformat(),
                    session.get('user_email'),
                    id
                ))
                
                conn.commit()
                
                return jsonify({
                    'success': True,
                    'isFixed': bool(new_fixed_status)
                })
                
            finally:
                if conn:
                    conn.close()
            
        except Exception as e:
            app.logger.error(f"Toggle fixed assignment error: {str(e)}")
            return jsonify({'error': f'Fehler beim Sperren/Entsperren: {str(e)}'}), 500
    
    @app.route('/api/shifts/export/csv', methods=['GET'])
    def export_schedule_csv():
        """Export schedule to CSV format"""
        start_date_str = request.args.get('startDate')
        end_date_str = request.args.get('endDate')
        
        if not start_date_str or not end_date_str:
            return jsonify({'error': 'startDate and endDate are required'}), 400
        
        try:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Get assignments
            cursor.execute("""
                SELECT sa.Date, e.Vorname, e.Name, e.Personalnummer, 
                       t.Name as TeamName, st.Code, st.Name as ShiftName
                FROM ShiftAssignments sa
                JOIN Employees e ON sa.EmployeeId = e.Id
                LEFT JOIN Teams t ON e.TeamId = t.Id
                JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
                WHERE sa.Date >= ? AND sa.Date <= ?
                ORDER BY sa.Date, t.Name, e.Name, e.Vorname
            """, (start_date.isoformat(), end_date.isoformat()))
            
            # Build CSV
            import io
            output = io.StringIO()
            output.write("Datum,Team,Mitarbeiter,Personalnummer,Schichttyp,Schichtname\n")
            
            for row in cursor.fetchall():
                team_name = row['TeamName'] or 'Ohne Team'
                output.write(f"{row['Date']},{team_name},{row['Vorname']} {row['Name']},{row['Personalnummer']},{row['Code']},{row['ShiftName']}\n")
            
            conn.close()
            
            # Return as downloadable file
            from flask import make_response
            csv_data = output.getvalue()
            output.close()
            
            response = make_response(csv_data)
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = f'attachment; filename=Dienstplan_{start_date_str}_bis_{end_date_str}.csv'
            return response
            
        except Exception as e:
            app.logger.error(f"CSV export error: {str(e)}")
            return jsonify({'error': f'Export-Fehler: {str(e)}'}), 500
    
    def get_shift_color(shift_code: str) -> tuple:
        """
        Get background color and text color for a shift type.
        Returns (bg_color_hex, text_color_hex)
        Matches the colors from the UI (wwwroot/css/styles.css and database)
        """
        colors_map = {
            'F': ('#4CAF50', '#000000'),   # Früh - green with black text
            'S': ('#FF9800', '#FFFFFF'),   # Spät - orange with white text
            'N': ('#2196F3', '#FFFFFF'),   # Nacht - blue with white text
            'Z': ('#9C27B0', '#FFFFFF'),   # Zwischendienst - purple with white text
            'TD': ('#673AB7', '#FFFFFF'),  # Tagdienst - deep purple with white text
            'BMT': ('#F44336', '#FFFFFF'), # Brandmeldetechniker - red with white text
            'BSB': ('#E91E63', '#FFFFFF'), # Brandschutzbeauftragter - pink with white text
            'U': ('#64748b', '#FFFFFF'),   # Urlaub - gray with white text
            'AU': ('#dc2626', '#FFFFFF'),  # Krank - dark red with white text
            'L': ('#3b82f6', '#FFFFFF'),   # Lehrgang - blue with white text
        }
        return colors_map.get(shift_code, ('#E0E0E0', '#000000'))  # Default gray
    
    def group_data_by_team_and_employee(conn, start_date: date, end_date: date, view_type: str = 'week'):
        """
        Group shift assignments by team and employee, mirroring the UI's groupByTeamAndEmployee logic.
        Returns: (team_groups, dates, absences_by_employee)
        """
        cursor = conn.cursor()
        
        # Get all employees with their team info and special functions
        cursor.execute("""
            SELECT e.Id, e.Vorname, e.Name, e.Personalnummer, e.TeamId, 
                   t.Name as TeamName, e.IsSpringer,
                   e.IsBrandmeldetechniker, e.IsBrandschutzbeauftragter
            FROM Employees e
            LEFT JOIN Teams t ON e.TeamId = t.Id
            ORDER BY t.Name NULLS LAST, e.IsSpringer DESC, e.Name, e.Vorname
        """)
        employees = cursor.fetchall()
        
        # Get all shift assignments in the date range
        cursor.execute("""
            SELECT sa.Date, sa.EmployeeId, st.Code, st.Name as ShiftName, st.ColorCode
            FROM ShiftAssignments sa
            JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
            WHERE sa.Date >= ? AND sa.Date <= ?
        """, (start_date.isoformat(), end_date.isoformat()))
        assignments = cursor.fetchall()
        
        # Get all absences in the date range
        cursor.execute("""
            SELECT a.EmployeeId, a.StartDate, a.EndDate, a.Type, a.Notes
            FROM Absences a
            WHERE (a.StartDate <= ? AND a.EndDate >= ?)
               OR (a.StartDate >= ? AND a.StartDate <= ?)
        """, (end_date.isoformat(), start_date.isoformat(), 
              start_date.isoformat(), end_date.isoformat()))
        absences = cursor.fetchall()
        
        # Generate date range
        dates = []
        current = start_date
        while current <= end_date:
            dates.append(current.isoformat())
            current += timedelta(days=1)
        
        # Build absences lookup
        absences_by_employee = {}
        for absence in absences:
            emp_id = absence['EmployeeId']
            if emp_id not in absences_by_employee:
                absences_by_employee[emp_id] = []
            absences_by_employee[emp_id].append(absence)
        
        # Build assignments lookup
        assignments_by_emp_date = {}
        for assignment in assignments:
            key = (assignment['EmployeeId'], assignment['Date'])
            if key not in assignments_by_emp_date:
                assignments_by_emp_date[key] = []
            assignments_by_emp_date[key].append(assignment)
        
        # Group by team
        UNASSIGNED_TEAM_ID = -1
        
        teams = {}
        for emp in employees:
            # Format employee name once for reuse
            emp_name = f"{emp['Vorname']} {emp['Name']}"
            if emp['Personalnummer']:
                emp_name = f"{emp_name} ({emp['Personalnummer']})"
            
            # Add to regular team
            team_id = emp['TeamId'] if emp['TeamId'] else UNASSIGNED_TEAM_ID
            team_name = emp['TeamName'] if emp['TeamName'] else 'Ohne Team'
            
            if team_id not in teams:
                teams[team_id] = {
                    'teamId': team_id,
                    'teamName': team_name,
                    'employees': {}
                }
            
            teams[team_id]['employees'][emp['Id']] = {
                'id': emp['Id'],
                'name': emp_name,
                'shifts': {}
            }
        
        # Populate shifts for each employee
        for team in teams.values():
            for emp_id, emp_data in team['employees'].items():
                for date_str in dates:
                    key = (emp_id, date_str)
                    shifts = assignments_by_emp_date.get(key, [])
                    emp_data['shifts'][date_str] = shifts
        
        # Sort teams (regular -> Ohne Team)
        sorted_teams = []
        for team_id in sorted(teams.keys()):
            if team_id == UNASSIGNED_TEAM_ID:
                continue
            sorted_teams.append(teams[team_id])
        
        if UNASSIGNED_TEAM_ID in teams:
            sorted_teams.append(teams[UNASSIGNED_TEAM_ID])
        
        # Sort employees within each team by name
        for team in sorted_teams:
            team['employees'] = dict(sorted(
                team['employees'].items(),
                key=lambda x: x[1]['name']
            ))
        
        return sorted_teams, dates, absences_by_employee
    
    def get_absence_for_date(absences: list, date_str: str) -> Optional[dict]:
        """Check if an employee has an absence on a specific date"""
        target_date = date.fromisoformat(date_str)
        for absence in absences:
            start = date.fromisoformat(absence['StartDate'])
            end = date.fromisoformat(absence['EndDate'])
            if start <= target_date <= end:
                return absence
        return None
    
    def get_absence_code(absence_type: int) -> str:
        """Convert absence type to code (U, AU, L)"""
        # From entities.py: U=1 (Urlaub), AU=2 (Krank), L=3 (Lehrgang/Fortbildung)
        codes = {1: 'U', 2: 'AU', 3: 'L'}
        return codes.get(absence_type, 'U')

    @app.route('/api/shifts/export/pdf', methods=['GET'])
    def export_schedule_pdf():
        """Export schedule to PDF format matching the UI view structure"""
        start_date_str = request.args.get('startDate')
        end_date_str = request.args.get('endDate')
        view_type = request.args.get('view', 'week')  # week, month, or year
        
        if not start_date_str or not end_date_str:
            return jsonify({'error': 'startDate and endDate are required'}), 400
        
        try:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
            
            conn = db.get_connection()
            
            # Get grouped data matching UI structure
            team_groups, dates, absences_by_employee = group_data_by_team_and_employee(conn, start_date, end_date, view_type)
            
            # Create PDF
            import io
            from reportlab.lib.colors import HexColor
            buffer = io.BytesIO()
            
            # Determine appropriate page size based on view type and number of columns
            # For month/year views with many columns, we need to adjust the page size or scale
            num_columns = len(dates) + 1  # +1 for employee name column
            
            # Calculate required table width
            employee_col_width = 5*cm
            if view_type == 'year':
                date_col_width = 0.8*cm  # Smaller for year view (365 days)
            elif view_type == 'month' and len(dates) > 28:
                date_col_width = 1.2*cm  # Compressed for month view
            else:
                date_col_width = 1.8*cm  # Normal for week view
            
            required_width = employee_col_width + (len(dates) * date_col_width)
            
            # Standard landscape A4 width for comparison
            landscape_a4_width = landscape(A4)[0]
            
            # Determine page size - use A3 for large tables
            if required_width > landscape_a4_width - 2*cm:
                # Use landscape A3 for larger tables
                pagesize = landscape(A3)
            else:
                pagesize = landscape(A4)
            
            # Set margins to maximize usable space
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=pagesize,
                leftMargin=0.5*cm,
                rightMargin=0.5*cm,
                topMargin=1*cm,
                bottomMargin=1*cm
            )
            elements = []
            
            # Title
            styles = getSampleStyleSheet()
            if view_type == 'week':
                # Get week number
                first_date_obj = datetime.fromisoformat(dates[0])
                week_num = first_date_obj.isocalendar()[1]
                year = first_date_obj.year
                title_text = f"Dienstplan - Woche: KW {week_num} {year}"
            elif view_type == 'month':
                first_date_obj = datetime.fromisoformat(dates[0])
                month_name = first_date_obj.strftime('%B %Y')
                title_text = f"Dienstplan - Monat: {month_name}"
            else:  # year
                year = datetime.fromisoformat(dates[0]).year
                title_text = f"Dienstplan - Jahr: {year}"
            
            title = Paragraph(title_text, styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.3*cm))
            
            # Build table data matching UI structure
            table_data = []
            
            # Header row
            header_row = ['Team / Mitarbeiter']
            for date_str in dates:
                date_obj = datetime.fromisoformat(date_str)
                if view_type == 'year':
                    # For year view, show only date number
                    header_row.append(date_obj.strftime('%d.%m'))
                else:
                    # For week/month view, show day name and date
                    day_name = date_obj.strftime('%a')
                    day_num = date_obj.strftime('%d.%m')
                    header_row.append(f"{day_name}\n{day_num}")
            table_data.append(header_row)
            
            # Data rows - grouped by team
            for team in team_groups:
                # Team header row
                team_row = [team['teamName']] + [''] * len(dates)
                table_data.append(team_row)
                
                # Employee rows
                for emp_id, emp_data in team['employees'].items():
                    emp_row = [f"  - {emp_data['name']}"]
                    
                    for date_str in dates:
                        # Check for absence first
                        absences = absences_by_employee.get(emp_id, [])
                        absence = get_absence_for_date(absences, date_str)
                        
                        if absence:
                            absence_code = get_absence_code(absence['Type'])
                            emp_row.append(absence_code)
                        else:
                            # Get shifts for this date
                            shifts = emp_data['shifts'].get(date_str, [])
                            if shifts:
                                shift_codes = ' '.join([s['Code'] for s in shifts])
                                emp_row.append(shift_codes)
                            else:
                                emp_row.append('-')
                    
                    table_data.append(emp_row)
            
            conn.close()
            
            # Create table with styling
            # Use the dynamically calculated column widths
            col_widths = [employee_col_width] + [date_col_width] * len(dates)
            
            table = Table(table_data, colWidths=col_widths)
            
            # Apply styling
            # Adjust font sizes based on view type
            if view_type == 'year':
                header_font_size = 6
                data_font_size = 5
            elif view_type == 'month':
                header_font_size = 7
                data_font_size = 6
            else:  # week
                header_font_size = 9
                data_font_size = 8
            
            style_commands = [
                # Header row styling
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                # First column (employee names) - left aligned
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTSIZE', (0, 1), (-1, -1), data_font_size),
                ('LEFTPADDING', (0, 1), (0, -1), 3),
                ('RIGHTPADDING', (0, 1), (0, -1), 3),
                ('TOPPADDING', (0, 1), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ]
            
            # Add team header row styling
            row_idx = 1
            for team in team_groups:
                # Team header background with gradient-like color
                style_commands.append(
                    ('BACKGROUND', (0, row_idx), (-1, row_idx), HexColor('#2563eb'))
                )
                style_commands.append(
                    ('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.white)
                )
                style_commands.append(
                    ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold')
                )
                style_commands.append(
                    ('ALIGN', (0, row_idx), (0, row_idx), 'LEFT')
                )
                row_idx += 1 + len(team['employees'])
            
            table.setStyle(TableStyle(style_commands))
            elements.append(table)
            doc.build(elements)
            
            # Return PDF
            buffer.seek(0)
            return send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'Dienstplan_{start_date_str}_bis_{end_date_str}.pdf'
            )
            
        except Exception as e:
            app.logger.error(f"PDF export error: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': f'PDF-Export-Fehler: {str(e)}'}), 500
    
    @app.route('/api/shifts/export/excel', methods=['GET'])
    def export_schedule_excel():
        """Export schedule to Excel format matching the UI view structure"""
        start_date_str = request.args.get('startDate')
        end_date_str = request.args.get('endDate')
        view_type = request.args.get('view', 'week')  # week, month, or year
        
        if not start_date_str or not end_date_str:
            return jsonify({'error': 'startDate and endDate are required'}), 400
        
        try:
            # Import Excel library
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                return jsonify({
                    'error': 'Excel-Export erfordert openpyxl. Bitte installieren Sie es mit: pip install openpyxl'
                }), 501
            
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
            
            conn = db.get_connection()
            
            # Get grouped data matching UI structure
            team_groups, dates, absences_by_employee = group_data_by_team_and_employee(conn, start_date, end_date, view_type)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Set title based on view type
            if view_type == 'week':
                first_date_obj = datetime.fromisoformat(dates[0])
                week_num = first_date_obj.isocalendar()[1]
                year = first_date_obj.year
                ws.title = f"KW{week_num} {year}"
            elif view_type == 'month':
                first_date_obj = datetime.fromisoformat(dates[0])
                ws.title = first_date_obj.strftime('%B %Y')
            else:  # year
                year = datetime.fromisoformat(dates[0]).year
                ws.title = f"Jahr {year}"
            
            # Header row
            header_row = ['Team / Mitarbeiter']
            for date_str in dates:
                date_obj = datetime.fromisoformat(date_str)
                if view_type == 'year':
                    header_row.append(date_obj.strftime('%d.%m'))
                else:
                    day_name = date_obj.strftime('%a')
                    day_num = date_obj.strftime('%d.%m')
                    header_row.append(f"{day_name}\n{day_num}")
            ws.append(header_row)
            
            # Style header row
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Set row height for header
            ws.row_dimensions[1].height = 30
            
            # Data rows - grouped by team
            current_row = 2
            for team in team_groups:
                # Team header row
                team_row = [team['teamName']] + [''] * len(dates)
                ws.append(team_row)
                
                # Style team header
                team_font = Font(bold=True, color="FFFFFF", size=10)
                team_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
                team_alignment = Alignment(horizontal="left", vertical="center")
                
                for col_idx in range(1, len(dates) + 2):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = team_font
                    cell.fill = team_fill
                    cell.alignment = team_alignment
                    cell.border = border
                
                current_row += 1
                
                # Employee rows
                for emp_id, emp_data in team['employees'].items():
                    emp_row = [f"  - {emp_data['name']}"]
                    
                    for date_str in dates:
                        # Check for absence first
                        absences = absences_by_employee.get(emp_id, [])
                        absence = get_absence_for_date(absences, date_str)
                        
                        if absence:
                            absence_code = get_absence_code(absence['Type'])
                            emp_row.append(absence_code)
                        else:
                            # Get shifts for this date
                            shifts = emp_data['shifts'].get(date_str, [])
                            if shifts:
                                shift_codes = ' '.join([s['Code'] for s in shifts])
                                emp_row.append(shift_codes)
                            else:
                                emp_row.append('-')
                    
                    ws.append(emp_row)
                    
                    # Style employee row
                    emp_font = Font(size=9)
                    emp_alignment_left = Alignment(horizontal="left", vertical="center")
                    emp_alignment_center = Alignment(horizontal="center", vertical="center")
                    
                    # First cell (employee name) - left aligned
                    cell = ws.cell(row=current_row, column=1)
                    cell.font = emp_font
                    cell.alignment = emp_alignment_left
                    cell.border = border
                    
                    # Shift cells - with color coding
                    for col_idx in range(2, len(dates) + 2):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.font = Font(size=8, bold=True)
                        cell.alignment = emp_alignment_center
                        cell.border = border
                        
                        # Get the shift code to apply color
                        cell_value = str(cell.value) if cell.value else ''
                        if cell_value and cell_value != '-':
                            # Split multiple shifts and use first one for color
                            first_shift = cell_value.split()[0]
                            bg_color, text_color = get_shift_color(first_shift)
                            # Remove # from hex colors for openpyxl
                            bg_hex = bg_color.replace('#', '')
                            text_hex = text_color.replace('#', '')
                            cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                            cell.font = Font(size=8, bold=True, color=text_hex)
                    
                    current_row += 1
            
            conn.close()
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 30  # Employee names column
            for col_idx in range(2, len(dates) + 2):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if view_type == 'year':
                    ws.column_dimensions[col_letter].width = 6
                else:
                    ws.column_dimensions[col_letter].width = 8
            
            # Save to BytesIO
            import io
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Return Excel file
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'Dienstplan_{start_date_str}_bis_{end_date_str}.xlsx'
            )
            
        except Exception as e:
            app.logger.error(f"Excel export error: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': f'Excel-Export-Fehler: {str(e)}'}), 500
    
    # ============================================================================
    # ABSENCE ENDPOINTS
    # ============================================================================
    
    @app.route('/api/absences', methods=['GET'])
    def get_absences():
        """Get all absences"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT a.*, e.Vorname, e.Name, e.TeamId
            FROM Absences a
            JOIN Employees e ON a.EmployeeId = e.Id
            ORDER BY a.StartDate DESC
        """)
        
        absences = []
        for row in cursor.fetchall():
            # Map type: 1=AU (Krank), 2=U (Urlaub), 3=L (Lehrgang)
            type_names = ['', 'Krank / AU', 'Urlaub', 'Lehrgang']
            absences.append({
                'id': row['Id'],
                'employeeId': row['EmployeeId'],
                'employeeName': f"{row['Vorname']} {row['Name']}",
                'teamId': row['TeamId'],
                'type': type_names[row['Type']] if row['Type'] < len(type_names) else 'Unbekannt',
                'startDate': row['StartDate'],
                'endDate': row['EndDate'],
                'notes': row['Notes'],
                'createdAt': row['CreatedAt']
            })
        
        conn.close()
        return jsonify(absences)
    
    @app.route('/api/absences', methods=['POST'])
    @require_role('Admin')
    def create_absence():
        """Create new absence (AU or Lehrgang)"""
        try:
            data = request.get_json()
            
            if not data.get('employeeId') or not data.get('type') or not data.get('startDate') or not data.get('endDate'):
                return jsonify({'error': 'EmployeeId, Type, StartDate und EndDate sind erforderlich'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO Absences 
                (EmployeeId, Type, StartDate, EndDate, Notes, CreatedAt, CreatedBy)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get('employeeId'),
                data.get('type'),
                data.get('startDate'),
                data.get('endDate'),
                data.get('notes'),
                datetime.utcnow().isoformat(),
                session.get('user_email')
            ))
            
            absence_id = cursor.lastrowid
            
            # Log audit entry
            changes = json.dumps({
                'employeeId': data.get('employeeId'),
                'type': data.get('type'),
                'startDate': data.get('startDate'),
                'endDate': data.get('endDate'),
                'notes': data.get('notes')
            }, ensure_ascii=False)
            log_audit(conn, 'Absence', absence_id, 'Created', changes)
            
            # Check for understaffing and create notifications
            try:
                start_date_obj = date.fromisoformat(data.get('startDate'))
                end_date_obj = date.fromisoformat(data.get('endDate'))
                
                notification_ids = process_absence_for_notifications(
                    conn,
                    absence_id,
                    data.get('employeeId'),
                    start_date_obj,
                    end_date_obj,
                    data.get('type'),
                    session.get('user_email')
                )
                
                if notification_ids:
                    app.logger.info(f"Created {len(notification_ids)} understaffing notifications for absence {absence_id}")
                
                # NEW: Automatically assign springers for affected shifts
                springer_results = process_absence_with_springer_assignment(
                    conn,
                    absence_id,
                    data.get('employeeId'),
                    start_date_obj,
                    end_date_obj,
                    data.get('type'),
                    session.get('user_email')
                )
                
                if springer_results['assignmentsCreated'] > 0:
                    app.logger.info(
                        f"Automatically assigned {springer_results['assignmentsCreated']} springers " +
                        f"for {springer_results['shiftsNeedingCoverage']} affected shifts " +
                        f"(Absence ID: {absence_id})"
                    )
                    
                    # Include springer results in response
                    conn.commit()
                    conn.close()
                    
                    return jsonify({
                        'success': True,
                        'id': absence_id,
                        'springerAssignments': {
                            'assignmentsCreated': springer_results['assignmentsCreated'],
                            'notificationsSent': springer_results['notificationsSent'],
                            'shiftsNeedingCoverage': springer_results['shiftsNeedingCoverage'],
                            'details': springer_results['details']
                        }
                    }), 201
                    
            except Exception as notif_error:
                # Log notification error but don't fail the absence creation
                app.logger.error(f"Error processing absence notifications/springers: {notif_error}")
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'id': absence_id}), 201
            
        except Exception as e:
            app.logger.error(f"Create absence error: {str(e)}")
            return jsonify({'error': f'Fehler beim Erstellen: {str(e)}'}), 500
    
    @app.route('/api/absences/<int:id>', methods=['DELETE'])
    @require_role('Admin')
    def delete_absence(id):
        """Delete an absence"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Get absence info for audit before deleting
            cursor.execute("""
                SELECT EmployeeId, Type, StartDate, EndDate 
                FROM Absences WHERE Id = ?
            """, (id,))
            absence_row = cursor.fetchone()
            
            if not absence_row:
                conn.close()
                return jsonify({'error': 'Abwesenheit nicht gefunden'}), 404
            
            cursor.execute("DELETE FROM Absences WHERE Id = ?", (id,))
            
            # Log audit entry
            changes = json.dumps({
                'employeeId': absence_row['EmployeeId'],
                'type': absence_row['Type'],
                'startDate': absence_row['StartDate'],
                'endDate': absence_row['EndDate']
            }, ensure_ascii=False)
            log_audit(conn, 'Absence', id, 'Deleted', changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Delete absence error: {str(e)}")
            return jsonify({'error': f'Fehler beim Löschen: {str(e)}'}), 500
    
    # ============================================================================
    # VACATION REQUEST ENDPOINTS
    # ============================================================================
    
    @app.route('/api/vacationrequests', methods=['GET'])
    def get_vacation_requests():
        """Get all vacation requests or pending ones"""
        status_filter = request.args.get('status')
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        if status_filter == 'pending':
            cursor.execute("""
                SELECT vr.*, e.Vorname, e.Name, e.TeamId
                FROM VacationRequests vr
                JOIN Employees e ON vr.EmployeeId = e.Id
                WHERE vr.Status = 'InBearbeitung'
                ORDER BY vr.CreatedAt DESC
            """)
        else:
            cursor.execute("""
                SELECT vr.*, e.Vorname, e.Name, e.TeamId
                FROM VacationRequests vr
                JOIN Employees e ON vr.EmployeeId = e.Id
                ORDER BY vr.CreatedAt DESC
            """)
        
        requests = []
        for row in cursor.fetchall():
            requests.append({
                'id': row['Id'],
                'employeeId': row['EmployeeId'],
                'employeeName': f"{row['Vorname']} {row['Name']}",
                'teamId': row['TeamId'],
                'startDate': row['StartDate'],
                'endDate': row['EndDate'],
                'status': row['Status'],
                'notes': row['Notes'],
                'disponentResponse': row['DisponentResponse'],
                'createdAt': row['CreatedAt'],
                'processedAt': row['ProcessedAt']
            })
        
        conn.close()
        return jsonify(requests)
    
    @app.route('/api/vacationrequests', methods=['POST'])
    @require_auth
    def create_vacation_request():
        """Create new vacation request"""
        try:
            data = request.get_json()
            
            if not data.get('employeeId') or not data.get('startDate') or not data.get('endDate'):
                return jsonify({'error': 'EmployeeId, StartDate und EndDate sind erforderlich'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO VacationRequests 
                (EmployeeId, StartDate, EndDate, Status, Notes, CreatedAt, CreatedBy)
                VALUES (?, ?, ?, 'InBearbeitung', ?, ?, ?)
            """, (
                data.get('employeeId'),
                data.get('startDate'),
                data.get('endDate'),
                data.get('notes'),
                datetime.utcnow().isoformat(),
                session.get('user_email')
            ))
            
            request_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'id': request_id}), 201
            
        except Exception as e:
            app.logger.error(f"Create vacation request error: {str(e)}")
            return jsonify({'error': f'Fehler beim Erstellen: {str(e)}'}), 500
    
    @app.route('/api/vacationrequests/<int:id>/status', methods=['PUT'])
    @require_role('Admin')
    def update_vacation_request_status(id):
        """Update vacation request status (Admin only)"""
        try:
            data = request.get_json()
            status = data.get('status')
            response = data.get('response')
            
            if status not in ['Genehmigt', 'Abgelehnt', 'InBearbeitung']:
                return jsonify({'error': 'Ungültiger Status'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE VacationRequests 
                SET Status = ?, DisponentResponse = ?, ProcessedAt = ?, ProcessedBy = ?
                WHERE Id = ?
            """, (
                status,
                response,
                datetime.utcnow().isoformat(),
                session.get('user_email'),
                id
            ))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Update vacation request error: {str(e)}")
            return jsonify({'error': f'Fehler beim Aktualisieren: {str(e)}'}), 500
    
    # ============================================================================
    # VACATION YEAR APPROVAL ENDPOINTS
    # ============================================================================
    
    @app.route('/api/vacationyearapprovals', methods=['GET'])
    def get_vacation_year_approvals():
        """Get all vacation year approvals"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT * FROM VacationYearApprovals
            ORDER BY Year DESC
        """)
        
        approvals = []
        for row in cursor.fetchall():
            approvals.append({
                'id': row['Id'],
                'year': row['Year'],
                'isApproved': bool(row['IsApproved']),
                'approvedAt': row['ApprovedAt'],
                'approvedBy': row['ApprovedBy'],
                'createdAt': row['CreatedAt'],
                'modifiedAt': row['ModifiedAt'],
                'notes': row['Notes']
            })
        
        conn.close()
        return jsonify(approvals)
    
    @app.route('/api/vacationyearapprovals/<int:year>', methods=['GET'])
    def get_vacation_year_approval(year):
        """Get vacation year approval status for a specific year"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT * FROM VacationYearApprovals WHERE Year = ?
        """, (year,))
        
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            return jsonify({
                'year': year,
                'isApproved': False,
                'exists': False
            })
        
        return jsonify({
            'id': row['Id'],
            'year': row['Year'],
            'isApproved': bool(row['IsApproved']),
            'approvedAt': row['ApprovedAt'],
            'approvedBy': row['ApprovedBy'],
            'createdAt': row['CreatedAt'],
            'modifiedAt': row['ModifiedAt'],
            'notes': row['Notes'],
            'exists': True
        })
    
    @app.route('/api/vacationyearapprovals', methods=['POST'])
    @require_role('Admin')
    def create_or_update_vacation_year_approval():
        """Create or update vacation year approval (Admin only)"""
        try:
            data = request.get_json()
            year = data.get('year')
            is_approved = data.get('isApproved', False)
            notes = data.get('notes')
            
            if not year:
                return jsonify({'error': 'Jahr ist erforderlich'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if entry exists
            cursor.execute("""
                SELECT Id FROM VacationYearApprovals WHERE Year = ?
            """, (year,))
            
            existing = cursor.fetchone()
            
            if existing:
                # Update existing entry
                cursor.execute("""
                    UPDATE VacationYearApprovals 
                    SET IsApproved = ?,
                        ApprovedAt = ?,
                        ApprovedBy = ?,
                        ModifiedAt = ?,
                        Notes = ?
                    WHERE Year = ?
                """, (
                    1 if is_approved else 0,
                    datetime.utcnow().isoformat() if is_approved else None,
                    session.get('user_email') if is_approved else None,
                    datetime.utcnow().isoformat(),
                    notes,
                    year
                ))
                
                approval_id = existing['Id']
                action = 'Updated'
            else:
                # Create new entry
                cursor.execute("""
                    INSERT INTO VacationYearApprovals 
                    (Year, IsApproved, ApprovedAt, ApprovedBy, CreatedAt, Notes)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    year,
                    1 if is_approved else 0,
                    datetime.utcnow().isoformat() if is_approved else None,
                    session.get('user_email') if is_approved else None,
                    datetime.utcnow().isoformat(),
                    notes
                ))
                
                approval_id = cursor.lastrowid
                action = 'Created'
            
            # Log audit entry
            changes = json.dumps({
                'year': year,
                'isApproved': is_approved,
                'notes': notes
            }, ensure_ascii=False)
            log_audit(conn, 'VacationYearApproval', approval_id, action, changes)
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'id': approval_id, 'year': year}), 201
            
        except Exception as e:
            app.logger.error(f"Create/update vacation year approval error: {str(e)}")
            return jsonify({'error': f'Fehler beim Speichern: {str(e)}'}), 500
    
    @app.route('/api/vacationyearplan/<int:year>', methods=['GET'])
    def get_vacation_year_plan(year):
        """
        Get vacation plan for a specific year.
        Returns vacation data only if the year is approved by admin.
        All users can access this endpoint.
        """
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Check if year is approved
        cursor.execute("""
            SELECT IsApproved FROM VacationYearApprovals WHERE Year = ?
        """, (year,))
        
        approval_row = cursor.fetchone()
        
        # If year is not approved, return empty data
        if not approval_row or not approval_row['IsApproved']:
            conn.close()
            return jsonify({
                'year': year,
                'isApproved': False,
                'vacations': [],
                'message': 'Urlaubsdaten für dieses Jahr wurden noch nicht freigegeben.'
            })
        
        # Get all vacation data for the year (from VacationRequests and Absences)
        start_date = f"{year}-01-01"
        end_date = f"{year}-12-31"
        
        # Get approved vacation requests
        cursor.execute("""
            SELECT 
                vr.Id,
                vr.EmployeeId,
                e.Vorname,
                e.Name,
                e.TeamId,
                t.Name as TeamName,
                vr.StartDate,
                vr.EndDate,
                vr.Status,
                vr.Notes,
                'VacationRequest' as Source
            FROM VacationRequests vr
            JOIN Employees e ON vr.EmployeeId = e.Id
            LEFT JOIN Teams t ON e.TeamId = t.Id
            WHERE (vr.StartDate <= ? AND vr.EndDate >= ?)
            ORDER BY vr.StartDate, e.Name, e.Vorname
        """, (end_date, start_date))
        
        vacation_requests = []
        for row in cursor.fetchall():
            vacation_requests.append({
                'id': row['Id'],
                'employeeId': row['EmployeeId'],
                'employeeName': f"{row['Vorname']} {row['Name']}",
                'teamId': row['TeamId'],
                'teamName': row['TeamName'],
                'startDate': row['StartDate'],
                'endDate': row['EndDate'],
                'status': row['Status'],
                'notes': row['Notes'],
                'source': row['Source']
            })
        
        # Get vacation absences (Type 2 = Urlaub)
        cursor.execute("""
            SELECT 
                a.Id,
                a.EmployeeId,
                e.Vorname,
                e.Name,
                e.TeamId,
                t.Name as TeamName,
                a.StartDate,
                a.EndDate,
                a.Notes,
                'Absence' as Source
            FROM Absences a
            JOIN Employees e ON a.EmployeeId = e.Id
            LEFT JOIN Teams t ON e.TeamId = t.Id
            WHERE a.Type = 2
            AND (a.StartDate <= ? AND a.EndDate >= ?)
            ORDER BY a.StartDate, e.Name, e.Vorname
        """, (end_date, start_date))
        
        absences = []
        for row in cursor.fetchall():
            absences.append({
                'id': row['Id'],
                'employeeId': row['EmployeeId'],
                'employeeName': f"{row['Vorname']} {row['Name']}",
                'teamId': row['TeamId'],
                'teamName': row['TeamName'],
                'startDate': row['StartDate'],
                'endDate': row['EndDate'],
                'status': 'Genehmigt',  # Absences are always approved
                'notes': row['Notes'],
                'source': row['Source']
            })
        
        conn.close()
        
        return jsonify({
            'year': year,
            'isApproved': True,
            'vacationRequests': vacation_requests,
            'absences': absences
        })
    
    # ============================================================================
    # SHIFT EXCHANGE ENDPOINTS
    # ============================================================================
    
    @app.route('/api/shiftexchanges/available', methods=['GET'])
    def get_available_shift_exchanges():
        """Get available shift exchanges"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT se.*, 
                   sa.Date, sa.ShiftTypeId,
                   st.Code as ShiftCode, st.Name as ShiftName,
                   e.Vorname, e.Name, e.TeamId
            FROM ShiftExchanges se
            JOIN ShiftAssignments sa ON se.ShiftAssignmentId = sa.Id
            JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
            JOIN Employees e ON se.OfferingEmployeeId = e.Id
            WHERE se.Status = 'Angeboten'
            ORDER BY sa.Date
        """)
        
        exchanges = []
        for row in cursor.fetchall():
            exchanges.append({
                'id': row['Id'],
                'shiftAssignmentId': row['ShiftAssignmentId'],
                'offeringEmployeeId': row['OfferingEmployeeId'],
                'offeringEmployeeName': f"{row['Vorname']} {row['Name']}",
                'teamId': row['TeamId'],
                'date': row['Date'],
                'shiftCode': row['ShiftCode'],
                'shiftName': row['ShiftName'],
                'status': row['Status'],
                'offeringReason': row['OfferingReason'],
                'createdAt': row['CreatedAt']
            })
        
        conn.close()
        return jsonify(exchanges)
    
    @app.route('/api/shiftexchanges/pending', methods=['GET'])
    @require_role('Admin')
    def get_pending_shift_exchanges():
        """Get pending shift exchanges (Admin only)"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT se.*, 
                   sa.Date, sa.ShiftTypeId,
                   st.Code as ShiftCode, st.Name as ShiftName,
                   e1.Vorname as OfferingVorname, e1.Name as OfferingName,
                   e2.Vorname as RequestingVorname, e2.Name as RequestingName
            FROM ShiftExchanges se
            JOIN ShiftAssignments sa ON se.ShiftAssignmentId = sa.Id
            JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
            JOIN Employees e1 ON se.OfferingEmployeeId = e1.Id
            LEFT JOIN Employees e2 ON se.RequestingEmployeeId = e2.Id
            WHERE se.Status = 'Angefragt'
            ORDER BY sa.Date
        """)
        
        exchanges = []
        for row in cursor.fetchall():
            exchanges.append({
                'id': row['Id'],
                'shiftAssignmentId': row['ShiftAssignmentId'],
                'offeringEmployeeId': row['OfferingEmployeeId'],
                'offeringEmployeeName': f"{row['OfferingVorname']} {row['OfferingName']}",
                'requestingEmployeeId': row['RequestingEmployeeId'],
                'requestingEmployeeName': f"{row['RequestingVorname']} {row['RequestingName']}" if row['RequestingEmployeeId'] else None,
                'date': row['Date'],
                'shiftCode': row['ShiftCode'],
                'shiftName': row['ShiftName'],
                'status': row['Status'],
                'offeringReason': row['OfferingReason'],
                'createdAt': row['CreatedAt']
            })
        
        conn.close()
        return jsonify(exchanges)
    
    @app.route('/api/shiftexchanges', methods=['POST'])
    @require_auth
    def create_shift_exchange():
        """Create new shift exchange offer"""
        try:
            data = request.get_json()
            
            if not data.get('shiftAssignmentId') or not data.get('offeringEmployeeId'):
                return jsonify({'error': 'ShiftAssignmentId und OfferingEmployeeId sind erforderlich'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO ShiftExchanges 
                (ShiftAssignmentId, OfferingEmployeeId, Status, OfferingReason, CreatedAt)
                VALUES (?, ?, 'Angeboten', ?, ?)
            """, (
                data.get('shiftAssignmentId'),
                data.get('offeringEmployeeId'),
                data.get('offeringReason'),
                datetime.utcnow().isoformat()
            ))
            
            exchange_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'id': exchange_id}), 201
            
        except Exception as e:
            app.logger.error(f"Create shift exchange error: {str(e)}")
            return jsonify({'error': f'Fehler beim Erstellen: {str(e)}'}), 500
    
    @app.route('/api/shiftexchanges/<int:id>/request', methods=['POST'])
    @require_auth
    def request_shift_exchange(id):
        """Request a shift exchange"""
        try:
            data = request.get_json()
            requesting_employee_id = data.get('requestingEmployeeId')
            
            if not requesting_employee_id:
                return jsonify({'error': 'RequestingEmployeeId ist erforderlich'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE ShiftExchanges 
                SET RequestingEmployeeId = ?, Status = 'Angefragt'
                WHERE Id = ? AND Status = 'Angeboten'
            """, (requesting_employee_id, id))
            
            if cursor.rowcount == 0:
                conn.close()
                return jsonify({'error': 'Tauschangebot nicht verfügbar'}), 404
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Request shift exchange error: {str(e)}")
            return jsonify({'error': f'Fehler beim Anfragen: {str(e)}'}), 500
    
    @app.route('/api/shiftexchanges/<int:id>/process', methods=['PUT'])
    @require_role('Admin')
    def process_shift_exchange(id):
        """Process shift exchange (approve/reject)"""
        try:
            data = request.get_json()
            status = data.get('status')
            notes = data.get('notes')
            
            if status not in ['Genehmigt', 'Abgelehnt']:
                return jsonify({'error': 'Status muss Genehmigt oder Abgelehnt sein'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Get exchange details
            cursor.execute("""
                SELECT ShiftAssignmentId, OfferingEmployeeId, RequestingEmployeeId
                FROM ShiftExchanges
                WHERE Id = ? AND Status = 'Angefragt'
            """, (id,))
            
            row = cursor.fetchone()
            if not row:
                conn.close()
                return jsonify({'error': 'Tauschangebot nicht gefunden oder bereits bearbeitet'}), 404
            
            shift_assignment_id = row['ShiftAssignmentId']
            requesting_employee_id = row['RequestingEmployeeId']
            
            # Update exchange status
            cursor.execute("""
                UPDATE ShiftExchanges 
                SET Status = ?, DisponentNotes = ?, ProcessedAt = ?, ProcessedBy = ?
                WHERE Id = ?
            """, (
                status,
                notes,
                datetime.utcnow().isoformat(),
                session.get('user_email'),
                id
            ))
            
            # If approved, update the shift assignment
            if status == 'Genehmigt' and requesting_employee_id:
                cursor.execute("""
                    UPDATE ShiftAssignments
                    SET EmployeeId = ?, ModifiedAt = ?, ModifiedBy = ?
                    WHERE Id = ?
                """, (
                    requesting_employee_id,
                    datetime.utcnow().isoformat(),
                    session.get('user_email'),
                    shift_assignment_id
                ))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Process shift exchange error: {str(e)}")
            return jsonify({'error': f'Fehler beim Bearbeiten: {str(e)}'}), 500
    
    # ============================================================================
    # STATISTICS ENDPOINTS
    # ============================================================================
    
    @app.route('/api/statistics/dashboard', methods=['GET'])
    def get_dashboard_stats():
        """Get dashboard statistics"""
        start_date_str = request.args.get('startDate')
        end_date_str = request.args.get('endDate')
        
        if not start_date_str or not end_date_str:
            # Default to current month
            today = date.today()
            start_date = date(today.year, today.month, 1)
            if today.month == 12:
                end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)
        else:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Employee work hours
        cursor.execute("""
            SELECT e.Id, e.Vorname, e.Name, e.TeamId,
                   COUNT(sa.Id) as ShiftCount,
                   COUNT(sa.Id) * 8.0 as TotalHours
            FROM Employees e
            LEFT JOIN ShiftAssignments sa ON e.Id = sa.EmployeeId 
                AND sa.Date >= ? AND sa.Date <= ?
            GROUP BY e.Id, e.Vorname, e.Name, e.TeamId
            HAVING ShiftCount > 0
            ORDER BY TotalHours DESC
        """, (start_date.isoformat(), end_date.isoformat()))
        
        employee_work_hours = []
        for row in cursor.fetchall():
            employee_work_hours.append({
                'employeeId': row['Id'],
                'employeeName': f"{row['Vorname']} {row['Name']}",
                'teamId': row['TeamId'],
                'shiftCount': row['ShiftCount'],
                'totalHours': row['TotalHours']
            })
        
        # Team shift distribution
        cursor.execute("""
            SELECT t.Id, t.Name,
                   st.Code,
                   COUNT(sa.Id) as ShiftCount
            FROM Teams t
            LEFT JOIN Employees e ON t.Id = e.TeamId
            LEFT JOIN ShiftAssignments sa ON e.Id = sa.EmployeeId 
                AND sa.Date >= ? AND sa.Date <= ?
            LEFT JOIN ShiftTypes st ON sa.ShiftTypeId = st.Id
            WHERE st.Code IS NOT NULL
            GROUP BY t.Id, t.Name, st.Code
            ORDER BY t.Name, st.Code
        """, (start_date.isoformat(), end_date.isoformat()))
        
        team_shift_data = {}
        for row in cursor.fetchall():
            team_id = row['Id']
            if team_id not in team_shift_data:
                team_shift_data[team_id] = {
                    'teamId': team_id,
                    'teamName': row['Name'],
                    'shiftCounts': {}
                }
            team_shift_data[team_id]['shiftCounts'][row['Code']] = row['ShiftCount']
        
        team_shift_distribution = list(team_shift_data.values())
        
        # Employee absence days
        cursor.execute("""
            SELECT e.Id, e.Vorname, e.Name,
                   SUM(julianday(a.EndDate) - julianday(a.StartDate) + 1) as TotalDays
            FROM Employees e
            JOIN Absences a ON e.Id = a.EmployeeId
            WHERE (a.StartDate <= ? AND a.EndDate >= ?)
               OR (a.StartDate >= ? AND a.StartDate <= ?)
            GROUP BY e.Id, e.Vorname, e.Name
            HAVING TotalDays > 0
            ORDER BY TotalDays DESC
        """, (end_date.isoformat(), start_date.isoformat(),
              start_date.isoformat(), end_date.isoformat()))
        
        employee_absence_days = []
        for row in cursor.fetchall():
            employee_absence_days.append({
                'employeeId': row['Id'],
                'employeeName': f"{row['Vorname']} {row['Name']}",
                'totalDays': int(row['TotalDays'])
            })
        
        # Team workload
        cursor.execute("""
            SELECT t.Id, t.Name,
                   COUNT(DISTINCT e.Id) as EmployeeCount,
                   COUNT(sa.Id) as TotalShifts,
                   CASE WHEN COUNT(DISTINCT e.Id) > 0 
                        THEN CAST(COUNT(sa.Id) AS REAL) / COUNT(DISTINCT e.Id)
                        ELSE 0 END as AvgShiftsPerEmployee
            FROM Teams t
            LEFT JOIN Employees e ON t.Id = e.TeamId
            LEFT JOIN ShiftAssignments sa ON e.Id = sa.EmployeeId 
                AND sa.Date >= ? AND sa.Date <= ?
            GROUP BY t.Id, t.Name
            HAVING EmployeeCount > 0
            ORDER BY t.Name
        """, (start_date.isoformat(), end_date.isoformat()))
        
        team_workload = []
        for row in cursor.fetchall():
            team_workload.append({
                'teamId': row['Id'],
                'teamName': row['Name'],
                'employeeCount': row['EmployeeCount'],
                'totalShifts': row['TotalShifts'],
                'averageShiftsPerEmployee': row['AvgShiftsPerEmployee']
            })
        
        conn.close()
        
        return jsonify({
            'startDate': start_date.isoformat(),
            'endDate': end_date.isoformat(),
            'employeeWorkHours': employee_work_hours,
            'teamShiftDistribution': team_shift_distribution,
            'employeeAbsenceDays': employee_absence_days,
            'teamWorkload': team_workload
        })
    
    # ============================================================================
    # AUDIT LOG ENDPOINTS
    # ============================================================================
    
    @app.route('/api/auditlogs', methods=['GET'])
    @require_role('Admin')
    def get_audit_logs():
        """Get audit logs with pagination and filters"""
        try:
            # Get and validate pagination parameters
            try:
                page = int(request.args.get('page', 1))
                page_size = int(request.args.get('pageSize', 50))
            except (ValueError, TypeError):
                return jsonify({'error': 'Invalid pagination parameters'}), 400
            
            # Validate pagination ranges
            if page < 1:
                page = 1
            if page_size < 1 or page_size > 100:
                page_size = min(max(page_size, 1), 100)
            
            # Get filter parameters
            entity_name = request.args.get('entityName')
            action = request.args.get('action')
            start_date = request.args.get('startDate')
            end_date = request.args.get('endDate')
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Build WHERE clause with parameterized queries for safety
            where_clauses = []
            params = []
            
            # Whitelist valid filters to prevent any potential SQL injection
            if entity_name:
                where_clauses.append("EntityName = ?")
                params.append(entity_name)
            
            if action:
                where_clauses.append("Action = ?")
                params.append(action)
            
            if start_date:
                where_clauses.append("DATE(Timestamp) >= ?")
                params.append(start_date)
            
            if end_date:
                where_clauses.append("DATE(Timestamp) <= ?")
                params.append(end_date)
            
            # Safe: only joining static WHERE clauses, all values are parameterized
            where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
            
            # Get total count
            count_query = f"SELECT COUNT(*) as total FROM AuditLogs WHERE {where_sql}"
            cursor.execute(count_query, params)
            total_count = cursor.fetchone()['total']
            
            # Calculate pagination
            total_pages = (total_count + page_size - 1) // page_size if total_count > 0 else 1
            offset = (page - 1) * page_size
            
            # Get paginated results - safe: WHERE clause uses only parameterized queries
            select_query = f"""
                SELECT Id, Timestamp, UserId, UserName, EntityName, EntityId, Action, Changes
                FROM AuditLogs
                WHERE {where_sql}
                ORDER BY Timestamp DESC
                LIMIT ? OFFSET ?
            """
            cursor.execute(select_query, params + [page_size, offset])
            
            items = []
            for row in cursor.fetchall():
                items.append({
                    'id': row['Id'],
                    'timestamp': row['Timestamp'],
                    'userId': row['UserId'],
                    'userName': row['UserName'],
                    'entityName': row['EntityName'],
                    'entityId': row['EntityId'],
                    'action': row['Action'],
                    'changes': row['Changes']
                })
            
            conn.close()
            
            return jsonify({
                'items': items,
                'page': page,
                'pageSize': page_size,
                'totalCount': total_count,
                'totalPages': total_pages,
                'hasPreviousPage': page > 1,
                'hasNextPage': page < total_pages
            })
            
        except Exception as e:
            app.logger.error(f"Get audit logs error: {str(e)}")
            return jsonify({'error': f'Fehler beim Laden der Audit-Logs: {str(e)}'}), 500
    
    @app.route('/api/auditlogs/recent/<int:count>', methods=['GET'])
    @require_role('Admin')
    def get_recent_audit_logs(count):
        """Get recent audit logs (simplified endpoint for backwards compatibility)"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT Id, Timestamp, UserId, UserName, EntityName, EntityId, Action, Changes
                FROM AuditLogs
                ORDER BY Timestamp DESC
                LIMIT ?
            """, (count,))
            
            logs = []
            for row in cursor.fetchall():
                logs.append({
                    'id': row['Id'],
                    'timestamp': row['Timestamp'],
                    'userId': row['UserId'],
                    'userName': row['UserName'],
                    'entityName': row['EntityName'],
                    'entityId': row['EntityId'],
                    'action': row['Action'],
                    'changes': row['Changes']
                })
            
            conn.close()
            return jsonify(logs)
            
        except Exception as e:
            app.logger.error(f"Get recent audit logs error: {str(e)}")
            return jsonify({'error': f'Fehler beim Laden der Audit-Logs: {str(e)}'}), 500
    
    # ============================================================================
    # NOTIFICATION ENDPOINTS
    # ============================================================================
    
    @app.route('/api/notifications', methods=['GET'])
    @require_role('Admin', 'Disponent')
    def get_notifications():
        """Get admin notifications (for Admins and Disponents only)"""
        try:
            unread_only = request.args.get('unreadOnly', 'false').lower() == 'true'
            limit = int(request.args.get('limit', 50))
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            if unread_only:
                cursor.execute("""
                    SELECT 
                        n.Id, n.Type, n.Severity, n.Title, n.Message,
                        n.ShiftDate, n.ShiftCode, n.RequiredStaff, n.ActualStaff,
                        n.CreatedAt, n.IsRead, n.ReadAt, n.ReadBy,
                        e.Vorname, e.Name, 
                        t.Name as TeamName
                    FROM AdminNotifications n
                    LEFT JOIN Employees e ON n.EmployeeId = e.Id
                    LEFT JOIN Teams t ON n.TeamId = t.Id
                    WHERE n.IsRead = 0
                    ORDER BY n.CreatedAt DESC
                    LIMIT ?
                """, (limit,))
            else:
                cursor.execute("""
                    SELECT 
                        n.Id, n.Type, n.Severity, n.Title, n.Message,
                        n.ShiftDate, n.ShiftCode, n.RequiredStaff, n.ActualStaff,
                        n.CreatedAt, n.IsRead, n.ReadAt, n.ReadBy,
                        e.Vorname, e.Name, 
                        t.Name as TeamName
                    FROM AdminNotifications n
                    LEFT JOIN Employees e ON n.EmployeeId = e.Id
                    LEFT JOIN Teams t ON n.TeamId = t.Id
                    ORDER BY n.CreatedAt DESC
                    LIMIT ?
                """, (limit,))
            
            notifications = []
            for row in cursor.fetchall():
                notifications.append({
                    'id': row[0],
                    'type': row[1],
                    'severity': row[2],
                    'title': row[3],
                    'message': row[4],
                    'shiftDate': row[5],
                    'shiftCode': row[6],
                    'requiredStaff': row[7],
                    'actualStaff': row[8],
                    'createdAt': row[9],
                    'isRead': bool(row[10]),
                    'readAt': row[11],
                    'readBy': row[12],
                    'employeeName': f"{row[13]} {row[14]}" if row[13] else None,
                    'teamName': row[15]
                })
            
            conn.close()
            return jsonify(notifications)
            
        except Exception as e:
            app.logger.error(f"Get notifications error: {str(e)}")
            return jsonify({'error': f'Fehler beim Laden der Benachrichtigungen: {str(e)}'}), 500
    
    @app.route('/api/notifications/count', methods=['GET'])
    @require_role('Admin', 'Disponent')
    def get_notification_count_endpoint():
        """Get count of unread notifications"""
        try:
            conn = db.get_connection()
            count = get_notification_count(conn, unread_only=True)
            conn.close()
            
            return jsonify({'count': count})
            
        except Exception as e:
            app.logger.error(f"Get notification count error: {str(e)}")
            return jsonify({'error': f'Fehler: {str(e)}'}), 500
    
    @app.route('/api/notifications/<int:id>/read', methods=['POST'])
    @require_role('Admin', 'Disponent')
    def mark_notification_read(id):
        """Mark notification as read"""
        try:
            conn = db.get_connection()
            success = mark_notification_as_read(conn, id, session.get('user_email'))
            conn.close()
            
            if success:
                return jsonify({'success': True})
            else:
                return jsonify({'error': 'Benachrichtigung nicht gefunden'}), 404
            
        except Exception as e:
            app.logger.error(f"Mark notification read error: {str(e)}")
            return jsonify({'error': f'Fehler: {str(e)}'}), 500
    
    @app.route('/api/notifications/mark-all-read', methods=['POST'])
    @require_role('Admin', 'Disponent')
    def mark_all_notifications_read():
        """Mark all notifications as read"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE AdminNotifications
                SET IsRead = 1, ReadAt = CURRENT_TIMESTAMP, ReadBy = ?
                WHERE IsRead = 0
            """, (session.get('user_email'),))
            
            conn.commit()
            count = cursor.rowcount
            conn.close()
            
            return jsonify({'success': True, 'count': count})
            
        except Exception as e:
            app.logger.error(f"Mark all notifications read error: {str(e)}")
            return jsonify({'error': f'Fehler: {str(e)}'}), 500
    
    # ============================================================================
    # EMAIL SETTINGS ENDPOINTS
    # ============================================================================
    
    @app.route('/api/email-settings', methods=['GET'])
    @require_role('Admin')
    def get_email_settings():
        """Get email settings (Admin only)"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT SmtpHost, SmtpPort, UseSsl, RequiresAuthentication, 
                       Username, SenderEmail, SenderName, ReplyToEmail, IsEnabled
                FROM EmailSettings
                WHERE Id = 1
            """)
            
            row = cursor.fetchone()
            conn.close()
            
            if row:
                return jsonify({
                    'smtpHost': row[0],
                    'smtpPort': row[1],
                    'useSsl': bool(row[2]),
                    'requiresAuthentication': bool(row[3]),
                    'username': row[4],
                    # Don't send password for security
                    'senderEmail': row[5],
                    'senderName': row[6],
                    'replyToEmail': row[7],
                    'isEnabled': bool(row[8])
                })
            else:
                # Return default values if not configured
                return jsonify({
                    'smtpHost': '',
                    'smtpPort': 587,
                    'useSsl': True,
                    'requiresAuthentication': True,
                    'username': '',
                    'senderEmail': '',
                    'senderName': 'Dienstplan',
                    'replyToEmail': '',
                    'isEnabled': False
                })
                
        except Exception as e:
            app.logger.error(f"Get email settings error: {str(e)}")
            return jsonify({'error': f'Fehler: {str(e)}'}), 500
    
    @app.route('/api/email-settings', methods=['POST'])
    @require_role('Admin')
    def save_email_settings():
        """Save email settings (Admin only)"""
        try:
            data = request.get_json()
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Check if settings exist
            cursor.execute("SELECT Id FROM EmailSettings WHERE Id = 1")
            exists = cursor.fetchone()
            
            if exists:
                # Update existing settings
                # Only update password if provided
                if data.get('password'):
                    cursor.execute("""
                        UPDATE EmailSettings
                        SET SmtpHost = ?, SmtpPort = ?, UseSsl = ?, RequiresAuthentication = ?,
                            Username = ?, Password = ?, SenderEmail = ?, SenderName = ?, 
                            ReplyToEmail = ?, IsEnabled = ?, ModifiedAt = ?, ModifiedBy = ?
                        WHERE Id = 1
                    """, (
                        data.get('smtpHost'),
                        data.get('smtpPort', 587),
                        1 if data.get('useSsl') else 0,
                        1 if data.get('requiresAuthentication') else 0,
                        data.get('username'),
                        data.get('password'),
                        data.get('senderEmail'),
                        data.get('senderName'),
                        data.get('replyToEmail'),
                        1 if data.get('isEnabled') else 0,
                        datetime.utcnow().isoformat(),
                        session.get('user_email')
                    ))
                else:
                    cursor.execute("""
                        UPDATE EmailSettings
                        SET SmtpHost = ?, SmtpPort = ?, UseSsl = ?, RequiresAuthentication = ?,
                            Username = ?, SenderEmail = ?, SenderName = ?, 
                            ReplyToEmail = ?, IsEnabled = ?, ModifiedAt = ?, ModifiedBy = ?
                        WHERE Id = 1
                    """, (
                        data.get('smtpHost'),
                        data.get('smtpPort', 587),
                        1 if data.get('useSsl') else 0,
                        1 if data.get('requiresAuthentication') else 0,
                        data.get('username'),
                        data.get('senderEmail'),
                        data.get('senderName'),
                        data.get('replyToEmail'),
                        1 if data.get('isEnabled') else 0,
                        datetime.utcnow().isoformat(),
                        session.get('user_email')
                    ))
            else:
                # Insert new settings
                cursor.execute("""
                    INSERT INTO EmailSettings 
                    (Id, SmtpHost, SmtpPort, UseSsl, RequiresAuthentication, 
                     Username, Password, SenderEmail, SenderName, ReplyToEmail, 
                     IsEnabled, ModifiedBy)
                    VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    data.get('smtpHost'),
                    data.get('smtpPort', 587),
                    1 if data.get('useSsl') else 0,
                    1 if data.get('requiresAuthentication') else 0,
                    data.get('username'),
                    data.get('password'),
                    data.get('senderEmail'),
                    data.get('senderName'),
                    data.get('replyToEmail'),
                    1 if data.get('isEnabled') else 0,
                    session.get('user_email')
                ))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Save email settings error: {str(e)}")
            return jsonify({'error': f'Fehler: {str(e)}'}), 500
    
    @app.route('/api/email-settings/test', methods=['POST'])
    @require_role('Admin')
    def test_email_settings():
        """Send test email to verify settings (Admin only)"""
        try:
            data = request.get_json()
            test_email = data.get('testEmail')
            
            if not test_email:
                return jsonify({'error': 'Test-E-Mail-Adresse erforderlich'}), 400
            
            from email_service import send_test_email
            
            conn = db.get_connection()
            success, error = send_test_email(conn, test_email)
            conn.close()
            
            if success:
                return jsonify({'success': True})
            else:
                return jsonify({'error': error}), 500
                
        except Exception as e:
            app.logger.error(f"Test email error: {str(e)}")
            return jsonify({'error': f'Fehler: {str(e)}'}), 500
    
    # ============================================================================
    # PASSWORD MANAGEMENT ENDPOINTS
    # ============================================================================
    
    @app.route('/api/auth/change-password', methods=['POST'])
    @require_auth
    def change_password():
        """Change password for currently logged in user"""
        try:
            data = request.get_json()
            current_password = data.get('currentPassword')
            new_password = data.get('newPassword')
            
            if not current_password or not new_password:
                return jsonify({'error': 'Aktuelles und neues Passwort sind erforderlich'}), 400
            
            if len(new_password) < 8:
                return jsonify({'error': 'Neues Passwort muss mindestens 8 Zeichen lang sein'}), 400
            
            user_id = session.get('user_id')
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Get current password hash
            cursor.execute("SELECT PasswordHash FROM Employees WHERE Id = ?", (user_id,))
            row = cursor.fetchone()
            
            if not row or not row[0]:
                conn.close()
                return jsonify({'error': 'Benutzer hat kein Passwort gesetzt'}), 400
            
            # Verify current password
            if not verify_password(current_password, row[0]):
                conn.close()
                return jsonify({'error': 'Aktuelles Passwort ist falsch'}), 401
            
            # Update password
            new_password_hash = hash_password(new_password)
            security_stamp = secrets.token_hex(16)
            
            cursor.execute("""
                UPDATE Employees
                SET PasswordHash = ?, SecurityStamp = ?
                WHERE Id = ?
            """, (new_password_hash, security_stamp, user_id))
            
            # Log audit entry
            log_audit(conn, 'Employee', user_id, 'PasswordChanged', 
                     json.dumps({'action': 'User changed own password'}, ensure_ascii=False))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Change password error: {str(e)}")
            return jsonify({'error': f'Fehler: {str(e)}'}), 500
    
    @app.route('/api/auth/forgot-password', methods=['POST'])
    def forgot_password():
        """Request password reset link"""
        try:
            data = request.get_json()
            email = data.get('email')
            
            if not email:
                return jsonify({'error': 'E-Mail-Adresse ist erforderlich'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Find employee by email
            cursor.execute("""
                SELECT Id, Vorname, Name, Email
                FROM Employees
                WHERE Email = ? AND PasswordHash IS NOT NULL
            """, (email,))
            
            employee = cursor.fetchone()
            
            # Always return success to prevent email enumeration
            if not employee:
                conn.close()
                return jsonify({'success': True, 'message': 'Falls die E-Mail-Adresse existiert, wurde eine Anleitung zum Zurücksetzen des Passworts gesendet.'})
            
            # Generate reset token
            import secrets as sec
            reset_token = sec.token_urlsafe(32)
            expires_at = (datetime.utcnow() + timedelta(hours=24)).isoformat()
            
            # Store reset token
            cursor.execute("""
                INSERT INTO PasswordResetTokens (EmployeeId, Token, ExpiresAt)
                VALUES (?, ?, ?)
            """, (employee[0], reset_token, expires_at))
            
            conn.commit()
            
            # Send reset email
            from email_service import send_password_reset_email
            employee_name = f"{employee[1]} {employee[2]}"
            base_url = request.host_url.rstrip('/')
            
            success, error = send_password_reset_email(
                conn, employee[3], reset_token, employee_name, base_url
            )
            
            conn.close()
            
            if not success:
                app.logger.error(f"Failed to send password reset email: {error}")
                # Don't expose email errors to user
            
            return jsonify({'success': True, 'message': 'Falls die E-Mail-Adresse existiert, wurde eine Anleitung zum Zurücksetzen des Passworts gesendet.'})
            
        except Exception as e:
            app.logger.error(f"Forgot password error: {str(e)}")
            return jsonify({'error': f'Fehler: {str(e)}'}), 500
    
    @app.route('/api/auth/reset-password', methods=['POST'])
    def reset_password():
        """Reset password using token"""
        try:
            data = request.get_json()
            token = data.get('token')
            new_password = data.get('newPassword')
            
            if not token or not new_password:
                return jsonify({'error': 'Token und neues Passwort sind erforderlich'}), 400
            
            if len(new_password) < 8:
                return jsonify({'error': 'Passwort muss mindestens 8 Zeichen lang sein'}), 400
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # Find valid token
            cursor.execute("""
                SELECT Id, EmployeeId, ExpiresAt
                FROM PasswordResetTokens
                WHERE Token = ? AND IsUsed = 0
            """, (token,))
            
            token_row = cursor.fetchone()
            
            if not token_row:
                conn.close()
                return jsonify({'error': 'Ungültiger oder bereits verwendeter Token'}), 400
            
            token_id = token_row[0]
            employee_id = token_row[1]
            expires_at = datetime.fromisoformat(token_row[2])
            
            # Check if token is expired
            if expires_at < datetime.utcnow():
                conn.close()
                return jsonify({'error': 'Token ist abgelaufen'}), 400
            
            # Update password
            new_password_hash = hash_password(new_password)
            security_stamp = secrets.token_hex(16)
            
            cursor.execute("""
                UPDATE Employees
                SET PasswordHash = ?, SecurityStamp = ?, AccessFailedCount = 0, LockoutEnd = NULL
                WHERE Id = ?
            """, (new_password_hash, security_stamp, employee_id))
            
            # Mark token as used
            cursor.execute("""
                UPDATE PasswordResetTokens
                SET IsUsed = 1, UsedAt = ?
                WHERE Id = ?
            """, (datetime.utcnow().isoformat(), token_id))
            
            # Log audit entry
            log_audit(conn, 'Employee', employee_id, 'PasswordReset', 
                     json.dumps({'action': 'Password reset via email token'}, ensure_ascii=False))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True})
            
        except Exception as e:
            app.logger.error(f"Reset password error: {str(e)}")
            return jsonify({'error': f'Fehler: {str(e)}'}), 500
    
    @app.route('/api/auth/validate-reset-token', methods=['POST'])
    def validate_reset_token():
        """Validate if reset token is valid"""
        try:
            data = request.get_json()
            token = data.get('token')
            
            if not token:
                return jsonify({'valid': False})
            
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT ExpiresAt
                FROM PasswordResetTokens
                WHERE Token = ? AND IsUsed = 0
            """, (token,))
            
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                return jsonify({'valid': False})
            
            expires_at = datetime.fromisoformat(row[0])
            if expires_at < datetime.utcnow():
                return jsonify({'valid': False})
            
            return jsonify({'valid': True})
            
        except Exception as e:
            app.logger.error(f"Validate reset token error: {str(e)}")
            return jsonify({'valid': False})
    
    # ============================================================================
    # STATIC FILES (Web UI)
    # ============================================================================
    
    @app.route('/')
    def index():
        """Serve the main web UI"""
        return app.send_static_file('index.html')
    
    return app


if __name__ == "__main__":
    import os
    # Only enable debug in development (not in production)
    debug_mode = os.environ.get('FLASK_ENV') == 'development'
    app = create_app()
    app.run(debug=debug_mode, port=5000)
